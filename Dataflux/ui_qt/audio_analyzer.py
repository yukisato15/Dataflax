#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PySide6-based Audio Analysis and Processing Tool
Enhanced audio analyzer with detailed metadata analysis and comprehensive sorting options
Based on the general analyzer UI structure
"""

from PySide6.QtWidgets import *
from PySide6.QtCore import *
from PySide6.QtGui import *
from pathlib import Path
import sys
import json
import csv
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from collections import defaultdict
from threading import Event
import shutil
import contextlib
import wave
import copy
import uuid
import os
import time
import unicodedata
import subprocess
import hashlib
try:
    import aifc
except ImportError:
    aifc = None

from .folder_tools import (
    FolderNameDeleteDialog,
    MATCH_EXACT,
    remove_folders_matching_query,
)
from utils.ffprobe_finder import find_ffprobe

PROJECT_ROOT = Path(__file__).resolve().parent.parent
FFPROBE_PATH = find_ffprobe()
CACHE_DIR = PROJECT_ROOT / "cache"
FFPROBE_CACHE_FILE = CACHE_DIR / "ffprobe_cache.json"
ANALYSIS_CACHE_FILE = CACHE_DIR / "analysis_cache.json"


def _load_json(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    try:
        with open(path, "r", encoding="utf-8") as fp:
            data = json.load(fp)
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    return {}


def _save_json(path: Path, data: Dict[str, Any]):
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as fp:
            json.dump(data, fp, ensure_ascii=False, indent=2)
    except Exception:
        pass

UNKNOWN_DURATION_KEY = "len_unknown"

DEFAULT_DURATION_RANGES = [
    {"key": "len_0_5s", "label": "0秒〜5秒", "min_seconds": 0, "max_seconds": 5},
    {"key": "len_5_30s", "label": "5秒〜30秒", "min_seconds": 5, "max_seconds": 30},
    {"key": "len_30_60s", "label": "30秒〜1分", "min_seconds": 30, "max_seconds": 60},
    {"key": "len_1_2m", "label": "1分〜2分", "min_seconds": 60, "max_seconds": 120},
    {"key": "len_2_3m", "label": "2分〜3分", "min_seconds": 120, "max_seconds": 180},
    {"key": "len_3_4m", "label": "3分〜4分", "min_seconds": 180, "max_seconds": 240},
    {"key": "len_4_5m", "label": "4分〜5分", "min_seconds": 240, "max_seconds": 300},
    {"key": "len_5_10m", "label": "5分〜10分", "min_seconds": 300, "max_seconds": 600},
    {"key": "len_10_30m", "label": "10分〜30分", "min_seconds": 600, "max_seconds": 1800},
    {"key": "len_30_60m", "label": "30分〜1時間", "min_seconds": 1800, "max_seconds": 3600},
    {"key": "len_60m_plus", "label": "1時間以上", "min_seconds": 3600, "max_seconds": None},
]


CATEGORY_LABELS = {
    "format": "フォーマット",
    "samplerate": "サンプルレート",
    "channels": "チャンネル",
    "duration": "時間",
    "bitrate": "ビットレート",
    "date": "日付",
}

BASE_CATEGORY_NAME_MAP = {
    "format": {"fmt_wav": "WAV", "fmt_mp3": "MP3", "fmt_flac": "FLAC", "fmt_aac": "AAC/M4A", "fmt_aiff": "AIFF", "fmt_other": "その他"},
    "samplerate": {"sr_low": "低品質 (~22kHz)", "sr_cd": "CD品質 (44.1kHz)", "sr_dvd": "DVD品質 (48kHz)", "sr_hd": "ハイレゾ (96kHz)", "sr_ultra": "超高解像度 (96kHz+)", "sr_unknown": "不明"},
    "channels": {"ch_mono": "モノラル", "ch_stereo": "ステレオ", "ch_unknown": "不明"},
    "bitrate": {"br_low": "低ビットレート (<128kbps)", "br_medium": "中ビットレート (128-192kbps)", "br_high": "高ビットレート (192-320kbps)", "br_very_high": "超高ビットレート (320kbps+)", "br_lossless_or_unknown": "ロスレス/不明"},
    "date": {},
    "duration": {},
}

CATEGORY_ORDER = ["format", "samplerate", "channels", "duration", "bitrate", "date"]

CATEGORY_DISPLAY_HEADERS = [
    "カテゴリ種別",
    "カテゴリキー",
    "カテゴリ名",
    "ファイル数",
    "合計サイズ(MB)",
    "合計時間(秒)",
    "合計時間(表示)",
]

DETAIL_COLUMN_DEFS = (
    [{"id": "name", "label": "データ名"}]
    + [{"id": f"folder_level_{i}", "label": f"第{i}階層フォルダ"} for i in range(0, 11)]
    + [
        {"id": "format", "label": "フォーマット"},
        {"id": "samplerate", "label": "サンプルレート(Hz)"},
        {"id": "channels", "label": "チャンネル"},
        {"id": "duration_seconds", "label": "再生時間(秒)"},
        {"id": "duration", "label": "再生時間"},
        {"id": "bitrate", "label": "ビットレート(kbps)"},
        {"id": "codec", "label": "コーデック"},
        {"id": "size", "label": "ファイルサイズ(バイト)"},
    ]
)

DEFAULT_DETAIL_COLUMN_IDS = [
    "name",
    "format",
    "samplerate",
    "channels",
    "duration_seconds",
    "duration",
    "bitrate",
    "codec",
    "size",
]


def _bps_to_kbps(value) -> Optional[float]:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    if numeric <= 0:
        return None
    if numeric > 1000:
        numeric = numeric / 1000.0
    return round(numeric, 3)


def ffprobe_audio_metadata(path: Path) -> Dict[str, Any]:
    if not FFPROBE_PATH:
        return {}
    try:
        cmd = [
            FFPROBE_PATH,
            "-v",
            "quiet",
            "-print_format",
            "json",
            "-show_streams",
            "-select_streams",
            "a:0",
            str(path),
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
        if result.returncode != 0 or not result.stdout:
            return {}
        data = json.loads(result.stdout)
        streams = data.get("streams", [])
        stream = next((s for s in streams if s.get("codec_type") == "audio"), None)
        if not stream:
            return {}
        metadata: Dict[str, Any] = {}
        for key in (
            "codec_name",
            "codec_long_name",
            "sample_fmt",
            "bits_per_sample",
            "bits_per_raw_sample",
            "sample_rate",
            "channels",
            "duration",
            "bit_rate",
        ):
            if key in stream and stream[key] not in (None, ""):
                metadata[key] = stream[key]
        return metadata
    except Exception:
        return {}


def deep_copy_duration_ranges(ranges: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Utility to clone duration range definitions."""
    return copy.deepcopy(ranges)


def get_duration_settings_path() -> Path:
    """Return a writable path for storing duration range settings."""
    candidate = PROJECT_ROOT / "audio_duration_ranges.json"
    try:
        candidate.parent.mkdir(parents=True, exist_ok=True)
        if candidate.exists() or os.access(candidate.parent, os.W_OK):
            return candidate
    except Exception:
        pass

    app_config_dir = Path(QStandardPaths.writableLocation(QStandardPaths.AppConfigLocation) or str(Path.home() / ".dataflux"))
    app_config_dir.mkdir(parents=True, exist_ok=True)
    return app_config_dir / "audio_duration_ranges.json"


def load_duration_ranges(settings_path: Path) -> List[Dict[str, Any]]:
    """Load duration ranges from json if available, otherwise defaults."""
    if settings_path.exists():
        try:
            with open(settings_path, "r", encoding="utf-8") as fp:
                data = json.load(fp)
            ranges = []
            for entry in data:
                if not isinstance(entry, dict):
                    continue
                key = entry.get("key") or f"len_custom_{uuid.uuid4().hex[:8]}"
                label = entry.get("label") or key
                min_seconds = entry.get("min_seconds")
                max_seconds = entry.get("max_seconds")
                if min_seconds is not None:
                    try:
                        min_seconds = float(min_seconds)
                    except Exception:
                        min_seconds = None
                if max_seconds is not None:
                    try:
                        max_seconds = float(max_seconds)
                    except Exception:
                        max_seconds = None
                ranges.append({
                    "key": key,
                    "label": label,
                    "min_seconds": min_seconds,
                    "max_seconds": max_seconds,
                })
            if ranges:
                return ranges
        except Exception:
            pass
    return deep_copy_duration_ranges(DEFAULT_DURATION_RANGES)


def save_duration_ranges(settings_path: Path, ranges: List[Dict[str, Any]]):
    """Persist duration ranges to json."""
    with open(settings_path, "w", encoding="utf-8") as fp:
        json.dump(ranges, fp, ensure_ascii=False, indent=2)


def format_seconds_value(seconds: Optional[float]) -> str:
    """Convert seconds to a human-friendly short string (e.g. 5s, 3m, 1h)."""
    if seconds is None:
        return ""
    try:
        seconds = float(seconds)
    except Exception:
        return ""
    if seconds % 3600 == 0 and seconds >= 3600:
        return f"{int(seconds // 3600)}h"
    if seconds % 60 == 0 and seconds >= 60:
        return f"{int(seconds // 60)}m"
    if seconds.is_integer():
        return f"{int(seconds)}s"
    return f"{seconds:.2f}s"


def parse_duration_value(text: str) -> Optional[float]:
    """Parse user input like '30s', '1m', '2h', '90' (seconds) into seconds."""
    if text is None:
        return None
    value = text.strip().lower()
    if not value:
        return None

    unit_multipliers = {
        "s": 1,
        "sec": 1,
        "秒": 1,
        "m": 60,
        "min": 60,
        "分": 60,
        "h": 3600,
        "hr": 3600,
        "hour": 3600,
        "時間": 3600,
    }

    # Support colon formats like HH:MM:SS or MM:SS
    if ":" in value and value.replace(":", "").replace(".", "").isdigit():
        parts = [float(p) for p in value.split(":")]
        seconds_total = 0
        for part in parts:
            seconds_total = seconds_total * 60 + part
        return float(seconds_total)

    # Extract numeric part and unit suffix
    for unit, multiplier in unit_multipliers.items():
        if value.endswith(unit):
            number_part = value[:-len(unit)].strip()
            if not number_part:
                return None
            try:
                return float(number_part) * multiplier
            except Exception:
                return None

    try:
        return float(value)
    except Exception:
        return None


def ensure_unique_duration_key(label: str, existing_keys: List[str]) -> str:
    """Generate a stable key for a label while avoiding duplicates."""
    base = "len_" + "".join(ch for ch in label if ch.isalnum() or ch in ("_", "-"))
    if not base or base == "len_":
        base = f"len_custom_{uuid.uuid4().hex[:6]}"
    candidate = base
    counter = 1
    while candidate in existing_keys:
        candidate = f"{base}_{counter}"
        counter += 1
    return candidate


def path_is_within(child: Path, parent: Path) -> bool:
    """Return True if child path is under parent path (inclusive)."""
    try:
        child.resolve().relative_to(parent.resolve())
        return True
    except Exception:
        return False


def is_hidden_name(name: str) -> bool:
    """Return True if a filename or directory name should be treated as hidden."""
    return name.startswith(".") or name.startswith("._")

# Audio analysis imports
try:
    from mutagen.mp3 import MP3
    from mutagen.flac import FLAC
    from mutagen.mp4 import MP4
    from mutagen.id3 import ID3NoHeaderError
except ImportError:
    MP3 = FLAC = MP4 = None
    ID3NoHeaderError = Exception

# Import the scanner from core module
sys.path.append(str(Path(__file__).parent.parent))

# Audio processing utilities
def unique_name(dest_dir: Path, filename: str) -> Path:
    """Generate unique filename to avoid overwriting"""
    dest_dir.mkdir(parents=True, exist_ok=True)
    base = Path(filename).stem
    ext = Path(filename).suffix
    candidate = dest_dir / f"{base}{ext}"
    counter = 1
    while candidate.exists():
        candidate = dest_dir / f"{base}_{counter:02d}{ext}"
        counter += 1
    return candidate

def send_to_trash(path: Path):
    """Move file to macOS trash"""
    trash = Path.home() / ".Trash"
    target = unique_name(trash, path.name)
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(path), str(target))

def get_file_hash(path: Path) -> str:
    """Calculate MD5 hash of file for duplicate detection."""
    try:
        h = hashlib.md5()
        with open(path, "rb") as fp:
            for chunk in iter(lambda: fp.read(4096), b""):
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return ""

def audio_probe(path: Path, include_ffprobe: bool = False) -> Dict[str, Any]:
    """Extract comprehensive audio metadata"""
    info = {
        "path": str(path),
        "name": path.name,
        "ext": path.suffix.lower(),
        "size": 0,
        "mtime": None,
        "samplerate": None,
        "channels": None,
        "duration": None,
        "bitrate": None,
        "format": None,
        "codec_name": None,
        "sample_fmt": None,
        "bits_per_sample": None,
        "title": None,
        "artist": None,
        "album": None,
        "genre": None,
        "year": None,
        "analysis_error": None
    }
    
    try:
        stat = path.stat()
        info["size"] = stat.st_size
        info["mtime"] = stat.st_mtime
    except:
        pass
    
    try:
        ext = path.suffix.lower()
        
        if ext == ".wav":
            with contextlib.closing(wave.open(str(path), "rb")) as w:
                info["samplerate"] = w.getframerate()
                info["channels"] = w.getnchannels()
                frames = w.getnframes()
                framerate = w.getframerate()
                sample_width = w.getsampwidth()
                if framerate:
                    info["duration"] = frames / framerate
                if sample_width:
                    info["bits_per_sample"] = sample_width * 8
                if info["samplerate"] and info["channels"] and sample_width:
                    bps = info["samplerate"] * info["channels"] * sample_width * 8
                    bitrate_kbps = _bps_to_kbps(bps)
                    if bitrate_kbps:
                        info["bitrate"] = bitrate_kbps
                info["format"] = "PCM WAV"
                info["codec_name"] = "pcm"

        elif ext in (".aif", ".aiff") and aifc:
            with contextlib.closing(aifc.open(str(path), "rb")) as w:
                info["samplerate"] = w.getframerate()
                info["channels"] = w.getnchannels()
                frames = w.getnframes()
                framerate = w.getframerate()
                sample_width = w.getsampwidth()
                if framerate:
                    info["duration"] = frames / framerate
                if sample_width:
                    info["bits_per_sample"] = sample_width * 8
                if info["samplerate"] and info["channels"] and sample_width:
                    bps = info["samplerate"] * info["channels"] * sample_width * 8
                    bitrate_kbps = _bps_to_kbps(bps)
                    if bitrate_kbps:
                        info["bitrate"] = bitrate_kbps
                info["format"] = "AIFF"
                info["codec_name"] = "pcm"

        elif ext == ".mp3" and MP3:
            audio = MP3(str(path))
            if audio.info:
                info["samplerate"] = getattr(audio.info, "sample_rate", None)
                info["channels"] = getattr(audio.info, "channels", None)
                info["duration"] = getattr(audio.info, "length", None)
                bitrate = getattr(audio.info, "bitrate", None)
                bitrate_kbps = _bps_to_kbps(bitrate)
                if bitrate_kbps:
                    info["bitrate"] = bitrate_kbps
                info["format"] = f"MP3 {info['bitrate']}kbps" if info["bitrate"] else "MP3"
                info["codec_name"] = "mp3"

            # ID3 tags
            try:
                info["title"] = str(audio.get("TIT2", [""])[0])
                info["artist"] = str(audio.get("TPE1", [""])[0])
                info["album"] = str(audio.get("TALB", [""])[0])
                info["genre"] = str(audio.get("TCON", [""])[0])
                year_tag = audio.get("TDRC") or audio.get("TYER")
                if year_tag:
                    info["year"] = str(year_tag[0])
            except:
                pass
        
        elif ext == ".flac" and FLAC:
            audio = FLAC(str(path))
            if audio.info:
                info["samplerate"] = getattr(audio.info, "sample_rate", None)
                info["channels"] = getattr(audio.info, "channels", None)
                info["duration"] = getattr(audio.info, "length", None)
                bitrate = getattr(audio.info, "bitrate", None)
                bitrate_kbps = _bps_to_kbps(bitrate)
                if bitrate_kbps:
                    info["bitrate"] = bitrate_kbps
                info["format"] = f"FLAC {info['samplerate']}Hz" if info["samplerate"] else "FLAC"
                info["codec_name"] = "flac"
            
            # FLAC tags
            try:
                info["title"] = audio.get("TITLE", [""])[0]
                info["artist"] = audio.get("ARTIST", [""])[0]
                info["album"] = audio.get("ALBUM", [""])[0]
                info["genre"] = audio.get("GENRE", [""])[0]
                info["year"] = audio.get("DATE", [""])[0]
            except:
                pass
        
        elif ext in (".m4a", ".mp4") and MP4:
            audio = MP4(str(path))
            if audio.info:
                info["samplerate"] = getattr(audio.info, "sample_rate", None)
                info["channels"] = getattr(audio.info, "channels", None)
                info["duration"] = getattr(audio.info, "length", None)
                bitrate = getattr(audio.info, "bitrate", None)
                bitrate_kbps = _bps_to_kbps(bitrate)
                if bitrate_kbps:
                    info["bitrate"] = bitrate_kbps
                info["format"] = f"AAC {info['bitrate']}kbps" if info["bitrate"] else "AAC"
                info["codec_name"] = "aac"
            
            # MP4 tags
            try:
                info["title"] = audio.get("\\xa9nam", [""])[0]
                info["artist"] = audio.get("\\xa9ART", [""])[0]
                info["album"] = audio.get("\\xa9alb", [""])[0]
                info["genre"] = audio.get("\\xa9gen", [""])[0]
                info["year"] = str(audio.get("\\xa9day", [""])[0])
            except:
                pass
                
    except Exception as e:
        info["analysis_error"] = str(e)

    if include_ffprobe and FFPROBE_PATH and path.exists():
        ff_meta = ffprobe_audio_metadata(path)
        if ff_meta:
            codec_name = ff_meta.get("codec_name") or ff_meta.get("codec_long_name")
            if codec_name:
                info["codec_name"] = codec_name
            sample_fmt = ff_meta.get("sample_fmt")
            if sample_fmt:
                info["sample_fmt"] = sample_fmt
            bits_per_sample = ff_meta.get("bits_per_sample") or ff_meta.get("bits_per_raw_sample")
            if bits_per_sample is not None and info.get("bits_per_sample") is None:
                try:
                    info["bits_per_sample"] = int(bits_per_sample)
                except (TypeError, ValueError):
                    pass
            if not info.get("samplerate") and ff_meta.get("sample_rate"):
                try:
                    info["samplerate"] = int(float(ff_meta["sample_rate"]))
                except (TypeError, ValueError):
                    pass
            if not info.get("channels") and ff_meta.get("channels") is not None:
                try:
                    info["channels"] = int(ff_meta["channels"])
                except (TypeError, ValueError):
                    pass
            if not info.get("duration") and ff_meta.get("duration"):
                try:
                    info["duration"] = float(ff_meta["duration"])
                except (TypeError, ValueError):
                    pass
            bit_rate = ff_meta.get("bit_rate")
            bitrate_kbps = _bps_to_kbps(bit_rate)
            if bitrate_kbps and info.get("bitrate") is None:
                info["bitrate"] = bitrate_kbps
            if not info.get("format") and codec_name:
                info["format"] = codec_name.upper()

    if info.get("bitrate") is not None:
        try:
            info["bitrate"] = float(info["bitrate"])
        except (TypeError, ValueError):
            info["bitrate"] = None

    return info

def categorize_audio(info: Dict[str, Any], duration_ranges: List[Dict[str, Any]]) -> Dict[str, str]:
    """Categorize audio file by various criteria"""
    categories = {}
    
    # Sample rate category
    sr = info.get("samplerate")
    if sr:
        if sr <= 22050:
            categories["samplerate"] = "sr_low"
        elif sr <= 44100:
            categories["samplerate"] = "sr_cd"
        elif sr <= 48000:
            categories["samplerate"] = "sr_dvd"
        elif sr <= 96000:
            categories["samplerate"] = "sr_hd"
        else:
            categories["samplerate"] = "sr_ultra"
    else:
        categories["samplerate"] = "sr_unknown"
    
    # Channel category
    ch = info.get("channels")
    if ch == 1:
        categories["channels"] = "ch_mono"
    elif ch == 2:
        categories["channels"] = "ch_stereo"
    elif ch and ch > 2:
        categories["channels"] = f"ch_multi_{ch}"
    else:
        categories["channels"] = "ch_unknown"
    
    # Duration category
    duration = info.get("duration")
    if duration is not None:
        matched_key = None
        for duration_range in duration_ranges:
            min_seconds = duration_range.get("min_seconds")
            max_seconds = duration_range.get("max_seconds")
            lower_ok = True if min_seconds is None else duration >= min_seconds
            upper_ok = True if max_seconds is None else duration < max_seconds
            if lower_ok and upper_ok:
                matched_key = duration_range["key"]
                break
        categories["duration"] = matched_key or UNKNOWN_DURATION_KEY
    else:
        categories["duration"] = UNKNOWN_DURATION_KEY
    
    # Bitrate category (for compressed formats)
    bitrate = info.get("bitrate")
    try:
        bitrate_value = float(bitrate) if bitrate is not None else None
    except (TypeError, ValueError):
        bitrate_value = None
    if bitrate_value:
        if bitrate_value < 128:
            categories["bitrate"] = "br_low"
        elif bitrate_value < 192:
            categories["bitrate"] = "br_medium"
        elif bitrate_value < 320:
            categories["bitrate"] = "br_high"
        else:
            categories["bitrate"] = "br_very_high"
    else:
        categories["bitrate"] = "br_lossless_or_unknown"
    
    # Format category
    ext = info.get("ext", "").lower()
    if ext == ".wav":
        categories["format"] = "fmt_wav"
    elif ext == ".mp3":
        categories["format"] = "fmt_mp3"
    elif ext == ".flac":
        categories["format"] = "fmt_flac"
    elif ext in (".m4a", ".mp4"):
        categories["format"] = "fmt_aac"
    elif ext in (".aif", ".aiff"):
        categories["format"] = "fmt_aiff"
    else:
        categories["format"] = "fmt_other"
    
    # Date category
    mtime = info.get("mtime")
    if mtime:
        date = datetime.fromtimestamp(mtime)
        categories["date"] = f"{date.year}-{date.month:02d}"
    else:
        categories["date"] = "date_unknown"
    
    return categories


def aggregate_audio_data(audio_infos: List[Dict[str, Any]], duration_ranges: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """Aggregate audio info into categories using flexible duration ranges."""
    aggregated: Dict[str, Dict[str, Any]] = {}

    for audio_info in audio_infos:
        categories = categorize_audio(audio_info, duration_ranges)
        for category_type, category_key in categories.items():
            category_map = aggregated.setdefault(category_type, {})
            bucket = category_map.setdefault(category_key, {
                "count": 0,
                "total_size": 0,
                "total_duration": 0,
                "files": []
            })

            bucket["count"] += 1
            bucket["total_size"] += audio_info.get("size", 0)
            duration_value = audio_info.get("duration")
            if duration_value:
                bucket["total_duration"] += duration_value
            bucket["files"].append(audio_info)

    return aggregated


class DurationSettingsDialog(QDialog):
    """時間区分を柔軟に編集するための設定ダイアログ"""

    def __init__(self, duration_ranges: List[Dict[str, Any]], parent=None):
        super().__init__(parent)
        self.setWindowTitle("時間区分設定")
        self.setMinimumSize(560, 420)

        self._ranges = deep_copy_duration_ranges(duration_ranges)

        main_layout = QVBoxLayout(self)
        info_label = QLabel("開始/終了には 5s, 1m, 90, 1h, 1:30 などの形式で入力できます。終了を空欄にすると上限なしになります。")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: #cccccc")
        main_layout.addWidget(info_label)

        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(["表示名", "開始", "終了"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)

        main_layout.addWidget(self.table)

        button_row = QHBoxLayout()
        add_button = QPushButton("追加")
        add_button.clicked.connect(self.add_row)
        button_row.addWidget(add_button)

        remove_button = QPushButton("削除")
        remove_button.clicked.connect(self.remove_selected_row)
        button_row.addWidget(remove_button)

        up_button = QPushButton("上へ")
        up_button.clicked.connect(lambda: self.move_selected_row(-1))
        button_row.addWidget(up_button)

        down_button = QPushButton("下へ")
        down_button.clicked.connect(lambda: self.move_selected_row(1))
        button_row.addWidget(down_button)

        reset_button = QPushButton("デフォルトに戻す")
        reset_button.clicked.connect(self.reset_to_default)
        button_row.addWidget(reset_button)

        button_row.addStretch()
        main_layout.addLayout(button_row)

        action_buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        action_buttons.accepted.connect(self.accept)
        action_buttons.rejected.connect(self.reject)
        main_layout.addWidget(action_buttons)

        self.populate_table()

    def populate_table(self):
        self.table.setRowCount(0)
        for range_info in self._ranges:
            self.add_row(range_info)

    def add_row(self, range_info: Optional[Dict[str, Any]] = None):
        row = self.table.rowCount()
        self.table.insertRow(row)

        if range_info is None:
            existing_keys = [self.table.item(r, 0).data(Qt.UserRole) for r in range(self.table.rowCount()) if self.table.item(r, 0)]
            label = "新しい区分"
            key = ensure_unique_duration_key(label, [k for k in existing_keys if k])
            range_info = {
                "key": key,
                "label": label,
                "min_seconds": None,
                "max_seconds": None,
            }

        label_item = QTableWidgetItem(range_info.get("label", ""))
        label_item.setData(Qt.UserRole, range_info.get("key"))
        self.table.setItem(row, 0, label_item)

        min_item = QTableWidgetItem(format_seconds_value(range_info.get("min_seconds")))
        self.table.setItem(row, 1, min_item)

        max_item = QTableWidgetItem(format_seconds_value(range_info.get("max_seconds")))
        self.table.setItem(row, 2, max_item)

    def remove_selected_row(self):
        row = self.table.currentRow()
        if row >= 0:
            self.table.removeRow(row)

    def move_selected_row(self, offset: int):
        row = self.table.currentRow()
        if row < 0:
            return
        target_row = row + offset
        if target_row < 0 or target_row >= self.table.rowCount():
            return
        for column in range(self.table.columnCount()):
            current_item = self.table.takeItem(row, column)
            target_item = self.table.takeItem(target_row, column)
            self.table.setItem(row, column, target_item)
            self.table.setItem(target_row, column, current_item)
        self.table.setCurrentCell(target_row, 0)

    def reset_to_default(self):
        self._ranges = deep_copy_duration_ranges(DEFAULT_DURATION_RANGES)
        self.populate_table()

    def get_ranges(self) -> List[Dict[str, Any]]:
        return deep_copy_duration_ranges(self._ranges)

    def accept(self):
        row_count = self.table.rowCount()
        if row_count == 0:
            QMessageBox.warning(self, "警告", "時間区分が1件もありません。最低1件は設定してください。")
            return

        new_ranges: List[Dict[str, Any]] = []
        used_keys: List[str] = []

        for row in range(row_count):
            label_item = self.table.item(row, 0)
            min_item = self.table.item(row, 1)
            max_item = self.table.item(row, 2)

            label = label_item.text().strip() if label_item else ""
            if not label:
                QMessageBox.warning(self, "警告", f"{row + 1}行目の表示名を入力してください。")
                return

            min_seconds = parse_duration_value(min_item.text() if min_item else "")
            max_seconds = parse_duration_value(max_item.text() if max_item else "")

            if min_seconds is not None and max_seconds is not None and max_seconds <= min_seconds:
                QMessageBox.warning(self, "警告", f"{row + 1}行目の開始値は終了値より小さくしてください。")
                return

            key = label_item.data(Qt.UserRole)
            if not key:
                key = ensure_unique_duration_key(label, used_keys)
            if key in used_keys:
                key = ensure_unique_duration_key(label, used_keys)
            used_keys.append(key)

            new_ranges.append({
                "key": key,
                "label": label,
                "min_seconds": min_seconds,
                "max_seconds": max_seconds,
            })

        self._ranges = new_ranges
        super().accept()


class DetailExportOptionsDialog(QDialog):
    def __init__(self, current_selection: List[str], parent=None):
        super().__init__(parent)
        self.setWindowTitle("詳細エクスポート項目の選択")
        self.setMinimumWidth(360)
        self._checkboxes: Dict[str, QCheckBox] = {}

        layout = QVBoxLayout(self)

        info_label = QLabel("出力したい項目を選択してください。")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)

        checkbox_container = QWidget()
        checkbox_layout = QVBoxLayout(checkbox_container)
        checkbox_layout.setContentsMargins(0, 0, 0, 0)
        checkbox_layout.setSpacing(4)

        for column in DETAIL_COLUMN_DEFS:
            cb = QCheckBox(column["label"])
            cb.setChecked(column["id"] in current_selection)
            checkbox_layout.addWidget(cb)
            self._checkboxes[column["id"]] = cb

        checkbox_layout.addStretch()
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(checkbox_container)
        layout.addWidget(scroll, 1)

        button_row = QHBoxLayout()
        select_all_btn = QPushButton("全選択")
        select_all_btn.clicked.connect(self._select_all)
        button_row.addWidget(select_all_btn)

        select_default_btn = QPushButton("標準")
        select_default_btn.clicked.connect(self._select_default)
        button_row.addWidget(select_default_btn)

        button_row.addStretch()
        layout.addLayout(button_row)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self._on_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self._select_default_values = list(DEFAULT_DETAIL_COLUMN_IDS)

    def _select_all(self):
        for cb in self._checkboxes.values():
            cb.setChecked(True)

    def _select_default(self):
        for column_id, cb in self._checkboxes.items():
            cb.setChecked(column_id in self._select_default_values)

    def _on_accept(self):
        selected = self.selected_column_ids()
        if not selected:
            QMessageBox.warning(self, "警告", "少なくとも1項目は選択してください。")
            return
        self.accept()

    def selected_column_ids(self) -> List[str]:
        return [column_id for column_id, cb in self._checkboxes.items() if cb.isChecked()]


class CacheClearDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("キャッシュ削除")
        layout = QVBoxLayout(self)

        info_label = QLabel("削除したいキャッシュを選択してください。")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)

        self.analysis_cb = QCheckBox("解析キャッシュ")
        self.ffprobe_cb = QCheckBox("詳細(ffprobe)キャッシュ")
        layout.addWidget(self.analysis_cb)
        layout.addWidget(self.ffprobe_cb)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self._on_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _on_accept(self):
        if not (self.analysis_cb.isChecked() or self.ffprobe_cb.isChecked()):
            QMessageBox.warning(self, "警告", "少なくとも1つ選択してください。")
            return
        self.accept()

    def selection(self) -> Dict[str, bool]:
        return {
            "analysis": self.analysis_cb.isChecked(),
            "ffprobe": self.ffprobe_cb.isChecked(),
        }



class AudioAnalysisThread(QThread):
    """Audio analysis thread with progress, pause/resume, cancellation, and logging"""

    analysis_started = Signal(int)
    progress_updated = Signal(int, int, str)
    analysis_paused = Signal()
    analysis_resumed = Signal()
    analysis_completed = Signal(dict, list, float)
    analysis_cancelled = Signal(dict, list)
    error_occurred = Signal(str)
    log_ready = Signal(str)

    def __init__(self, paths: List[Path], duration_ranges: List[Dict[str, Any]], analysis_cache: Optional[Dict[str, Any]] = None):
        super().__init__()
        self.paths = paths if isinstance(paths, list) else [paths]
        self.audio_extensions = {'.wav', '.mp3', '.flac', '.m4a', '.mp4', '.aif', '.aiff'}
        self.duration_ranges = deep_copy_duration_ranges(duration_ranges)
        self.cancel_event = Event()
        self.pause_event = Event()
        self._log_entries: List[str] = []
        self._processed_files = 0
        self._total_files = 0
        self.analysis_cache_snapshot = analysis_cache or {}
        self.updated_cache: Dict[str, Any] = {}
        self.problem_files_found: List[Dict[str, Any]] = []

    def request_cancel(self):
        self.cancel_event.set()

    def request_pause(self):
        self.pause_event.set()

    def resume(self):
        if self.pause_event.is_set():
            self.pause_event.clear()

    def _wait_if_paused(self):
        was_paused = False
        if self.pause_event.is_set() and not self.cancel_event.is_set():
            self.analysis_paused.emit()
            was_paused = True
        while self.pause_event.is_set() and not self.cancel_event.is_set():
            time.sleep(0.1)
        if was_paused and not self.cancel_event.is_set():
            self.analysis_resumed.emit()

    def _append_log(self, message: str):
        timestamp = datetime.now().strftime('%H:%M:%S')
        self._log_entries.append(f"[{timestamp}] {message}")

    def _finalize_log(self, status: str, elapsed: float) -> Optional[Path]:
        try:
            log_dir = PROJECT_ROOT / "logs"
            log_dir.mkdir(exist_ok=True)
            summary = (
                f"status={status} total={self._total_files} processed={self._processed_files} "
                f"elapsed={elapsed:.2f}s"
            )
            self._log_entries.append(summary)
            log_path = log_dir / f"audio_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
            log_path.write_text("\n".join(self._log_entries), encoding="utf-8")
            return log_path
        except Exception:
            return None

    def _get_audio_info(self, file_path: Path) -> Tuple[Dict[str, Any], bool]:
        path_str = str(file_path)
        size = None
        mtime = None
        try:
            stat = file_path.stat()
            size = stat.st_size
            mtime = stat.st_mtime
        except Exception:
            pass

        cached = self.analysis_cache_snapshot.get(path_str)
        if cached and isinstance(cached, dict):
            if cached.get("size") == size and abs((cached.get("mtime") or 0) - (mtime or 0)) < 1:
                data = cached.get("data")
                if isinstance(data, dict):
                    return dict(data), True

        info = audio_probe(file_path, include_ffprobe=False)
        if info:
            self.updated_cache[path_str] = {
                "size": size,
                "mtime": mtime,
                "data": dict(info),
                "cached_at": datetime.now().isoformat(timespec="seconds"),
            }
        return info, False

    def _iter_audio_files(self, root: Path):
        stack = [root]
        while stack:
            if self.cancel_event.is_set():
                return
            current = stack.pop()
            if not current.exists():
                continue
            self._wait_if_paused()
            try:
                with os.scandir(current) as entries:
                    for entry in entries:
                        if self.cancel_event.is_set():
                            return
                        self._wait_if_paused()
                        if is_hidden_name(entry.name):
                            continue
                        entry_path = Path(entry.path)
                        try:
                            if entry.is_dir(follow_symlinks=False):
                                stack.append(entry_path)
                            elif entry.is_file(follow_symlinks=False) and entry_path.suffix.lower() in self.audio_extensions:
                                yield entry_path
                        except (OSError, PermissionError):
                            continue
            except (OSError, PermissionError):
                continue

    def run(self):
        start_time = time.monotonic()
        self._log_entries = []
        self._append_log(f"analysis start targets={len(self.paths)}")

        try:
            audio_files: List[Path] = []
            for root_path in self.paths:
                if self.cancel_event.is_set():
                    break
                self._wait_if_paused()
                if root_path.is_dir():
                    for file_path in self._iter_audio_files(root_path):
                        if self.cancel_event.is_set():
                            break
                        self._wait_if_paused()
                        audio_files.append(file_path)
                elif root_path.is_file() and root_path.suffix.lower() in self.audio_extensions:
                    audio_files.append(root_path)

            self._total_files = len(audio_files)
            self._append_log(f"discovered {self._total_files} audio files")

            if self.cancel_event.is_set():
                elapsed = time.monotonic() - start_time
                log_path = self._finalize_log("cancelled", elapsed)
                if log_path:
                    self.log_ready.emit(str(log_path))
                self.analysis_cancelled.emit({}, [])
                return

            if self._total_files == 0:
                elapsed = time.monotonic() - start_time
                log_path = self._finalize_log("empty", elapsed)
                if log_path:
                    self.log_ready.emit(str(log_path))
                self.analysis_completed.emit({}, [], elapsed)
                return

            self.analysis_started.emit(self._total_files)

            processed_infos: List[Dict[str, Any]] = []

            for index, file_path in enumerate(audio_files, start=1):
                if self.cancel_event.is_set():
                    break

                self._wait_if_paused()
                self.progress_updated.emit(index, self._total_files, str(file_path))

                try:
                    audio_info, used_cache = self._get_audio_info(file_path)
                    processed_infos.append(audio_info)
                    self._append_log(f"{'CACHE' if used_cache else 'ok'} {file_path}")
                except Exception as exc:
                    self._append_log(f"fail {file_path}: {exc}")
                    continue

                self._processed_files = index

            elapsed = time.monotonic() - start_time

            aggregated_results = aggregate_audio_data(processed_infos, self.duration_ranges)

            if self.cancel_event.is_set():
                log_path = self._finalize_log("cancelled", elapsed)
                if log_path:
                    self.log_ready.emit(str(log_path))
                self.analysis_cancelled.emit(aggregated_results, processed_infos)
                return

            self._processed_files = len(processed_infos)
            log_path = self._finalize_log("completed", elapsed)
            if log_path:
                self.log_ready.emit(str(log_path))

            self.analysis_completed.emit(aggregated_results, processed_infos, elapsed)

        except Exception as exc:
            elapsed = time.monotonic() - start_time
            self._append_log(f"error {exc}")
            log_path = self._finalize_log("error", elapsed)
            if log_path:
                self.log_ready.emit(str(log_path))
            self.error_occurred.emit(str(exc))


class AudioDetailExportWorker(QThread):
    """Gather ffprobe metadata on demand and export detailed CSV/Excel."""

    progress_updated = Signal(int, int, str)
    completed = Signal(str, dict, str)
    error_occurred = Signal(str)

    def __init__(
        self,
        files: List[Dict[str, Any]],
        cache_snapshot: Dict[str, Any],
        headers: List[str],
        output_path: str,
        export_format: str,
        column_ids: List[str],
        root_paths: List[Path],
    ):
        super().__init__()
        self.files = list(enumerate(files))
        self.cache_snapshot = cache_snapshot or {}
        self.headers = headers
        self.output_path = Path(output_path)
        self.export_format = export_format  # 'csv' or 'excel'
        self.column_ids = column_ids
        self.root_paths = [Path(p).resolve() for p in root_paths if p is not None]
        self.cancel_event = Event()
        self.updated_cache: Dict[str, Any] = {}
        self.log_entries: List[str] = []

    def request_cancel(self):
        self.cancel_event.set()

    def run(self):
        try:
            total = len(self.files)
            if total == 0:
                raise ValueError("詳細エクスポート対象がありません")

            rows: List[Tuple[int, Dict[str, Any]]] = []

            for index, info in self.files:
                if self.cancel_event.is_set():
                    self.log_entries.append("CANCELLED")
                    self.error_occurred.emit("ユーザーにより中止されました")
                    return

                row, cache_update, log_line = self._process_entry(info)
                rows.append((index, row))
                if cache_update:
                    key, entry = cache_update
                    self.updated_cache[key] = entry
                if log_line:
                    self.log_entries.append(log_line)

                self.progress_updated.emit(len(rows), total, info.get("name", ""))

            rows.sort(key=lambda x: x[0])
            ordered_rows = [row for _, row in rows]

            if self.export_format == "csv":
                self._write_csv(ordered_rows)
            else:
                self._write_excel(ordered_rows)

            log_path = self._write_log()
            self.completed.emit(str(self.output_path), self.updated_cache, log_path)

        except Exception as exc:
            self.error_occurred.emit(str(exc))

    def _process_entry(self, info: Dict[str, Any]):
        path_str = info.get("path", "")
        path = Path(path_str)
        cache_key = str(path)
        size = info.get("size")
        mtime = info.get("mtime")

        cached = self.cache_snapshot.get(cache_key)
        metadata = {}
        used_cache = False

        if cached and isinstance(cached, dict):
            cached_size = cached.get("size")
            cached_mtime = cached.get("mtime")
            if cached_size == size and abs((cached_mtime or 0) - (mtime or 0)) < 1:
                metadata = cached.get("data", {})
                used_cache = True

        if not metadata and FFPROBE_PATH and path.exists():
            metadata = ffprobe_audio_metadata(path)

        cache_update = None
        if metadata and not used_cache:
            cache_update = (
                cache_key,
                {
                    "size": size,
                    "mtime": mtime,
                    "data": metadata,
                    "cached_at": datetime.now().isoformat(timespec="seconds"),
                },
            )

        row = self._build_row(info, metadata)
        log_line = f"{'CACHE' if used_cache else 'ffprobe'} {path_str}" if FFPROBE_PATH else f"metadata {path_str}"
        return row, cache_update, log_line

    def _build_row(self, info: Dict[str, Any], metadata: Dict[str, Any]) -> Dict[str, Any]:
        duration_value = info.get("duration")
        try:
            duration_seconds = round(float(duration_value), 3) if duration_value is not None else ""
        except (TypeError, ValueError):
            duration_seconds = ""

        sample_rate = info.get("samplerate")
        try:
            sample_rate = int(float(sample_rate)) if sample_rate is not None else ""
        except (TypeError, ValueError):
            sample_rate = ""

        bitrate_value = info.get("bitrate")
        try:
            bitrate_value = float(bitrate_value) if bitrate_value is not None else None
        except (TypeError, ValueError):
            bitrate_value = None

        channel_label = AudioAnalyzerWindow._format_channel_label_static(info.get("channels"))

        rel_parts = self._relative_parts(Path(info.get("path", "")))

        codec_name = info.get("codec_name")
        if metadata:
            codec_name = metadata.get("codec_name") or metadata.get("codec_long_name") or codec_name
            if metadata.get("sample_rate") and not sample_rate:
                try:
                    sample_rate = int(float(metadata.get("sample_rate")))
                except (TypeError, ValueError):
                    pass
            if metadata.get("bit_rate") and bitrate_value is None:
                bitrate_value = _bps_to_kbps(metadata.get("bit_rate"))
            elif bitrate_value is not None:
                bitrate_value = round(bitrate_value, 3)
        elif bitrate_value is not None:
            bitrate_value = round(bitrate_value, 3)

        values_by_id: Dict[str, Any] = {
            "name": info.get("name", ""),
            "format": info.get("format") or info.get("ext", "").upper(),
            "samplerate": sample_rate or "",
            "channels": channel_label,
            "duration_seconds": duration_seconds,
            "duration": AudioAnalyzerWindow._format_total_duration_static(duration_value) if duration_value else "",
            "bitrate": bitrate_value if bitrate_value is not None else "",
            "codec": codec_name or (info.get("format") or info.get("ext", "").upper()),
            "size": int(info.get("size", 0) or 0),
        }

        for level in range(0, 11):
            key = f"folder_level_{level}"
            values_by_id[key] = rel_parts[level] if level < len(rel_parts) else ""

        row = {}
        for column_id, header in zip(self.column_ids, self.headers):
            row[header] = values_by_id.get(column_id, "")

        return row

    def _write_csv(self, rows: List[Dict[str, Any]]):
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(self.output_path, "w", newline="", encoding="utf-8") as fp:
            fp.write("\ufeff")  # BOM for Excel compatibility
            writer = csv.DictWriter(fp, fieldnames=self.headers)
            writer.writeheader()
            for row in rows:
                writer.writerow({key: row.get(key, "") for key in self.headers})

    def _write_excel(self, rows: List[Dict[str, Any]]):
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise RuntimeError("openpyxl がインストールされていません。") from exc

        wb = Workbook()
        ws = wb.active
        ws.title = "Audio Detail"
        ws.append(self.headers)
        for row in rows:
            ws.append([row.get(h, "") for h in self.headers])
        ws.freeze_panes = "A2"
        for idx, column_cells in enumerate(ws.columns, start=1):
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(idx)].width = min(max(length + 2, 12), 60)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.output_path)

    def _write_log(self) -> str:
        if not self.log_entries:
            return ""
        log_dir = PROJECT_ROOT / "logs"
        log_dir.mkdir(parents=True, exist_ok=True)
        log_path = log_dir / f"audio_detail_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        with open(log_path, "w", encoding="utf-8") as fp:
            fp.write("\n".join(self.log_entries))
        return str(log_path)

    def _relative_parts(self, path: Path) -> List[str]:
        try:
            resolved = path.resolve()
        except Exception:
            return []

        best_root = None
        best_relative = None
        for root in self.root_paths:
            try:
                rel = resolved.relative_to(root)
                if best_root is None or len(root.parts) > len(best_root.parts):
                    best_root = root
                    best_relative = rel
            except ValueError:
                continue

        parts: List[str] = []
        if best_root is not None:
            root_name = best_root.name or str(best_root)
            parts.append(root_name)
            if best_relative is not None:
                rel_parts = Path(best_relative).parts[:-1]
                parts.extend(rel_parts)
            return parts

        # Fallback: walk parents from top-level to nearest
        parent = resolved.parent
        while parent and parent != parent.parent:
            if parent.name:
                parts.insert(0, parent.name)
            parent = parent.parent
        return parts


class AudioProcessingThread(QThread):
    """Audio processing thread with progress, pause/resume, cancellation, and logging"""

    processing_started = Signal(int)
    progress_updated = Signal(int, int, str)
    processing_paused = Signal()
    processing_resumed = Signal()
    processing_completed = Signal(int, int, float, bool, str)
    processing_cancelled = Signal(int, int, bool, str)
    error_occurred = Signal(str)
    log_ready = Signal(str)

    def __init__(
        self,
        files: List[Dict[str, Any]],
        duration_ranges: List[Dict[str, Any]],
        mode: str,
        output_dir: Path,
        dry_run: bool,
        category_key: Optional[str],
    ):
        super().__init__()
        self.files = files
        self.duration_ranges = deep_copy_duration_ranges(duration_ranges)
        self.mode = mode
        self.output_dir = output_dir
        self.dry_run = dry_run
        self.category_key = category_key
        self.cancel_event = Event()
        self.pause_event = Event()
        self._log_entries: List[str] = []
        self._processed_files = 0

    def request_cancel(self):
        self.cancel_event.set()

    def request_pause(self):
        self.pause_event.set()

    def resume(self):
        if self.pause_event.is_set():
            self.pause_event.clear()

    def _wait_if_paused(self):
        was_paused = False
        if self.pause_event.is_set() and not self.cancel_event.is_set():
            self.processing_paused.emit()
            was_paused = True
        while self.pause_event.is_set() and not self.cancel_event.is_set():
            time.sleep(0.1)
        if was_paused and not self.cancel_event.is_set():
            self.processing_resumed.emit()

    def _append_log(self, message: str):
        timestamp = datetime.now().strftime('%H:%M:%S')
        self._log_entries.append(f"[{timestamp}] {message}")

    def _finalize_log(self, status: str, elapsed: float) -> Optional[Path]:
        try:
            log_dir = PROJECT_ROOT / "logs"
            log_dir.mkdir(exist_ok=True)
            summary = (
                f"status={status} mode={self.mode} dry_run={self.dry_run} "
                f"processed={self._processed_files} total={len(self.files)} elapsed={elapsed:.2f}s"
            )
            self._log_entries.append(summary)
            log_path = log_dir / f"audio_processing_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
            log_path.write_text("\n".join(self._log_entries), encoding="utf-8")
            return log_path
        except Exception:
            return None

    def run(self):
        start_time = time.monotonic()
        self._log_entries = []
        self._append_log(f"processing start mode={self.mode} targets={len(self.files)} dry_run={self.dry_run}")

        try:
            total_files = len(self.files)
            if total_files == 0:
                elapsed = time.monotonic() - start_time
                log_path = self._finalize_log("empty", elapsed)
                if log_path:
                    self.log_ready.emit(str(log_path))
                self.processing_completed.emit(0, 0, elapsed, self.dry_run, self.mode)
                return

            self.processing_started.emit(total_files)

            success_count = 0
            error_count = 0

            for index, file_info in enumerate(self.files, start=1):
                if self.cancel_event.is_set():
                    break

                self._wait_if_paused()
                source_path = Path(file_info.get('path', ''))
                self.progress_updated.emit(index, total_files, str(source_path))

                try:
                    if not source_path.exists():
                        raise FileNotFoundError(str(source_path))

                    if self.mode == "フラット化":
                        target_path = unique_name(self.output_dir, source_path.name)
                    elif self.mode == "音声整理" and self.category_key:
                        categories = categorize_audio(file_info, self.duration_ranges)
                        subdir_name = categories.get(self.category_key, "unknown")
                        target_path = unique_name(self.output_dir / subdir_name, source_path.name)
                    else:
                        target_path = unique_name(self.output_dir, source_path.name)

                    if not self.dry_run:
                        target_path.parent.mkdir(parents=True, exist_ok=True)
                        shutil.copy2(source_path, target_path)

                    success_count += 1
                    self._append_log(f"ok {source_path} -> {target_path}")

                except Exception as exc:
                    error_count += 1
                    self._append_log(f"fail {source_path}: {exc}")

                self._processed_files = index

            elapsed = time.monotonic() - start_time

            if self.cancel_event.is_set():
                log_path = self._finalize_log("cancelled", elapsed)
                if log_path:
                    self.log_ready.emit(str(log_path))
                self.processing_cancelled.emit(success_count, error_count, self.dry_run, self.mode)
                return

            log_path = self._finalize_log("completed", elapsed)
            if log_path:
                self.log_ready.emit(str(log_path))

            self.processing_completed.emit(success_count, error_count, elapsed, self.dry_run, self.mode)

        except Exception as exc:
            elapsed = time.monotonic() - start_time
            self._append_log(f"error {exc}")
            log_path = self._finalize_log("error", elapsed)
            if log_path:
                self.log_ready.emit(str(log_path))
            self.error_occurred.emit(str(exc))
class AudioAnalyzerWindow(QMainWindow):
    """Enhanced audio analyzer with comprehensive analysis and sorting capabilities"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("音声解析・整理ツール")
        self.setGeometry(200, 200, 1400, 900)
        self.setMinimumSize(1200, 800)
        
        # Data management
        self.selected_paths: List[Path] = []
        self.analysis_results: Dict[str, Any] = {}
        self.analysis_files: List[Dict[str, Any]] = []
        self.display_files: List[Dict[str, Any]] = []
        self.analysis_thread: Optional[AudioAnalysisThread] = None
        self.processing_thread: Optional["AudioProcessingThread"] = None
        self.detail_worker: Optional[AudioDetailExportWorker] = None
        self.analysis_buttons: List[QPushButton] = []
        self.is_analyzing: bool = False
        self.is_processing: bool = False
        self.current_operation: Optional[str] = None
        self.latest_log_path: Optional[str] = None
        self.operation_paused: bool = False
        self._pause_label_backup: Optional[str] = None
        self.folder_placeholder_text = "ここに音声フォルダをドラッグ&ドロップ"

        self.ffprobe_cache: Dict[str, Any] = _load_json(FFPROBE_CACHE_FILE)
        self.analysis_cache: Dict[str, Any] = _load_json(ANALYSIS_CACHE_FILE)
        self.detail_column_ids: List[str] = list(DEFAULT_DETAIL_COLUMN_IDS)
        self.pending_detail_export: Optional[Dict[str, Any]] = None
        self.problem_files = []

        self.duration_settings_path = get_duration_settings_path()
        self.duration_ranges: List[Dict[str, Any]] = load_duration_ranges(self.duration_settings_path)

        self.audio_extensions = {'.wav', '.mp3', '.flac', '.m4a', '.mp4', '.aif', '.aiff'}
        self.max_files_display_per_dir = 200

        self.init_ui()
        self.apply_pro_theme()
        self.setAcceptDrops(True)

    def closeEvent(self, event):
        try:
            self._save_analysis_cache()
            self._save_ffprobe_cache()
        except Exception:
            pass
        super().closeEvent(event)

    def init_ui(self):
        """Initialize the UI layout similar to analyzer"""
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(5, 5, 5, 5)
        main_layout.setSpacing(5)
        
        # Main splitter (vertical)
        vsplitter = QSplitter(Qt.Vertical)
        
        # Top: Audio folder tree
        folder_widget = self.create_folder_tree_widget()
        vsplitter.addWidget(folder_widget)
        
        # Middle: Toolbar and analysis results
        bottom_widget = QWidget()
        bottom_layout = QVBoxLayout(bottom_widget)
        bottom_layout.setContentsMargins(0, 0, 0, 0)
        bottom_layout.setSpacing(2)
        
        # Toolbar
        toolbar = self.create_toolbar()
        bottom_layout.addWidget(toolbar)
        
        # Analysis results and processing options in horizontal splitter
        hsplitter = QSplitter(Qt.Horizontal)
        
        # Left: Analysis results
        result_widget = self.create_result_widget()
        hsplitter.addWidget(result_widget)
        
        # Right: Processing options
        options_widget = self.create_options_widget()
        hsplitter.addWidget(options_widget)
        
        hsplitter.setSizes([700, 400])
        bottom_layout.addWidget(hsplitter)
        
        vsplitter.addWidget(bottom_widget)
        vsplitter.setSizes([300, 600])
        
        main_layout.addWidget(vsplitter)
        
        # Status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("音声ファイルフォルダを追加して解析を開始してください")
    
    def create_folder_tree_widget(self):
        """Create folder tree widget for audio folders"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # Header
        header_layout = QHBoxLayout()
        header_layout.addWidget(QLabel("音声フォルダ"))
        
        header_layout.addStretch()
        layout.addLayout(header_layout)
        
        # Tree view
        self.folder_tree = QTreeWidget()
        self.folder_tree.setHeaderHidden(True)
        self.folder_tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.folder_tree.setAcceptDrops(True)
        self.folder_tree.setMinimumHeight(200)
        
        # Placeholder
        self._add_placeholder_if_empty()
        
        layout.addWidget(self.folder_tree)

        return widget

    def _add_placeholder_if_empty(self):
        """Ensure guidance placeholder exists when no folders are registered."""
        if self.folder_tree.topLevelItemCount() == 0:
            placeholder = QTreeWidgetItem(self.folder_tree)
            placeholder.setText(0, self.folder_placeholder_text)
            placeholder.setFlags(Qt.NoItemFlags)
            placeholder.setForeground(0, QBrush(QColor("#666666")))

    def _register_analysis_button(self, button: QPushButton):
        if button not in self.analysis_buttons:
            self.analysis_buttons.append(button)

    def _set_analysis_controls_enabled(self, enabled: bool):
        for button in self.analysis_buttons:
            button.setEnabled(enabled)
        if hasattr(self, "pause_button"):
            self.pause_button.setEnabled(not enabled and (self.is_analyzing or self.is_processing))
            self.pause_button.setVisible(not enabled and (self.is_analyzing or self.is_processing))
        if hasattr(self, "resume_button"):
            self.resume_button.setVisible(False)
        if hasattr(self, "stop_button"):
            self.stop_button.setVisible(False)

    def _reset_operation_progress_ui(self, message: str = "準備完了"):
        if hasattr(self, "progress_bar") and self.progress_bar:
            self.progress_bar.setVisible(False)
            self.progress_bar.setRange(0, 1)
            self.progress_bar.setValue(0)

        if hasattr(self, "progress_label") and self.progress_label:
            self.progress_label.setText(message)
            self.progress_label.setVisible(bool(message))

        if hasattr(self, "pause_button"):
            self.pause_button.setEnabled(False)
            self.pause_button.setVisible(False)
        if hasattr(self, "resume_button"):
            self.resume_button.setEnabled(True)
            self.resume_button.setVisible(False)
        if hasattr(self, "stop_button"):
            self.stop_button.setEnabled(True)
            self.stop_button.setVisible(False)

        self.is_analyzing = False
        self.is_processing = False
        self.current_operation = None
        self.operation_paused = False
        self._pause_label_backup = None
        self.detail_worker = None
        self._set_analysis_controls_enabled(True)

    def _save_ffprobe_cache(self):
        _save_json(FFPROBE_CACHE_FILE, self.ffprobe_cache)

    def _save_analysis_cache(self):
        _save_json(ANALYSIS_CACHE_FILE, self.analysis_cache)

    def clear_ffprobe_cache(self):
        self.ffprobe_cache = {}
        try:
            if FFPROBE_CACHE_FILE.exists():
                FFPROBE_CACHE_FILE.unlink()
            self._save_ffprobe_cache()
        except Exception:
            pass

    def clear_analysis_cache(self):
        self.analysis_cache = {}
        try:
            if ANALYSIS_CACHE_FILE.exists():
                ANALYSIS_CACHE_FILE.unlink()
            self._save_analysis_cache()
        except Exception:
            pass

    def prompt_cache_clear(self):
        dialog = CacheClearDialog(self)
        if dialog.exec() != QDialog.Accepted:
            return
        selection = dialog.selection()
        messages = []
        if selection.get("analysis"):
            self.clear_analysis_cache()
            messages.append("解析キャッシュ")
        if selection.get("ffprobe"):
            self.clear_ffprobe_cache()
            messages.append("詳細キャッシュ")
        if messages:
            QMessageBox.information(self, "完了", "、".join(messages) + "を削除しました")
        else:
            QMessageBox.information(self, "情報", "キャッシュは削除されませんでした")

    def _merge_analysis_cache(self, cache_update: Optional[Dict[str, Any]]):
        if not cache_update:
            return
        self.analysis_cache.update(cache_update)
        self._save_analysis_cache()

    def _merge_ffprobe_cache(self, cache_update: Optional[Dict[str, Any]]):
        if not cache_update:
            return
        self.ffprobe_cache.update(cache_update)
        self._save_ffprobe_cache()
    
    def create_toolbar(self):
        """Create toolbar with audio-specific options"""
        toolbar = QWidget()
        toolbar.setMaximumHeight(40)
        layout = QHBoxLayout(toolbar)
        layout.setContentsMargins(5, 2, 5, 2)
        layout.setSpacing(5)

        # Folder selection
        add_btn = QPushButton("フォルダ選択")
        add_btn.clicked.connect(self.select_audio_folders)
        layout.addWidget(add_btn)
        
        # Remove selected
        remove_btn = QPushButton("選択削除")
        remove_btn.clicked.connect(self.remove_selected_folders)
        layout.addWidget(remove_btn)

        name_remove_btn = QPushButton("名前で削除")
        name_remove_btn.clicked.connect(self.remove_folders_by_name)
        layout.addWidget(name_remove_btn)

        # Analysis
        analyze_btn = QPushButton("音声解析実行")
        analyze_btn.setStyleSheet("background-color: #2d5a2d; color: white; font-weight: bold;")
        analyze_btn.clicked.connect(self.run_audio_analysis)
        layout.addWidget(analyze_btn)
        self._register_analysis_button(analyze_btn)

        duplicate_pick_btn = QPushButton("重複候補抽出")
        duplicate_pick_btn.clicked.connect(self.pickup_audio_duplicates)
        layout.addWidget(duplicate_pick_btn)

        corruption_pick_btn = QPushButton("破損候補抽出")
        corruption_pick_btn.clicked.connect(self.pickup_audio_corruption_candidates)
        layout.addWidget(corruption_pick_btn)

        duplicate_cleanup_btn = QPushButton("重複整理(1件残す)")
        duplicate_cleanup_btn.clicked.connect(self.cleanup_audio_duplicates_keep_one)
        layout.addWidget(duplicate_cleanup_btn)

        corruption_quarantine_btn = QPushButton("破損候補退避")
        corruption_quarantine_btn.clicked.connect(self.quarantine_audio_corruption_candidates)
        layout.addWidget(corruption_quarantine_btn)

        duration_btn = QPushButton("時間区分設定")
        duration_btn.clicked.connect(self.open_duration_settings)
        layout.addWidget(duration_btn)

        partial_btn = QPushButton("部分解析更新")
        partial_btn.clicked.connect(self.run_partial_analysis)
        layout.addWidget(partial_btn)
        self._register_analysis_button(partial_btn)

        layout.addWidget(QLabel("|"))
        
        # Processing mode
        layout.addWidget(QLabel("処理モード:"))
        self.processing_mode = QComboBox()
        self.processing_mode.addItems(["音声整理", "フラット化"])
        layout.addWidget(self.processing_mode)
        
        # Dry run
        self.dry_run_check = QCheckBox("シミュレーション")
        self.dry_run_check.setChecked(True)
        layout.addWidget(self.dry_run_check)
        
        layout.addStretch()
        
        # Clear all
        clear_btn = QPushButton("全クリア")
        clear_btn.setStyleSheet("color: #a94442;")
        clear_btn.clicked.connect(self.clear_all)
        layout.addWidget(clear_btn)
        
        return toolbar
    
    def create_result_widget(self):
        """Create analysis results widget"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # Header
        header = QLabel("音声解析結果")
        layout.addWidget(header)

        export_layout = QHBoxLayout()
        export_layout.setSpacing(8)

        export_csv_btn = QPushButton("CSVエクスポート")
        export_csv_btn.clicked.connect(self.export_results_csv)
        export_layout.addWidget(export_csv_btn)

        export_excel_btn = QPushButton("Excelエクスポート")
        export_excel_btn.clicked.connect(self.export_results_excel)
        export_layout.addWidget(export_excel_btn)

        export_layout.addWidget(QLabel("|"))

        detail_csv_btn = QPushButton("詳細CSV")
        detail_csv_btn.clicked.connect(self.export_detailed_csv)
        export_layout.addWidget(detail_csv_btn)
        self._register_analysis_button(detail_csv_btn)

        detail_excel_btn = QPushButton("詳細Excel")
        detail_excel_btn.clicked.connect(self.export_detailed_excel)
        export_layout.addWidget(detail_excel_btn)
        self._register_analysis_button(detail_excel_btn)

        cache_clear_btn = QPushButton("キャッシュ削除")
        cache_clear_btn.clicked.connect(self.prompt_cache_clear)
        export_layout.addWidget(cache_clear_btn)

        export_layout.addStretch()
        layout.addLayout(export_layout)

        # Category tabs
        self.result_tabs = QTabWidget()
        
        # Create tabs for different analysis categories
        self.create_analysis_tabs()
        
        layout.addWidget(self.result_tabs)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)

        progress_row = QHBoxLayout()
        self.progress_label = QLabel("準備完了")
        self.progress_label.setVisible(False)
        progress_row.addWidget(self.progress_label, 1)

        button_box = QHBoxLayout()
        button_box.setSpacing(4)

        self.pause_button = QPushButton("一時停止")
        self.pause_button.setVisible(False)
        self.pause_button.clicked.connect(self.pause_current_operation)
        button_box.addWidget(self.pause_button)

        self.resume_button = QPushButton("再開")
        self.resume_button.setVisible(False)
        self.resume_button.clicked.connect(self.resume_current_operation)
        button_box.addWidget(self.resume_button)

        self.stop_button = QPushButton("終了")
        self.stop_button.setVisible(False)
        self.stop_button.clicked.connect(self.stop_current_operation)
        button_box.addWidget(self.stop_button)

        progress_row.addLayout(button_box, 0)

        layout.addLayout(progress_row)

        return widget
    
    def create_analysis_tabs(self):
        """Create tabs for different audio analysis categories"""
        categories = [
            ("フォーマット", "format"),
            ("サンプルレート", "samplerate"), 
            ("チャンネル", "channels"),
            ("時間", "duration"),
            ("ビットレート", "bitrate"),
            ("日付", "date")
        ]
        
        self.category_trees = {}
        
        for tab_name, category_key in categories:
            tab_widget = QWidget()
            tab_layout = QVBoxLayout(tab_widget)
            
            tree = QTreeWidget()
            tree.setHeaderLabels(["カテゴリ", "ファイル数", "合計サイズ", "合計時間"])
            tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
            tree.setAlternatingRowColors(True)
            
            tab_layout.addWidget(tree)
            self.result_tabs.addTab(tab_widget, tab_name)
            self.category_trees[category_key] = tree
    
    def create_options_widget(self):
        """Create processing options widget"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # Processing options
        options_group = QGroupBox("処理オプション")
        options_layout = QVBoxLayout(options_group)
        
        # Format selection for flattening
        format_group = QGroupBox("保持フォーマット (フラット化時)")
        format_layout = QVBoxLayout(format_group)
        
        self.format_checks = {}
        formats = [
            ("wav", "WAV (PCM)", True),
            ("mp3", "MP3", False),
            ("flac", "FLAC", False),
            ("m4a", "AAC/M4A", False),
            ("aiff", "AIFF", False)
        ]
        
        for fmt_key, fmt_label, default in formats:
            check = QCheckBox(fmt_label)
            check.setChecked(default)
            self.format_checks[fmt_key] = check
            format_layout.addWidget(check)
        
        options_layout.addWidget(format_group)
        
        # Sorting criteria with advanced options
        sort_group = QGroupBox("整理基準")
        sort_layout = QVBoxLayout(sort_group)
        
        self.sort_criterion = QComboBox()
        self.sort_criterion.addItems([
            "フォーマット別",
            "サンプルレート別", 
            "チャンネル別",
            "時間別",
            "ビットレート別",
            "日付別",
            "アーティスト別",
            "アルバム別",
            "ジャンル別"
        ])
        sort_layout.addWidget(self.sort_criterion)
        
        # Advanced sorting options
        advanced_sort_group = QGroupBox("条件整理オプション")
        advanced_sort_layout = QVBoxLayout(advanced_sort_group)
        
        # Multi-criteria sorting
        multi_sort_layout = QHBoxLayout()
        multi_sort_layout.addWidget(QLabel("複数条件:"))
        
        self.multi_sort_check = QCheckBox("有効")
        multi_sort_layout.addWidget(self.multi_sort_check)
        
        self.secondary_criterion = QComboBox()
        self.secondary_criterion.addItems([
            "なし",
            "フォーマット別",
            "サンプルレート別", 
            "チャンネル別",
            "時間別",
            "ビットレート別",
            "日付別"
        ])
        self.secondary_criterion.setEnabled(False)
        multi_sort_layout.addWidget(self.secondary_criterion)
        
        self.multi_sort_check.toggled.connect(self.secondary_criterion.setEnabled)
        advanced_sort_layout.addLayout(multi_sort_layout)
        
        # Conditional filtering
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("フィルター:"))
        
        self.filter_enabled = QCheckBox("有効")
        filter_layout.addWidget(self.filter_enabled)
        
        self.filter_type = QComboBox()
        self.filter_type.addItems(["時間", "サイズ", "ビットレート"])
        self.filter_type.setEnabled(False)
        filter_layout.addWidget(self.filter_type)
        
        self.filter_condition = QComboBox()
        self.filter_condition.addItems(["以上", "以下", "範囲"])
        self.filter_condition.setEnabled(False)
        filter_layout.addWidget(self.filter_condition)
        
        self.filter_value = QLineEdit()
        self.filter_value.setPlaceholderText("値を入力")
        self.filter_value.setEnabled(False)
        filter_layout.addWidget(self.filter_value)
        
        def toggle_filter_controls(enabled):
            self.filter_type.setEnabled(enabled)
            self.filter_condition.setEnabled(enabled)
            self.filter_value.setEnabled(enabled)
        
        self.filter_enabled.toggled.connect(toggle_filter_controls)
        advanced_sort_layout.addLayout(filter_layout)
        
        sort_layout.addWidget(advanced_sort_group)
        options_layout.addWidget(sort_group)
        
        # Additional options
        additional_group = QGroupBox("追加オプション")
        additional_layout = QVBoxLayout(additional_group)
        
        self.delete_zip_check = QCheckBox("ZIP/圧縮ファイルを削除")
        self.delete_zip_check.setChecked(True)
        additional_layout.addWidget(self.delete_zip_check)
        
        self.remove_empty_check = QCheckBox("空フォルダを削除")
        self.remove_empty_check.setChecked(True)
        additional_layout.addWidget(self.remove_empty_check)
        
        self.use_trash_check = QCheckBox("不要ファイルをゴミ箱へ")
        additional_layout.addWidget(self.use_trash_check)
        
        options_layout.addWidget(additional_group)
        
        layout.addWidget(options_group)
        
        # Execute buttons
        button_layout = QHBoxLayout()
        
        execute_btn = QPushButton("実行")
        execute_btn.setStyleSheet("background-color: #28a745; color: white; font-weight: bold; padding: 8px 16px;")
        execute_btn.clicked.connect(self.execute_processing)
        button_layout.addWidget(execute_btn)
        
        layout.addLayout(button_layout)
        layout.addStretch()
        
        return widget
    
    def apply_pro_theme(self):
        """Apply Pro (dark) theme"""
        pro_theme_file = Path("themes/pro.qss")
        if pro_theme_file.exists():
            with open(pro_theme_file, "r", encoding="utf-8") as f:
                base_style = f.read()
        else:
            base_style = self.get_fallback_theme()
            
        # Audio analyzer specific styles
        audio_style = """
            QTabWidget::pane {
                border: 1px solid #5c5c5c;
                background-color: #2b2b2b;
            }
            
            QTabBar::tab {
                background-color: #3c3c3c;
                color: #cccccc;
                padding: 8px 16px;
                margin-right: 2px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            
            QTabBar::tab:selected {
                background-color: #007acc;
                color: #ffffff;
                font-weight: bold;
            }
            
            QTabBar::tab:hover:!selected {
                background-color: #4c4c4c;
            }
        """
        
        self.setStyleSheet(base_style + audio_style)
    
    def get_fallback_theme(self) -> str:
        """Fallback theme for Pro style"""
        return """
            QMainWindow { background-color: #2b2b2b; color: #ffffff; }
            QPushButton { 
                background-color: #3c3c3c; color: white; border: none; 
                padding: 8px 16px; border-radius: 4px; 
            }
            QPushButton:hover { background-color: #4c4c4c; }
            QTreeWidget { 
                background-color: #1e1e1e; color: #cccccc; 
                border: 1px solid #3c3c3c; 
            }
            QGroupBox { 
                color: #ffffff; border: 1px solid #5c5c5c; 
                border-radius: 5px; margin-top: 15px; font-weight: bold;
            }
            QGroupBox::title {
                subcontrol-origin: margin; left: 10px; padding: 0 5px; color: #4ec9b0;
            }
        """
    
    def select_audio_folders(self):
        """Select audio folders for analysis"""
        folder = QFileDialog.getExistingDirectory(
            self, "音声フォルダを選択", 
            str(Path.home()),
            QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks
        )
        
        if folder:
            self.add_audio_folder(Path(folder))
    
    def remove_selected_folders(self):
        """Remove selected folders from the list"""
        selected_items = self.folder_tree.selectedItems()
        if not selected_items:
            QMessageBox.information(self, "情報", "削除するフォルダを選択してください")
            return
            
        for item in selected_items:
            if item.parent() is None:  # Top level only
                path_str = item.data(0, Qt.UserRole)
                if path_str:
                    path_to_remove = Path(path_str)
                    if path_to_remove in self.selected_paths:
                        self.selected_paths.remove(path_to_remove)
                
                index = self.folder_tree.indexOfTopLevelItem(item)
                if index >= 0:
                    self.folder_tree.takeTopLevelItem(index)
        
        # Add placeholder if empty
        self._add_placeholder_if_empty()

        self.status_bar.showMessage("選択したフォルダを削除しました")

    def remove_folders_by_name(self):
        """Remove folders by matching their names via dialog."""
        dialog = FolderNameDeleteDialog(self)
        if dialog.exec() != QDialog.Accepted:
            return

        query = dialog.get_query()
        match_mode = dialog.get_match_mode()

        removed_paths = remove_folders_matching_query(
            self.folder_tree,
            self.selected_paths,
            query,
            match_mode=match_mode,
        )

        if not removed_paths:
            QMessageBox.information(self, "情報", f"『{query}』に該当するフォルダは見つかりませんでした。")
            return

        self._add_placeholder_if_empty()

        match_label = "完全一致" if match_mode == MATCH_EXACT else "部分一致"
        preview_names = ", ".join(path.name for path in removed_paths[:3])
        if len(removed_paths) > 3:
            preview_names += " ..."

        if preview_names:
            message = f"{len(removed_paths)}件のフォルダを削除 ({match_label}): {preview_names}"
        else:
            message = f"{len(removed_paths)}件のフォルダを削除 ({match_label})"

        self.status_bar.showMessage(message)

    def run_audio_analysis(self):
        """Run detailed audio analysis"""
        if self.is_analyzing or self.is_processing:
            QMessageBox.information(self, "情報", "別の処理が進行中です。中止するか完了をお待ちください。")
            return

        self.problem_files = []
        if not self.selected_paths:
            QMessageBox.warning(self, "警告", "解析する音声フォルダがありません")
            return
        
        for tree in self.category_trees.values():
            tree.clear()

        self.is_analyzing = True
        self.current_operation = "analysis"
        self.operation_paused = False
        self._pause_label_backup = None
        self.latest_log_path = None
        self.display_files = []
        self.analysis_results = {}
        self._set_analysis_controls_enabled(False)

        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)
        self.progress_bar.setValue(0)
        self.progress_label.setText("ファイル数を計測しています…")
        self.progress_label.setVisible(True)
        self.pause_button.setVisible(True)
        self.pause_button.setEnabled(True)
        self.resume_button.setVisible(False)
        self.stop_button.setVisible(False)
        self.status_bar.showMessage("音声ファイルをスキャンしています…")

        self.analysis_thread = AudioAnalysisThread(self.selected_paths, self.duration_ranges, analysis_cache=self.analysis_cache)
        self.analysis_thread.problem_files_found = []
        self.analysis_thread.analysis_started.connect(self.on_analysis_started)
        self.analysis_thread.progress_updated.connect(self.update_analysis_progress)
        self.analysis_thread.analysis_paused.connect(self.on_analysis_paused)
        self.analysis_thread.analysis_resumed.connect(self.on_analysis_resumed)
        self.analysis_thread.analysis_completed.connect(self.display_analysis_results)
        self.analysis_thread.analysis_cancelled.connect(self.on_analysis_cancelled)
        self.analysis_thread.error_occurred.connect(self.handle_analysis_error)
        self.analysis_thread.log_ready.connect(self.on_analysis_log_ready)
        self.analysis_thread.finished.connect(self.on_analysis_thread_finished)
        self.analysis_thread.start()
    
    def pause_current_operation(self):
        """Pause current analysis or processing."""
        if self.operation_paused:
            return
        if self.analysis_thread and self.analysis_thread.isRunning():
            self.pause_button.setEnabled(False)
            self.status_bar.showMessage("音声解析を一時停止しています…", 3000)
            self.analysis_thread.request_pause()
        elif self.processing_thread and self.processing_thread.isRunning():
            self.pause_button.setEnabled(False)
            self.status_bar.showMessage("音声整理を一時停止しています…", 3000)
            self.processing_thread.request_pause()

    def resume_current_operation(self):
        """Resume paused analysis or processing."""
        if not self.operation_paused:
            return
        if self.analysis_thread and self.analysis_thread.isRunning():
            self.resume_button.setEnabled(False)
            self.analysis_thread.resume()
        elif self.processing_thread and self.processing_thread.isRunning():
            self.resume_button.setEnabled(False)
            self.processing_thread.resume()

    def stop_current_operation(self):
        """Cancel the current operation regardless of state."""
        if self.analysis_thread and self.analysis_thread.isRunning():
            self.stop_button.setEnabled(False)
            self.status_bar.showMessage("音声解析を終了しています…", 5000)
            self.analysis_thread.request_cancel()
        if self.processing_thread and self.processing_thread.isRunning():
            self.stop_button.setEnabled(False)
            self.status_bar.showMessage("音声整理を終了しています…", 5000)
            self.processing_thread.request_cancel()
        if self.detail_worker and self.detail_worker.isRunning():
            self.stop_button.setEnabled(False)
            self.status_bar.showMessage("詳細エクスポートを終了しています…", 5000)
            self.detail_worker.request_cancel()

    def on_analysis_started(self, total_files: int):
        label_prefix = "音声整理" if self.current_operation == "processing" else "音声解析"
        if total_files > 0:
            self.progress_bar.setRange(0, total_files)
            self.progress_bar.setValue(0)
            self.progress_label.setText(f"0.0% (0/{total_files}) - {label_prefix}開始")
        else:
            self.progress_bar.setRange(0, 0)
            self.progress_label.setText("対象の音声ファイルが見つかりませんでした")
        self.progress_label.setVisible(True)

    def update_analysis_progress(self, processed: int, total: int, current_path: str):
        """Update analysis progress"""
        label_prefix = "音声整理" if self.current_operation == "processing" else "音声解析"
        if total > 0:
            self.progress_bar.setRange(0, total)
            self.progress_bar.setValue(min(processed, total))
            percent = min(100.0, (processed / total) * 100) if total else 0.0
            label_text = f"{percent:5.1f}% ({processed}/{total})"
        else:
            self.progress_bar.setRange(0, 0)
            label_text = f"{processed} 件処理済み"

        file_name = Path(current_path).name if current_path else "(解析中)"
        self.progress_label.setText(f"{label_text} - {file_name}")
        total_display = total if total > 0 else "?"
        self.status_bar.showMessage(f"{label_prefix}中: {file_name} ({processed}/{total_display})")

    def display_analysis_results(self, results: Dict[str, Any], audio_infos: List[Dict[str, Any]], elapsed: float):
        """Display detailed analysis results in category tabs"""
        if self.analysis_thread and getattr(self.analysis_thread, "updated_cache", None):
            self._merge_analysis_cache(self.analysis_thread.updated_cache)

        self.analysis_results = results or {}
        self.analysis_files = audio_infos or []
        self.display_files = list(audio_infos or [])

        if not results:
            message = "音声ファイルが見つかりませんでした"
            if self.latest_log_path:
                message += f" | ログ: {Path(self.latest_log_path).name}"
            QMessageBox.information(self, "結果", message)
            self.pending_detail_export = None
            self._reset_operation_progress_ui(message)
            self.status_bar.showMessage(message, 7000)
            return

        self._populate_result_trees()

        category_count = sum(len(cat_data) for cat_data in results.values())
        summary = f"音声解析完了: {category_count}カテゴリ ({elapsed:.1f}秒)"
        if self.latest_log_path:
            summary += f" | ログ: {Path(self.latest_log_path).name}"

        self._reset_operation_progress_ui(summary)
        self.status_bar.showMessage(summary, 7000)

        if self.pending_detail_export:
            params = self.pending_detail_export
            self.pending_detail_export = None
            files = self.display_files or self.analysis_files
            if files:
                self._start_detail_export(files, params["column_ids"], params["file_path"], params["format"])
            else:
                QMessageBox.information(self, "情報", "詳細出力に必要なファイルがありませんでした")

    def on_analysis_cancelled(self, results: Dict[str, Any], audio_infos: List[Dict[str, Any]]):
        """Handle user cancellation"""
        self.analysis_results = results or {}
        self.analysis_files = audio_infos or []
        self.display_files = list(audio_infos or [])

        if results:
            self._populate_result_trees()
            message = "音声解析を中止しました（途中結果を表示）"
        else:
            message = "音声解析を中止しました"

        if self.latest_log_path:
            message += f" | ログ: {Path(self.latest_log_path).name}"

        self.pending_detail_export = None
        self._merge_analysis_cache(getattr(self.analysis_thread, "updated_cache", None))
        self.pending_detail_export = None
        self._reset_operation_progress_ui(message)
        self.status_bar.showMessage(message, 7000)

    def on_analysis_log_ready(self, log_path: str):
        self.latest_log_path = log_path
        if not self.is_analyzing and not self.is_processing:
            self.status_bar.showMessage(f"ログを保存しました: {Path(log_path).name}", 5000)

    def on_analysis_thread_finished(self):
        self.analysis_thread = None
        self.is_analyzing = False
        self._set_analysis_controls_enabled(True)

    def handle_analysis_error(self, error_message: str):
        """Handle analysis errors"""
        self._reset_operation_progress_ui("解析エラー")
        QMessageBox.critical(self, "解析エラー", f"音声解析中にエラーが発生しました:\n\n{error_message}")
        message = "音声解析エラー"
        if self.latest_log_path:
            message += f" | ログ: {Path(self.latest_log_path).name}"
        self.status_bar.showMessage(message, 7000)
        self._merge_analysis_cache(getattr(self.analysis_thread, "updated_cache", None))
        self.pending_detail_export = None

    def on_analysis_paused(self):
        if self.operation_paused:
            return
        self.operation_paused = True
        self._pause_label_backup = self.progress_label.text()
        if self._pause_label_backup:
            self.progress_label.setText(f"{self._pause_label_backup} [一時停止中]")
        self.pause_button.setVisible(False)
        self.resume_button.setEnabled(True)
        self.resume_button.setVisible(True)
        self.stop_button.setEnabled(True)
        self.stop_button.setVisible(True)
        self.status_bar.showMessage("音声解析を一時停止しました。再開または終了を選択してください。")

    def on_analysis_resumed(self):
        if not self.operation_paused:
            return
        self.operation_paused = False
        if self._pause_label_backup is not None:
            self.progress_label.setText(self._pause_label_backup)
        self._pause_label_backup = None
        self.resume_button.setVisible(False)
        self.resume_button.setEnabled(True)
        self.stop_button.setVisible(False)
        self.stop_button.setEnabled(True)
        self.pause_button.setEnabled(True)
        self.pause_button.setVisible(True)
        self.status_bar.showMessage("音声解析を再開しました", 3000)

    def on_processing_paused(self):
        if self.operation_paused:
            return
        self.operation_paused = True
        self._pause_label_backup = self.progress_label.text()
        if self._pause_label_backup:
            self.progress_label.setText(f"{self._pause_label_backup} [一時停止中]")
        self.pause_button.setVisible(False)
        self.resume_button.setEnabled(True)
        self.resume_button.setVisible(True)
        self.stop_button.setEnabled(True)
        self.stop_button.setVisible(True)
        self.status_bar.showMessage("音声整理を一時停止しました。再開または終了を選択してください。")

    def on_processing_resumed(self):
        if not self.operation_paused:
            return
        self.operation_paused = False
        if self._pause_label_backup is not None:
            self.progress_label.setText(self._pause_label_backup)
        self._pause_label_backup = None
        self.resume_button.setVisible(False)
        self.resume_button.setEnabled(True)
        self.stop_button.setVisible(False)
        self.stop_button.setEnabled(True)
        self.pause_button.setEnabled(True)
        self.pause_button.setVisible(True)
        self.status_bar.showMessage("音声整理を再開しました", 3000)

    def on_processing_completed(self, success_count: int, error_count: int, elapsed: float, dry_run: bool, mode: str):
        """Handle processing completion"""
        mode_text = f"{mode} {'シミュレーション' if dry_run else '実行'}"
        message = (
            f"{mode_text}が完了しました\n\n"
            f"成功: {success_count} ファイル\n"
            f"エラー: {error_count} ファイル\n"
            f"所要時間: {elapsed:.1f} 秒"
        )
        QMessageBox.information(self, "処理完了", message)

        summary = f"{mode_text}: 成功{success_count} エラー{error_count} ({elapsed:.1f}秒)"
        if self.latest_log_path:
            summary += f" | ログ: {Path(self.latest_log_path).name}"

        self._reset_operation_progress_ui(summary)
        self.status_bar.showMessage(summary, 7000)

    def on_processing_cancelled(self, success_count: int, error_count: int, dry_run: bool, mode: str):
        """Handle processing cancellation"""
        mode_text = f"{mode} {'シミュレーション' if dry_run else '実行'}"
        summary = f"{mode_text}を中止しました"
        if success_count or error_count:
            summary += f"（成功{success_count} / エラー{error_count}）"
        if self.latest_log_path:
            summary += f" | ログ: {Path(self.latest_log_path).name}"

        self._reset_operation_progress_ui(summary)
        self.status_bar.showMessage(summary, 7000)

    def on_processing_thread_finished(self):
        self.processing_thread = None
        self.is_processing = False
        if not self.is_analyzing:
            self._set_analysis_controls_enabled(True)

    def handle_processing_error(self, error_message: str):
        """Handle processing errors"""
        self._reset_operation_progress_ui("処理エラー")
        QMessageBox.critical(self, "処理エラー", f"音声整理中にエラーが発生しました:\n\n{error_message}")
        message = "音声整理エラー"
        if self.latest_log_path:
            message += f" | ログ: {Path(self.latest_log_path).name}"
        self.status_bar.showMessage(message, 7000)

    def _populate_result_trees(self):
        """Populate the results tabs using current analysis data."""
        category_names = self._get_category_display_names()

        for category, tree in self.category_trees.items():
            tree.clear()
            category_data = self.analysis_results.get(category, {}) if self.analysis_results else {}
            if not category_data:
                continue

            names = category_names.get(category, {})
            subcategory_keys = list(category_data.keys())
            if names:
                ordered = [key for key in names.keys() if key in category_data]
                extras = [key for key in category_data.keys() if key not in names]
                subcategory_keys = ordered + extras

            for subcategory in subcategory_keys:
                data = category_data[subcategory]
                item = QTreeWidgetItem(tree)
                item.setText(0, names.get(subcategory, subcategory))
                item.setText(1, f"{data['count']:,}")

                size_mb = data['total_size'] / (1024 * 1024)
                item.setText(2, f"{size_mb:.1f} MB" if size_mb >= 0.1 else "< 0.1 MB")

                total_duration = data.get('total_duration', 0)
                item.setText(3, self._format_total_duration(total_duration) if total_duration > 0 else "不明")

                item.setData(0, Qt.UserRole, subcategory)

        for tree in self.category_trees.values():
            tree.expandAll()
            tree.resizeColumnToContents(0)

    @staticmethod
    def _format_total_duration_static(total_seconds: Optional[float]) -> str:
        if total_seconds is None:
            return ""
        try:
            total = float(total_seconds)
        except (TypeError, ValueError):
            return ""
        if total < 0:
            total = 0
        hours = int(total // 3600)
        minutes = int((total % 3600) // 60)
        seconds = int(total % 60)
        if hours > 0:
            return f"{hours}h {minutes}m {seconds}s"
        if minutes > 0:
            return f"{minutes}m {seconds}s"
        return f"{seconds}s"

    def _format_total_duration(self, total_seconds: float) -> str:
        return AudioAnalyzerWindow._format_total_duration_static(total_seconds)

    @staticmethod
    def _format_channel_label_static(channels: Optional[int]) -> str:
        try:
            ch = int(channels) if channels is not None else None
        except (TypeError, ValueError):
            ch = None
        if ch == 1:
            return "モノラル"
        if ch == 2:
            return "ステレオ"
        if ch is not None:
            return f"{ch}ch"
        return ""

    def _format_channel_label(self, channels: Optional[int]) -> str:
        return AudioAnalyzerWindow._format_channel_label_static(channels)

    @staticmethod
    def _detail_column_label(column_id: str) -> str:
        for column in DETAIL_COLUMN_DEFS:
            if column["id"] == column_id:
                return column["label"]
        return column_id

    def rebuild_analysis_results(self):
        """Recalculate aggregation after duration ranges are updated."""
        source_files = self.display_files or self.analysis_files
        if not source_files:
            return
        self.analysis_results = aggregate_audio_data(source_files, self.duration_ranges)
        self._populate_result_trees()
        self.status_bar.showMessage("時間区分設定を適用しました")

    def open_duration_settings(self):
        """Open dialog to edit duration buckets dynamically."""
        dialog = DurationSettingsDialog(self.duration_ranges, self)
        if dialog.exec() != QDialog.Accepted:
            return

        new_ranges = dialog.get_ranges()
        self.duration_ranges = deep_copy_duration_ranges(new_ranges)

        try:
            save_duration_ranges(self.duration_settings_path, self.duration_ranges)
        except Exception as exc:
            QMessageBox.warning(self, "保存エラー", f"設定を保存できませんでした:\n{exc}")

        if self.analysis_files:
            self.rebuild_analysis_results()

    def run_partial_analysis(self):
        """Recalculate results based on selected folders only"""
        if self.is_analyzing or self.is_processing:
            QMessageBox.information(self, "情報", "処理中は部分解析を実行できません")
            return

        if not self.analysis_files:
            QMessageBox.warning(self, "警告", "先に音声解析を実行してください")
            return

        selected_items = self.folder_tree.selectedItems()
        if not selected_items:
            QMessageBox.information(self, "情報", "解析対象のフォルダを選択してください")
            return

        selected_paths = self._collect_selected_folder_paths(selected_items)
        if not selected_paths:
            QMessageBox.information(self, "情報", "フォルダを選択してください")
            return

        filtered_infos = [
            info for info in self.analysis_files
            if any(path_is_within(Path(info['path']), sel_path) for sel_path in selected_paths)
        ]

        if not filtered_infos:
            QMessageBox.information(self, "情報", "選択されたフォルダに解析済み音声がありません")
            return

        self.analysis_results = aggregate_audio_data(filtered_infos, self.duration_ranges)
        self.display_files = filtered_infos
        self._populate_result_trees()
        self.status_bar.showMessage("選択フォルダの部分解析を実行しました")

    def _get_category_display_names(self) -> Dict[str, Dict[str, str]]:
        names = copy.deepcopy(BASE_CATEGORY_NAME_MAP)
        duration_map = {entry["key"]: entry["label"] for entry in self.duration_ranges}
        duration_map[UNKNOWN_DURATION_KEY] = "不明"
        names["duration"] = duration_map
        return names

    def _collect_selected_folder_paths(self, items: List[QTreeWidgetItem]) -> List[Path]:
        """Extract directory paths from selected tree items."""
        paths: List[Path] = []
        seen = set()

        for item in items:
            data = item.data(0, Qt.UserRole)
            if not data:
                continue

            path = Path(str(data))
            if path.is_file():
                path = path.parent

            resolved = path.resolve() if path.exists() else path
            if resolved in seen:
                continue

            seen.add(resolved)
            paths.append(path)

        return paths

    def _prepare_summary_export_rows(self) -> Tuple[List[str], Dict[str, List[Dict[str, Any]]], List[Dict[str, Any]]]:
        headers = CATEGORY_DISPLAY_HEADERS
        if not self.analysis_results:
            return headers, {}, []

        category_names = self._get_category_display_names()
        per_category: Dict[str, List[Dict[str, Any]]] = {}
        summary_rows: List[Dict[str, Any]] = []

        processed_keys = set()
        for cat_key in CATEGORY_ORDER + [key for key in self.analysis_results.keys() if key not in CATEGORY_ORDER]:
            cat_data = self.analysis_results.get(cat_key)
            if not cat_data:
                continue
            processed_keys.add(cat_key)
            cat_rows: List[Dict[str, Any]] = []
            for sub_key, data in cat_data.items():
                total_size = data.get("total_size", 0) or 0
                total_duration = data.get("total_duration", 0) or 0
                size_mb = total_size / (1024 * 1024) if total_size else 0.0
                duration_label = self._format_total_duration(total_duration) if total_duration else "0s"
                row = {
                    "カテゴリ種別": CATEGORY_LABELS.get(cat_key, cat_key),
                    "カテゴリキー": sub_key,
                    "カテゴリ名": category_names.get(cat_key, {}).get(sub_key, sub_key),
                    "ファイル数": data.get("count", 0),
                    "合計サイズ(MB)": round(size_mb, 3),
                    "合計時間(秒)": round(float(total_duration), 3),
                    "合計時間(表示)": duration_label,
                }
                cat_rows.append(row)
            per_category[cat_key] = cat_rows
            summary_rows.extend(cat_rows)

        # include any leftover categories not in order list
        for cat_key, cat_data in self.analysis_results.items():
            if cat_key in processed_keys:
                continue
            cat_rows = []
            for sub_key, data in cat_data.items():
                total_size = data.get("total_size", 0) or 0
                total_duration = data.get("total_duration", 0) or 0
                size_mb = total_size / (1024 * 1024) if total_size else 0.0
                duration_label = self._format_total_duration(total_duration) if total_duration else "0s"
                cat_rows.append({
                    "カテゴリ種別": CATEGORY_LABELS.get(cat_key, cat_key),
                    "カテゴリキー": sub_key,
                    "カテゴリ名": category_names.get(cat_key, {}).get(sub_key, sub_key),
                    "ファイル数": data.get("count", 0),
                    "合計サイズ(MB)": round(size_mb, 3),
                    "合計時間(秒)": round(float(total_duration), 3),
                    "合計時間(表示)": duration_label,
                })
            per_category[cat_key] = cat_rows
            summary_rows.extend(cat_rows)

        return headers, per_category, summary_rows

    def _prepare_detail_export_rows(self) -> Tuple[List[str], List[Dict[str, Any]]]:
        files = self.display_files or self.analysis_files
        column_ids = DEFAULT_DETAIL_COLUMN_IDS
        headers = [self._detail_column_label(column_id) for column_id in column_ids]

        rows: List[Dict[str, Any]] = []
        for info in files:
            duration_value = info.get("duration")
            if duration_value and duration_value < 0:
                duration_value = None

            bitrate_value = info.get("bitrate")
            try:
                bitrate_value = float(bitrate_value) if bitrate_value is not None else None
            except (TypeError, ValueError):
                bitrate_value = None

            sample_rate = info.get("samplerate")
            try:
                sample_rate = int(float(sample_rate)) if sample_rate is not None else None
            except (TypeError, ValueError):
                sample_rate = None

            channel_label = self._format_channel_label(info.get("channels"))

            duration_seconds = None
            if duration_value is not None:
                try:
                    duration_seconds = round(float(duration_value), 3)
                except (TypeError, ValueError):
                    duration_seconds = None

            row = {
                self._detail_column_label("name"): info.get("name", ""),
                self._detail_column_label("format"): info.get("format") or info.get("ext", "").upper(),
                self._detail_column_label("samplerate"): sample_rate or "",
                self._detail_column_label("channels"): channel_label,
                self._detail_column_label("duration_seconds"): duration_seconds if duration_seconds is not None else "",
                self._detail_column_label("duration"): self._format_total_duration(duration_value) if duration_value else "",
                self._detail_column_label("bitrate"): bitrate_value if bitrate_value is not None else "",
                self._detail_column_label("codec"): info.get("codec_name") or (info.get("format") or info.get("ext", "").upper()),
                self._detail_column_label("size"): int(info.get("size", 0) or 0),
            }
            rows.append(row)

        return headers, rows

    def export_results_csv(self):
        """Export current analysis summary as CSV."""
        if not self.analysis_results:
            QMessageBox.information(self, "情報", "先に音声解析を実行してください")
            return

        headers, _, summary_rows = self._prepare_summary_export_rows()
        if not summary_rows:
            QMessageBox.information(self, "情報", "出力可能な解析結果がありません")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "CSVエクスポート",
            str(Path.home() / "analysis_summary.csv"),
            "CSVファイル (*.csv)"
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".csv"):
            file_path += ".csv"

        try:
            with open(file_path, "w", newline="", encoding="utf-8-sig") as fp:
                writer = csv.DictWriter(fp, fieldnames=headers)
                writer.writeheader()
                for row in summary_rows:
                    writer.writerow({key: row.get(key, "") for key in headers})
        except Exception as exc:
            QMessageBox.critical(self, "保存エラー", f"CSVを書き出せませんでした:\n{exc}")
            return

        QMessageBox.information(self, "完了", f"CSVを保存しました:\n{file_path}")

    def export_results_excel(self):
        """Export current analysis summary/detail as Excel workbook."""
        if not self.analysis_results:
            QMessageBox.information(self, "情報", "先に音声解析を実行してください")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.warning(
                self,
                "依存関係エラー",
                "Excel出力には openpyxl が必要です。\n`pip install openpyxl` を実行してから再試行してください。"
            )
            return

        headers, per_category, summary_rows = self._prepare_summary_export_rows()
        detail_headers, detail_rows = self._prepare_detail_export_rows()

        if not summary_rows:
            QMessageBox.information(self, "情報", "出力可能な解析結果がありません")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Excelエクスポート",
            str(Path.home() / "analysis_summary.xlsx"),
            "Excelファイル (*.xlsx)"
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            wb = Workbook()
            # Remove default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)

            for cat_key in CATEGORY_ORDER:
                rows = per_category.get(cat_key)
                if not rows:
                    continue
                sheet_name = CATEGORY_LABELS.get(cat_key, cat_key)
                ws = wb.create_sheet(title=sheet_name[:31])
                ws.append(headers)
                for row in rows:
                    ws.append([row.get(h, "") for h in headers])
                ws.freeze_panes = "A2"
                for idx, column_cells in enumerate(ws.columns, start=1):
                    length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                    ws.column_dimensions[get_column_letter(idx)].width = min(max(length + 2, 12), 60)

            for cat_key, rows in per_category.items():
                if cat_key in CATEGORY_ORDER:
                    continue
                if not rows:
                    continue
                sheet_name = CATEGORY_LABELS.get(cat_key, cat_key)
                ws = wb.create_sheet(title=sheet_name[:31])
                ws.append(headers)
                for row in rows:
                    ws.append([row.get(h, "") for h in headers])
                ws.freeze_panes = "A2"
                for idx, column_cells in enumerate(ws.columns, start=1):
                    length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                    ws.column_dimensions[get_column_letter(idx)].width = min(max(length + 2, 12), 60)

            if detail_rows:
                ws_detail = wb.create_sheet(title="ファイル詳細")
                ws_detail.append(detail_headers)
                for detail in detail_rows:
                    ws_detail.append([detail.get(h, "") for h in detail_headers])
                ws_detail.freeze_panes = "A2"
                for idx, column_cells in enumerate(ws_detail.columns, start=1):
                    length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                    ws_detail.column_dimensions[get_column_letter(idx)].width = min(max(length + 2, 12), 80)

            wb.save(file_path)
        except Exception as exc:
            QMessageBox.critical(self, "保存エラー", f"Excelファイルを書き出せませんでした:\n{exc}")
            return

        QMessageBox.information(self, "完了", f"Excelファイルを保存しました:\n{file_path}")

    def _get_all_audio_files(self) -> List[Dict[str, Any]]:
        files = self.display_files or self.analysis_files
        if not files:
            return []
        seen = set()
        unique_files: List[Dict[str, Any]] = []
        for info in files:
            path = info.get("path")
            if not path or path in seen:
                continue
            seen.add(path)
            unique_files.append(info)
        return unique_files

    def _show_pickup_dialog(self, *, title: str, headers: List[str], rows: List[List[Any]], default_filename: str):
        dialog = QDialog(self)
        dialog.setWindowTitle(title)
        dialog.setMinimumSize(980, 620)
        layout = QVBoxLayout(dialog)

        summary = QLabel(f"{len(rows):,} 件")
        summary.setStyleSheet("font-weight: bold;")
        layout.addWidget(summary)

        table = QTableWidget()
        table.setColumnCount(len(headers))
        table.setRowCount(len(rows))
        table.setHorizontalHeaderLabels(headers)
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)

        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                table.setItem(r, c, QTableWidgetItem("" if val is None else str(val)))

        table.horizontalHeader().setStretchLastSection(True)
        table.resizeColumnsToContents()
        layout.addWidget(table)

        buttons = QHBoxLayout()
        export_btn = QPushButton("CSV出力")
        close_btn = QPushButton("閉じる")
        buttons.addWidget(export_btn)
        buttons.addStretch()
        buttons.addWidget(close_btn)
        layout.addLayout(buttons)

        def _export_csv():
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "CSVとして保存",
                default_filename,
                "CSV files (*.csv)"
            )
            if not file_path:
                return
            try:
                with open(file_path, "w", newline="", encoding="utf-8-sig") as fp:
                    writer = csv.writer(fp)
                    writer.writerow(headers)
                    writer.writerows(rows)
                QMessageBox.information(self, "出力完了", f"CSVを保存しました:\n{file_path}")
            except Exception as exc:
                QMessageBox.critical(self, "出力エラー", f"CSV出力に失敗しました:\n{exc}")

        export_btn.clicked.connect(_export_csv)
        close_btn.clicked.connect(dialog.accept)
        dialog.exec()

    def _score_keep_audio_candidate(self, info: Dict[str, Any]) -> tuple:
        path = (info.get("path") or "").lower()
        preferred_keywords = ("納品", "deliver", "delivery", "master", "final")
        preferred = any(k in path for k in preferred_keywords)
        mtime = float(info.get("mtime", 0) or 0)
        size = int(info.get("size", 0) or 0)
        name = info.get("name") or Path(info.get("path", "")).name
        return (1 if preferred else 0, mtime, size, name)

    def _build_audio_duplicate_groups(self, mode: str, files: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        for info in files:
            path = info.get("path")
            if not path:
                continue
            if mode == "内容重複":
                key = get_file_hash(Path(path))
            else:
                key = info.get("name") or Path(path).name
            if key:
                groups[key].append(info)
        return {k: v for k, v in groups.items() if len(v) >= 2}

    def pickup_audio_duplicates(self):
        files = self._get_all_audio_files()
        if not files:
            QMessageBox.information(self, "情報", "先に音声解析を実行してください")
            return

        by_name: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        by_hash: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

        for info in files:
            path = info.get("path")
            if not path:
                continue
            name = info.get("name") or Path(path).name
            by_name[name].append(info)
            h = get_file_hash(Path(path))
            if h:
                by_hash[h].append(info)

        rows: List[List[Any]] = []
        for name, group in by_name.items():
            if len(group) < 2:
                continue
            for info in group:
                rows.append(["名前重複", name, len(group), info.get("path", ""), int(info.get("size", 0) or 0)])

        for h, group in by_hash.items():
            if len(group) < 2:
                continue
            for info in group:
                rows.append(["内容重複", h, len(group), info.get("path", ""), int(info.get("size", 0) or 0)])

        if not rows:
            QMessageBox.information(self, "結果", "重複候補は見つかりませんでした")
            return

        self._show_pickup_dialog(
            title="音声重複候補一覧",
            headers=["種別", "重複キー", "同一件数", "ファイルパス", "サイズ(バイト)"],
            rows=rows,
            default_filename=f"audio_duplicates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )
        self.status_bar.showMessage(f"音声重複候補を抽出: {len(rows):,}件")

    def _collect_audio_corruption_actions(self) -> List[Dict[str, Any]]:
        files = self._get_all_audio_files()
        actions: List[Dict[str, Any]] = []
        for info in files:
            path = info.get("path", "")
            if not path:
                continue
            size = int(info.get("size", 0) or 0)
            ext = (info.get("ext") or "").lower()
            reasons: List[str] = []

            if size == 0:
                reasons.append("0バイト")
            if info.get("analysis_error"):
                reasons.append(f"解析エラー: {info.get('analysis_error')}")
            if size > 0:
                if info.get("duration") is None:
                    reasons.append("再生時間取得不可")
                if info.get("samplerate") is None:
                    reasons.append("サンプルレート取得不可")
                if info.get("channels") is None:
                    reasons.append("チャンネル取得不可")
                h = get_file_hash(Path(path))
                if not h:
                    reasons.append("ハッシュ計算不可")

            if reasons:
                actions.append({
                    "path": path,
                    "ext": ext,
                    "size": size,
                    "reason": " / ".join(dict.fromkeys(reasons)),
                })
        return actions

    def pickup_audio_corruption_candidates(self):
        actions = self._collect_audio_corruption_actions()
        if not actions:
            QMessageBox.information(self, "結果", "破損候補は見つかりませんでした")
            return

        rows = [[a["path"], a["ext"], a["size"], a["reason"]] for a in actions]
        self._show_pickup_dialog(
            title="音声破損候補一覧",
            headers=["ファイルパス", "拡張子", "サイズ(バイト)", "判定理由"],
            rows=rows,
            default_filename=f"audio_corruption_candidates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )
        self.status_bar.showMessage(f"音声破損候補を抽出: {len(rows):,}件")

    def cleanup_audio_duplicates_keep_one(self):
        files = self._get_all_audio_files()
        if not files:
            QMessageBox.information(self, "情報", "先に音声解析を実行してください")
            return

        mode, ok = QInputDialog.getItem(
            self,
            "重複整理モード",
            "整理対象:",
            ["内容重複", "名前重複"],
            0,
            False,
        )
        if not ok:
            return

        groups = self._build_audio_duplicate_groups(mode, files)
        if not groups:
            QMessageBox.information(self, "結果", f"{mode} の重複グループは見つかりませんでした")
            return

        actions: List[Dict[str, Any]] = []
        for dup_key, group in groups.items():
            sorted_group = sorted(group, key=self._score_keep_audio_candidate, reverse=True)
            keep = sorted_group[0]
            for info in sorted_group[1:]:
                actions.append({
                    "mode": mode,
                    "dup_key": dup_key,
                    "keep_path": keep.get("path", ""),
                    "remove_path": info.get("path", ""),
                    "size": int(info.get("size", 0) or 0),
                })

        if not actions:
            QMessageBox.information(self, "結果", "整理対象はありませんでした")
            return

        preview_rows = [[a["mode"], a["dup_key"], a["keep_path"], a["remove_path"], a["size"]] for a in actions]
        if self.dry_run_check.isChecked():
            self._show_pickup_dialog(
                title=f"音声重複整理プレビュー ({mode})",
                headers=["種別", "重複キー", "残すファイル", "移動対象", "サイズ(バイト)"],
                rows=preview_rows,
                default_filename=f"audio_duplicate_cleanup_preview_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            )
            self.status_bar.showMessage(f"音声重複整理プレビュー: {len(actions):,}件")
            return

        reply = QMessageBox.question(
            self,
            "実行確認",
            f"{mode} の重複を整理します。\n"
            f"重複グループ: {len(groups):,}\n"
            f"移動対象: {len(actions):,}\n\n"
            "各グループで1件だけ残し、残りを退避フォルダへ移動します。実行しますか？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        base_dir = QFileDialog.getExistingDirectory(
            self,
            "重複音声の退避先フォルダを選択",
            str(Path.home()),
            QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks
        )
        if not base_dir:
            return

        quarantine = Path(base_dir) / f"audio_duplicates_removed_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        quarantine.mkdir(parents=True, exist_ok=True)

        success = 0
        errors = 0
        for idx, a in enumerate(actions, start=1):
            src = Path(a["remove_path"])
            if not src.exists():
                errors += 1
                continue
            try:
                dst = unique_name(quarantine, src.name)
                shutil.move(str(src), str(dst))
                success += 1
            except Exception:
                errors += 1

            if idx % 20 == 0 or idx == len(actions):
                self.status_bar.showMessage(f"音声重複整理中... {idx}/{len(actions)}")
                QApplication.processEvents()

        QMessageBox.information(
            self,
            "音声重複整理完了",
            f"完了しました。\n\n移動成功: {success}\nエラー: {errors}\n退避先: {quarantine}"
        )
        self.status_bar.showMessage(f"音声重複整理完了: 成功{success} / エラー{errors}")

    def quarantine_audio_corruption_candidates(self):
        actions = self._collect_audio_corruption_actions()
        if not actions:
            QMessageBox.information(self, "結果", "退避対象の破損候補は見つかりませんでした")
            return

        rows = [[a["path"], a["ext"], a["size"], a["reason"]] for a in actions]
        if self.dry_run_check.isChecked():
            self._show_pickup_dialog(
                title="音声破損候補退避プレビュー",
                headers=["ファイルパス", "拡張子", "サイズ(バイト)", "判定理由"],
                rows=rows,
                default_filename=f"audio_corruption_quarantine_preview_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            )
            self.status_bar.showMessage(f"音声破損候補退避プレビュー: {len(actions):,}件")
            return

        base_dir = QFileDialog.getExistingDirectory(
            self,
            "破損候補音声の退避先フォルダを選択",
            str(Path.home()),
            QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks
        )
        if not base_dir:
            return

        quarantine = Path(base_dir) / f"audio_corruption_candidates_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        quarantine.mkdir(parents=True, exist_ok=True)

        success = 0
        errors = 0
        for idx, a in enumerate(actions, start=1):
            src = Path(a["path"])
            if not src.exists():
                errors += 1
                continue
            try:
                dst = unique_name(quarantine, src.name)
                shutil.move(str(src), str(dst))
                success += 1
            except Exception:
                errors += 1

            if idx % 20 == 0 or idx == len(actions):
                self.status_bar.showMessage(f"音声破損候補退避中... {idx}/{len(actions)}")
                QApplication.processEvents()

        QMessageBox.information(
            self,
            "音声破損候補退避完了",
            f"完了しました。\n\n移動成功: {success}\nエラー: {errors}\n退避先: {quarantine}"
        )
        self.status_bar.showMessage(f"音声破損候補退避完了: 成功{success} / エラー{errors}")

    def export_detailed_csv(self):
        if self.detail_worker and self.detail_worker.isRunning():
            QMessageBox.information(self, "情報", "詳細エクスポートが進行中です。完了をお待ちください。")
            return

        if not self.selected_paths:
            QMessageBox.information(self, "情報", "先に音声フォルダを追加してください")
            return

        files = self.display_files or self.analysis_files

        selected_columns = self._ask_detail_columns()
        if not selected_columns:
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "詳細CSVエクスポート",
            str(Path.home() / "audio_detail.csv"),
            "CSVファイル (*.csv)"
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".csv"):
            file_path += ".csv"

        if files:
            self._start_detail_export(files, selected_columns, file_path, "csv")
            return

        self.pending_detail_export = {
            "column_ids": list(selected_columns),
            "file_path": file_path,
            "format": "csv",
        }

        if self.is_analyzing or self.is_processing:
            QMessageBox.information(self, "情報", "解析が完了すると自動的に詳細CSVを出力します。")
            self.status_bar.showMessage("解析完了後に詳細CSVを出力します", 7000)
        else:
            QMessageBox.information(self, "情報", "解析を開始して詳細CSVを出力します。")
            self.status_bar.showMessage("解析を開始して詳細CSVを出力します", 7000)
            self.run_audio_analysis()

    def export_detailed_excel(self):
        if self.detail_worker and self.detail_worker.isRunning():
            QMessageBox.information(self, "情報", "詳細エクスポートが進行中です。完了をお待ちください。")
            return

        if not self.selected_paths:
            QMessageBox.information(self, "情報", "先に音声フォルダを追加してください")
            return

        files = self.display_files or self.analysis_files

        selected_columns = self._ask_detail_columns()
        if not selected_columns:
            return

        try:
            from openpyxl import Workbook  # noqa: F401
        except ImportError:
            QMessageBox.warning(
                self,
                "依存関係エラー",
                "詳細Excel出力には openpyxl が必要です。\n`pip install openpyxl` を実行してから再試行してください。"
            )
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "詳細Excelエクスポート",
            str(Path.home() / "audio_detail.xlsx"),
            "Excelファイル (*.xlsx)"
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        if files:
            self._start_detail_export(files, selected_columns, file_path, "excel")
            return

        self.pending_detail_export = {
            "column_ids": list(selected_columns),
            "file_path": file_path,
            "format": "excel",
        }

        if self.is_analyzing or self.is_processing:
            QMessageBox.information(self, "情報", "解析が完了すると自動的に詳細Excelを出力します。")
            self.status_bar.showMessage("解析完了後に詳細Excelを出力します", 7000)
        else:
            QMessageBox.information(self, "情報", "解析を開始して詳細Excelを出力します。")
            self.status_bar.showMessage("解析を開始して詳細Excelを出力します", 7000)
            self.run_audio_analysis()

    def _ask_detail_columns(self) -> Optional[List[str]]:
        dialog = DetailExportOptionsDialog(self.detail_column_ids, self)
        if dialog.exec() != QDialog.Accepted:
            return None
        selected = dialog.selected_column_ids()
        if selected:
            self.detail_column_ids = selected
        return self.detail_column_ids

    def _start_detail_export(self, files: List[Dict[str, Any]], column_ids: List[str], file_path: str, export_format: str):
        total = len(files)
        if total == 0:
            QMessageBox.information(self, "情報", "出力できる詳細データがありません")
            return

        column_ids = list(column_ids)
        headers = [self._detail_column_label(column_id) for column_id in column_ids]

        self.pending_detail_export = None

        self.current_operation = "detail"
        self.is_analyzing = False
        self.is_processing = False
        self.operation_paused = False
        self._pause_label_backup = None
        self._set_analysis_controls_enabled(False)

        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, total if total else 1)
        self.progress_bar.setValue(0)
        self.progress_label.setText("詳細データを収集中…")
        self.progress_label.setVisible(True)
        self.pause_button.setVisible(False)
        self.resume_button.setVisible(False)
        self.stop_button.setVisible(True)
        self.stop_button.setEnabled(True)
        self.status_bar.showMessage("詳細データを収集中…")
        self.latest_log_path = None

        cache_snapshot = dict(self.ffprobe_cache)
        self.detail_worker = AudioDetailExportWorker(
            files,
            cache_snapshot,
            headers,
            file_path,
            export_format,
            column_ids,
            list(self.selected_paths),
        )
        self.detail_worker.progress_updated.connect(self.on_detail_progress)
        self.detail_worker.completed.connect(self.on_detail_completed)
        self.detail_worker.error_occurred.connect(self.on_detail_error)
        self.detail_worker.start()

    def on_detail_progress(self, processed: int, total: int, current_name: str):
        if total > 0:
            self.progress_bar.setRange(0, total)
            self.progress_bar.setValue(processed)
            percent = min(100.0, (processed / total) * 100) if total else 0.0
            label_text = f"{percent:5.1f}% ({processed}/{total})"
        else:
            self.progress_bar.setRange(0, 0)
            label_text = f"{processed} 件処理済み"

        display_name = current_name or "(処理中)"
        self.progress_label.setText(f"{label_text} - {display_name}")
        total_display = total if total > 0 else "?"
        self.status_bar.showMessage(f"詳細出力中: {display_name} ({processed}/{total_display})")

    def on_detail_completed(self, output_path: str, cache_updates: Dict[str, Any], log_path: str):
        self._merge_ffprobe_cache(cache_updates)

        if log_path:
            self.latest_log_path = log_path

        self._reset_operation_progress_ui()

        message = f"詳細ファイルを保存しました:\n{output_path}"
        if log_path:
            message += f"\nログ: {Path(log_path).name}"

        QMessageBox.information(self, "完了", message)
        if log_path:
            self.status_bar.showMessage(f"詳細エクスポート完了: {Path(output_path).name} | ログ: {Path(log_path).name}", 7000)
        else:
            self.status_bar.showMessage(f"詳細エクスポート完了: {Path(output_path).name}", 7000)

    def on_detail_error(self, message: str):
        if self.detail_worker and getattr(self.detail_worker, "updated_cache", None):
            self._merge_ffprobe_cache(self.detail_worker.updated_cache)

        self._reset_operation_progress_ui("詳細エクスポート失敗")
        if message.strip() == "ユーザーにより中止されました":
            QMessageBox.information(self, "中止", "詳細エクスポートを中止しました")
            self.status_bar.showMessage("詳細エクスポートを中止しました", 7000)
        else:
            QMessageBox.critical(self, "エラー", f"詳細エクスポートでエラーが発生しました:\n{message}")
            self.status_bar.showMessage("詳細エクスポートでエラーが発生しました", 7000)

    def execute_processing(self):
        """Execute audio processing based on settings"""
        if not self.analysis_results:
            QMessageBox.warning(self, "警告", "先に音声解析を実行してください")
            return
        
        # Get current tab (category) and selected items
        current_tab = self.result_tabs.currentIndex()
        if current_tab < 0:
            return
        
        category_keys = list(self.category_trees.keys())
        if current_tab >= len(category_keys):
            return
        
        current_category = category_keys[current_tab]
        current_tree = self.category_trees[current_category]
        selected_items = current_tree.selectedItems()
        
        if not selected_items:
            QMessageBox.warning(self, "警告", "処理対象を選択してください")
            return
        
        # Get selected files
        selected_files = []
        for item in selected_items:
            subcategory = item.data(0, Qt.UserRole)
            if subcategory and current_category in self.analysis_results:
                category_data = self.analysis_results[current_category].get(subcategory, {})
                files = category_data.get('files', [])
                selected_files.extend(files)
        
        if not selected_files:
            QMessageBox.warning(self, "警告", "処理対象ファイルがありません")
            return
        
        # Get output directory
        output_dir = QFileDialog.getExistingDirectory(
            self, "出力先フォルダを選択", 
            str(Path.home()),
            QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks
        )
        
        if not output_dir:
            return
        
        # Execute processing
        self._execute_audio_processing(selected_files, Path(output_dir))
    
    def _execute_audio_processing(self, files: List[Dict], output_dir: Path):
        """Execute the actual audio processing"""
        if self.is_analyzing or self.is_processing:
            QMessageBox.information(self, "情報", "別の処理が進行中です。中止するか完了をお待ちください。")
            return

        mode = self.processing_mode.currentText()
        is_dry_run = self.dry_run_check.isChecked()
        current_tab = self.result_tabs.currentIndex()
        category_keys = list(self.category_trees.keys())
        category_key = category_keys[current_tab] if 0 <= current_tab < len(category_keys) else None

        self.is_processing = True
        self.current_operation = "processing"
        self.operation_paused = False
        self._pause_label_backup = None
        self.latest_log_path = None
        self._set_analysis_controls_enabled(False)

        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)
        self.progress_bar.setValue(0)
        self.progress_label.setText("音声整理を準備しています…")
        self.progress_label.setVisible(True)
        self.pause_button.setVisible(True)
        self.pause_button.setEnabled(True)
        self.resume_button.setVisible(False)
        self.stop_button.setVisible(False)
        self.status_bar.showMessage("音声整理を準備しています…")

        self.processing_thread = AudioProcessingThread(
            files,
            self.duration_ranges,
            mode,
            output_dir,
            is_dry_run,
            category_key,
        )

        self.processing_thread.processing_started.connect(self.on_analysis_started)
        self.processing_thread.progress_updated.connect(self.update_analysis_progress)
        self.processing_thread.processing_paused.connect(self.on_processing_paused)
        self.processing_thread.processing_resumed.connect(self.on_processing_resumed)
        self.processing_thread.processing_completed.connect(self.on_processing_completed)
        self.processing_thread.processing_cancelled.connect(self.on_processing_cancelled)
        self.processing_thread.error_occurred.connect(self.handle_processing_error)
        self.processing_thread.log_ready.connect(self.on_analysis_log_ready)
        self.processing_thread.finished.connect(self.on_processing_thread_finished)
        self.processing_thread.start()
    
    def clear_all(self):
        """Clear all data"""
        reply = QMessageBox.question(self, "確認", "すべてをクリアしますか？")
        if reply == QMessageBox.Yes:
            self.selected_paths.clear()
            self.analysis_results.clear()
            self.folder_tree.clear()
            for tree in self.category_trees.values():
                tree.clear()
            
            # Add placeholder
            self._add_placeholder_if_empty()
            
            self.status_bar.showMessage("すべてクリアしました")
    
    # Drag and drop support
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
    
    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
    
    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            for url in urls:
                path = Path(url.toLocalFile())
                if path.is_dir():
                    self.add_audio_folder(path)
            event.acceptProposedAction()
    
    def add_audio_folder(self, folder_path: Path):
        """Add audio folder to the analysis list"""
        # Remove placeholder if present
        if self.folder_tree.topLevelItemCount() == 1:
            item = self.folder_tree.topLevelItem(0)
            if item.text(0) == self.folder_placeholder_text:
                self.folder_tree.clear()
        
        # Check if already exists
        if folder_path in self.selected_paths:
            return
        
        # Add to paths list
        self.selected_paths.append(folder_path)
        
        # Add to tree
        root_item = QTreeWidgetItem(self.folder_tree, [folder_path.name])
        root_item.setData(0, Qt.UserRole, str(folder_path))
        root_item.setToolTip(0, str(folder_path))
        
        audio_count = self._populate_tree_with_structure(root_item, folder_path)

        root_item.setExpanded(True)
        if audio_count is None:
            self.status_bar.showMessage(f"音声フォルダを追加しました: {folder_path.name}")
        else:
            self.status_bar.showMessage(f"音声フォルダを追加しました: {folder_path.name} (推定 {audio_count} ファイル)")

    def _populate_tree_with_structure(self, parent_item: QTreeWidgetItem, folder_path: Path) -> Optional[int]:
        """再帰的にフォルダ構造を追加し、推定の音声ファイル数を返す"""
        try:
            entries = sorted(
                [p for p in folder_path.iterdir() if p.name not in {'.DS_Store'}],
                key=lambda p: (p.is_file(), p.name.lower())
            )
        except Exception:
            return None

        total_audio = 0
        displayed_files = 0

        for entry in entries:
            if entry.is_dir() and not entry.is_symlink():
                dir_item = QTreeWidgetItem([entry.name])
                dir_item.setData(0, Qt.UserRole, str(entry))
                dir_item.setToolTip(0, str(entry))
                parent_item.addChild(dir_item)

                child_total = self._populate_tree_with_structure(dir_item, entry)
                if child_total is not None:
                    total_audio += child_total

            elif entry.is_file() and entry.suffix.lower() in self.audio_extensions:
                total_audio += 1
                if displayed_files < self.max_files_display_per_dir:
                    file_item = QTreeWidgetItem([f"🎵 {entry.name}"])
                    file_item.setData(0, Qt.UserRole, str(entry))
                    file_item.setToolTip(0, str(entry))
                    parent_item.addChild(file_item)
                    displayed_files += 1

        if total_audio > displayed_files:
            remaining = total_audio - displayed_files
            if remaining > 0:
                more_item = QTreeWidgetItem([f"... 他{remaining}個の音声ファイル"])
                more_item.setFlags(Qt.NoItemFlags)
                more_item.setForeground(0, QBrush(QColor("#888888")))
                parent_item.addChild(more_item)

        return total_audio


if __name__ == "__main__":
    from PySide6.QtWidgets import QApplication
    import sys
    
    app = QApplication(sys.argv)
    window = AudioAnalyzerWindow()
    window.show()
    sys.exit(app.exec())
