#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PySide6-based Multimedia Analysis and Processing Tool
Ultimate file analyzer integrating audio, video, image, document, and 3D analysis
The most comprehensive file analysis and organization tool
Based on all previous analyzer UI structures
"""

from PySide6.QtWidgets import *
from PySide6.QtCore import *
from PySide6.QtGui import *
from pathlib import Path
import sys
import json
import csv
from datetime import datetime
from typing import List, Dict, Any, Optional, Set
import shutil
import subprocess
import hashlib
import mimetypes
import struct
import re
import zipfile
import tarfile
import unicodedata

# Import all analysis engines
sys.path.append(str(Path(__file__).parent.parent))

from .folder_tools import (
    FolderNameDeleteDialog,
    MATCH_EXACT,
    remove_folders_matching_query,
)

# Independent analysis engine - avoid circular imports by directly using libraries
COMBINED_LIBRARY_STATUS = {}

# Audio analysis libraries
try:
    import mutagen
    AUDIO_MUTAGEN_AVAILABLE = True
    AUDIO_AVAILABLE = True  # 統一フラグ
    COMBINED_LIBRARY_STATUS['mutagen'] = {
        'available': True, 'version': mutagen.version_string,
        'description': '音声メタデータ解析', 'install_cmd': 'pip install mutagen'
    }
except ImportError:
    AUDIO_MUTAGEN_AVAILABLE = False
    AUDIO_AVAILABLE = False  # 統一フラグ
    COMBINED_LIBRARY_STATUS['mutagen'] = {
        'available': False, 'error': 'Not installed',
        'description': '音声メタデータ解析', 'install_cmd': 'pip install mutagen'
    }

try:
    import wave, aifc
    AUDIO_BASIC_AVAILABLE = True
except ImportError:
    AUDIO_BASIC_AVAILABLE = False

# Video analysis libraries
try:
    import subprocess
    result = subprocess.run(['ffprobe', '-version'], capture_output=True, text=True, timeout=5)
    VIDEO_FFPROBE_AVAILABLE = result.returncode == 0
    VIDEO_AVAILABLE = VIDEO_FFPROBE_AVAILABLE  # 統一フラグ
    COMBINED_LIBRARY_STATUS['ffprobe'] = {
        'available': VIDEO_FFPROBE_AVAILABLE,
        'version': 'installed' if VIDEO_FFPROBE_AVAILABLE else 'not found',
        'description': '動画メタデータ解析', 'install_cmd': 'Install ffmpeg package'
    }
except (subprocess.TimeoutExpired, FileNotFoundError, OSError):
    VIDEO_FFPROBE_AVAILABLE = False
    VIDEO_AVAILABLE = False  # 統一フラグ
    COMBINED_LIBRARY_STATUS['ffprobe'] = {
        'available': False, 'error': 'ffprobe not found',
        'description': '動画メタデータ解析', 'install_cmd': 'Install ffmpeg package'
    }

# Image analysis libraries
try:
    from PIL import Image
    from PIL.ExifTags import TAGS
    IMAGE_PIL_AVAILABLE = True
    IMAGE_AVAILABLE = True  # 統一フラグ
    COMBINED_LIBRARY_STATUS['Pillow'] = {
        'available': True, 'version': Image.__version__ if hasattr(Image, '__version__') else 'available',
        'description': '画像解析とEXIFデータ抽出', 'install_cmd': 'pip install Pillow'
    }
except ImportError:
    IMAGE_PIL_AVAILABLE = False
    IMAGE_AVAILABLE = False  # 統一フラグ
    COMBINED_LIBRARY_STATUS['Pillow'] = {
        'available': False, 'error': 'Not installed',
        'description': '画像解析とEXIFデータ抽出', 'install_cmd': 'pip install Pillow'
    }

# Document analysis libraries
try:
    import PyPDF2
    DOCUMENT_PDF_AVAILABLE = True
    COMBINED_LIBRARY_STATUS['PyPDF2'] = {
        'available': True, 'version': PyPDF2.__version__ if hasattr(PyPDF2, '__version__') else 'available',
        'description': 'PDF文書解析', 'install_cmd': 'pip install PyPDF2'
    }
except ImportError:
    DOCUMENT_PDF_AVAILABLE = False
    COMBINED_LIBRARY_STATUS['PyPDF2'] = {
        'available': False, 'error': 'Not installed',
        'description': 'PDF文書解析', 'install_cmd': 'pip install PyPDF2'
    }

try:
    import docx
    DOCUMENT_DOCX_AVAILABLE = True
    COMBINED_LIBRARY_STATUS['python-docx'] = {
        'available': True, 'version': 'available',
        'description': 'Word文書解析', 'install_cmd': 'pip install python-docx'
    }
except ImportError:
    DOCUMENT_DOCX_AVAILABLE = False
    COMBINED_LIBRARY_STATUS['python-docx'] = {
        'available': False, 'error': 'Not installed',
        'description': 'Word文書解析', 'install_cmd': 'pip install python-docx'
    }

# Document統合フラグ
DOC_AVAILABLE = DOCUMENT_PDF_AVAILABLE and DOCUMENT_DOCX_AVAILABLE

try:
    import chardet
    DOCUMENT_CHARDET_AVAILABLE = True
    COMBINED_LIBRARY_STATUS['chardet'] = {
        'available': True, 'version': chardet.__version__ if hasattr(chardet, '__version__') else 'available',
        'description': '文字エンコーディング検出', 'install_cmd': 'pip install chardet'
    }
except ImportError:
    DOCUMENT_CHARDET_AVAILABLE = False
    COMBINED_LIBRARY_STATUS['chardet'] = {
        'available': False, 'error': 'Not installed',
        'description': '文字エンコーディング検出', 'install_cmd': 'pip install chardet'
    }

# 3D analysis libraries
try:
    import trimesh
    import numpy
    THREED_AVAILABLE = True
    MODEL3D_AVAILABLE = True  # 統一フラグ
    COMBINED_LIBRARY_STATUS['trimesh'] = {
        'available': True, 'version': trimesh.__version__ if hasattr(trimesh, '__version__') else 'available',
        'description': '3Dモデル解析', 'install_cmd': 'pip install trimesh'
    }
    COMBINED_LIBRARY_STATUS['numpy'] = {
        'available': True, 'version': numpy.__version__ if hasattr(numpy, '__version__') else 'available',
        'description': '数値計算ライブラリ', 'install_cmd': 'pip install numpy'
    }
except ImportError:
    THREED_AVAILABLE = False
    MODEL3D_AVAILABLE = False  # 統一フラグ
    COMBINED_LIBRARY_STATUS['trimesh'] = {
        'available': False, 'error': 'Not installed',
        'description': '3Dモデル解析', 'install_cmd': 'pip install trimesh'
    }
    COMBINED_LIBRARY_STATUS['numpy'] = {
        'available': False, 'error': 'Not installed',
        'description': '数値計算ライブラリ', 'install_cmd': 'pip install numpy'
    }

# Additional libraries for multimedia analysis
try:
    import magic
    PYTHON_MAGIC_AVAILABLE = True
    COMBINED_LIBRARY_STATUS['python-magic'] = {
        'available': True,
        'version': 'available',
        'description': 'より正確なファイル形式検出',
        'install_cmd': 'pip install python-magic'
    }
except ImportError as e:
    PYTHON_MAGIC_AVAILABLE = False
    COMBINED_LIBRARY_STATUS['python-magic'] = {
        'available': False,
        'error': str(e),
        'description': 'より正確なファイル形式検出',
        'install_cmd': 'pip install python-magic'
    }

# Archive handling
try:
    import rarfile
    RARFILE_AVAILABLE = True
    COMBINED_LIBRARY_STATUS['rarfile'] = {
        'available': True,
        'version': 'available',
        'description': 'RARアーカイブの解析',
        'install_cmd': 'pip install rarfile'
    }
except ImportError as e:
    RARFILE_AVAILABLE = False
    COMBINED_LIBRARY_STATUS['rarfile'] = {
        'available': False,
        'error': str(e),
        'description': 'RARアーカイブの解析',
        'install_cmd': 'pip install rarfile'
    }

# Comprehensive file type definitions
MULTIMEDIA_FILE_TYPES = {
    # Audio formats
    'audio': {
        '.mp3', '.wav', '.flac', '.aac', '.ogg', '.wma', '.m4a', '.aiff', '.ape', 
        '.ac3', '.dts', '.opus', '.ra', '.au', '.snd', '.mka'
    },
    
    # Video formats  
    'video': {
        '.mp4', '.mkv', '.avi', '.mov', '.webm', '.m4v', '.3gp', '.flv', '.wmv', 
        '.mpg', '.mpeg', '.m2v', '.vob', '.ts', '.mts', '.m2ts', '.f4v', '.asf',
        '.rmvb', '.rm', '.divx', '.xvid', '.ogv', '.mxf', '.dv', '.dat'
    },
    
    # Image formats
    'image': {
        '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp', '.svg', 
        '.raw', '.cr2', '.nef', '.arw', '.dng', '.heic', '.heif', '.ico', '.psd',
        '.ai', '.eps', '.pdf', '.xcf', '.tga', '.pcx', '.pbm', '.pgm', '.ppm',
        '.exr', '.hdr', '.jp2', '.j2k', '.avif'
    },
    
    # Document formats
    'document': {
        '.pdf', '.doc', '.docx', '.docm', '.xls', '.xlsx', '.xlsm', '.ppt', '.pptx', 
        '.pptm', '.odt', '.ods', '.odp', '.rtf', '.txt', '.md', '.markdown', '.csv', 
        '.json', '.xml', '.html', '.htm', '.tex', '.bib', '.epub', '.mobi', '.azw',
        '.fb2', '.djvu', '.chm', '.lit'
    },
    
    # Code/Script formats
    'code': {
        '.py', '.js', '.java', '.cpp', '.c', '.h', '.hpp', '.cs', '.php', '.rb', 
        '.go', '.rs', '.swift', '.kt', '.scala', '.pl', '.sh', '.bat', '.ps1', 
        '.r', '.m', '.sql', '.css', '.scss', '.less', '.ts', '.jsx', '.tsx',
        '.vue', '.svelte', '.dart', '.lua', '.perl', '.asm', '.vb', '.pas',
        '.f90', '.f95', '.for', '.cob', '.cbl', '.ada', '.ml', '.hs', '.elm',
        '.clj', '.lisp', '.scheme', '.coffee', '.nim', '.crystal', '.zig'
    },
    
    # 3D Model formats
    '3d': {
        '.obj', '.stl', '.ply', '.off', '.gltf', '.glb', '.fbx', '.dae', '.x3d', 
        '.3ds', '.blend', '.ma', '.mb', '.c4d', '.max', '.lwo', '.3mf', '.amf', 
        '.wrl', '.vrml', '.x', '.md2', '.md3', '.md5', '.ase', '.lxo', '.mesh',
        '.dxf', '.iges', '.igs', '.step', '.stp', '.dwg', '.skp', '.3dm'
    },
    
    # Archive formats
    'archive': {
        '.zip', '.rar', '.7z', '.tar', '.gz', '.bz2', '.xz', '.tar.gz', '.tar.bz2',
        '.tar.xz', '.tgz', '.tbz2', '.txz', '.Z', '.lz', '.lzma', '.cab', '.msi',
        '.deb', '.rpm', '.dmg', '.iso', '.img', '.bin', '.cue', '.nrg', '.mdf',
        '.cdi', '.bwt', '.b6t', '.pdi'
    },
    
    # Font formats
    'font': {
        '.ttf', '.otf', '.woff', '.woff2', '.eot', '.fon', '.fnt', '.bdf', '.pcf',
        '.snf', '.pfb', '.pfm', '.afm', '.pfa', '.gsf', '.psf', '.dfont'
    },
    
    # Data/Database formats
    'data': {
        '.db', '.sqlite', '.sqlite3', '.mdb', '.accdb', '.dbf', '.dat', '.sav',
        '.por', '.xpt', '.rdata', '.rds', '.mat', '.h5', '.hdf5', '.nc', '.cdf',
        '.fits', '.parquet', '.arrow', '.avro', '.orc', '.feather'
    },
    
    # Configuration formats
    'config': {
        '.ini', '.cfg', '.conf', '.config', '.properties', '.yaml', '.yml', 
        '.toml', '.env', '.plist', '.reg', '.inf', '.desktop', '.service'
    },
    
    # System/Executable formats
    'executable': {
        '.exe', '.msi', '.app', '.deb', '.rpm', '.pkg', '.dmg', '.run', '.bin',
        '.com', '.scr', '.dll', '.so', '.dylib', '.sys', '.drv', '.ocx'
    },
    
    # Virtual/Disk formats
    'virtual': {
        '.vmdk', '.vdi', '.vhd', '.vhdx', '.qcow', '.qcow2', '.vmx', '.ovf', '.ova'
    },
    
    # Log/Temporary formats
    'temp_log': {
        '.log', '.tmp', '.temp', '.bak', '.backup', '.old', '.orig', '.cache',
        '.lock', '.pid', '.swap', '.~'
    }
}

# All supported extensions (flattened)
ALL_SUPPORTED_EXTENSIONS = set()
for category_exts in MULTIMEDIA_FILE_TYPES.values():
    ALL_SUPPORTED_EXTENSIONS.update(category_exts)

# Utility functions
def unique_name(dest_dir: Path, filename: str) -> Path:
    """Generate unique filename to avoid overwriting"""
    dest_dir.mkdir(parents=True, exist_ok=True)
    base = Path(filename).stem
    ext = Path(filename).suffix
    candidate = dest_dir / f"{base}{ext}"
    counter = 1
    while candidate.exists():
        candidate = dest_dir / f"{base}_{counter:02d}{ext}"
        counter += 1
    return candidate

def send_to_trash(path: Path):
    """Move file to macOS trash"""
    trash = Path.home() / ".Trash"
    target = unique_name(trash, path.name)
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(path), str(target))

def get_file_hash(path: Path) -> str:
    """Calculate MD5 hash of file for duplicate detection"""
    try:
        hash_md5 = hashlib.md5()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    except:
        return ""

def detect_file_category(path: Path) -> str:
    """Detect which category a file belongs to"""
    ext = path.suffix.lower()
    
    for category, extensions in MULTIMEDIA_FILE_TYPES.items():
        if ext in extensions:
            return category
    
    return 'other'

MOJIBAKE_PATTERNS = [
    re.compile(r"Ã[\x80-\xBF]"),
    re.compile(r"Â[\x80-\xBF]"),
    re.compile(r"ãƒ"),
    re.compile(r"ã\x83"),
    re.compile(r"ã\x81"),
    re.compile(r"ã\x82"),
    re.compile(r"ã\x84"),
    re.compile(r"[åæç][\x80-\xBF]"),
]

CJK_CHAR_PATTERN = re.compile(r"[\u3040-\u30ff\u3400-\u4dbf\u4e00-\u9fff]")

def detect_name_anomalies(name: str) -> List[str]:
    """Detect common mojibake patterns or invalid characters in file names"""
    reasons: List[str] = []
    
    if not name:
        return reasons
    
    if "�" in name:
        reasons.append("名前に置換文字 (�) が含まれています")
    
    control_chars = [
        ch for ch in name
        if unicodedata.category(ch).startswith("C") and ch not in ("\n", "\r", "\t")
    ]
    if control_chars:
        reasons.append("名前に制御文字が含まれています")
    
    if any(pattern.search(name) for pattern in MOJIBAKE_PATTERNS):
        reasons.append("UTF-8/Shift_JIS の文字化けと見られるパターン (Ã, Â, ã 等)")
    
    high_ascii_chars = [ch for ch in name if 0x80 <= ord(ch) <= 0xFF]
    if high_ascii_chars:
        ratio = len(high_ascii_chars) / max(len(name), 1)
        if ratio > 0.6 and not CJK_CHAR_PATTERN.search(name):
            reasons.append("拡張ASCII文字が多く日本語情報が失われている可能性があります")
    
    if "??" in name:
        reasons.append("名前に ? が連続しておりエンコード失敗の可能性があります")
    
    # Preserve discovery order while removing duplicates
    return list(dict.fromkeys(reasons))

def detect_corruption_indicators(path: Path, info: Dict[str, Any]) -> List[str]:
    """Detect indicators that the file content may be corrupted"""
    reasons: List[str] = []
    
    analysis_error = info.get("analysis_error")
    if analysis_error:
        reasons.append(f"解析エンジンのエラー: {analysis_error}")
    
    archive_error = info.get("error")
    if archive_error:
        reasons.append(f"アーカイブ解析エラー: {archive_error}")
    
    size = info.get("size") or 0
    if size == 0:
        reasons.append("ファイルサイズが0バイトです")
    
    if not info.get("file_hash"):
        reasons.append("ファイルのハッシュを計算できませんでした (読み取り不可の可能性)")
    
    engine = info.get("analysis_engine")
    if engine == "audio" and info.get("duration") is None and size > 0:
        reasons.append("音声の再生時間を取得できませんでした (フォーマット非対応または破損の可能性)")
    
    if engine == "video" and size > 0:
        if info.get("width") is None or info.get("height") is None:
            reasons.append("動画の解像度を取得できませんでした (破損の可能性)")
        if info.get("duration") is None:
            reasons.append("動画の長さを取得できませんでした (破損の可能性)")
    
    if engine == "image" and size > 0:
        if info.get("width") is None or info.get("height") is None:
            reasons.append("画像のピクセルサイズを取得できませんでした (破損の可能性)")
    
    return list(dict.fromkeys(reasons))

# Independent analysis functions - avoid circular imports

def analyze_audio_file_independent(path: Path) -> Dict[str, Any]:
    """Independent audio file analysis using direct library access"""
    info = {"samplerate": None, "channels": None, "duration": None, "bitrate": None, 
            "format": None, "title": None, "artist": None, "album": None, "genre": None, "year": None}
    
    if not AUDIO_MUTAGEN_AVAILABLE:
        return info
    
    try:
        from mutagen import File
        from mutagen.wave import WAVE
        from mutagen.aiff import AIFF
        
        audio_file = File(str(path))
        if audio_file is None:
            return info
        
        # Get basic info
        if hasattr(audio_file, 'info') and audio_file.info:
            info["duration"] = getattr(audio_file.info, 'length', None)
            info["bitrate"] = getattr(audio_file.info, 'bitrate', None)
            
            # Sample rate and channels
            if hasattr(audio_file.info, 'sample_rate'):
                info["samplerate"] = audio_file.info.sample_rate
            elif hasattr(audio_file.info, 'samplerate'):
                info["samplerate"] = audio_file.info.samplerate
                
            if hasattr(audio_file.info, 'channels'):
                info["channels"] = audio_file.info.channels
        
        # Get metadata tags
        if hasattr(audio_file, 'tags') and audio_file.tags:
            tags = audio_file.tags
            
            # Title
            for key in ['TIT2', 'TITLE', '\xa9nam', 'Title']:
                if key in tags:
                    info["title"] = str(tags[key][0]) if isinstance(tags[key], list) else str(tags[key])
                    break
            
            # Artist  
            for key in ['TPE1', 'ARTIST', '\xa9ART', 'Artist']:
                if key in tags:
                    info["artist"] = str(tags[key][0]) if isinstance(tags[key], list) else str(tags[key])
                    break
            
            # Album
            for key in ['TALB', 'ALBUM', '\xa9alb', 'Album']:
                if key in tags:
                    info["album"] = str(tags[key][0]) if isinstance(tags[key], list) else str(tags[key])
                    break
            
            # Genre
            for key in ['TCON', 'GENRE', '\xa9gen', 'Genre']:
                if key in tags:
                    info["genre"] = str(tags[key][0]) if isinstance(tags[key], list) else str(tags[key])
                    break
            
            # Year
            for key in ['TDRC', 'DATE', '\xa9day', 'Date', 'YEAR']:
                if key in tags:
                    year_str = str(tags[key][0]) if isinstance(tags[key], list) else str(tags[key])
                    year_match = re.search(r'\d{4}', year_str)
                    if year_match:
                        info["year"] = int(year_match.group())
                    break
        
        # Format detection
        file_type = type(audio_file).__name__
        info["format"] = file_type.lower()
        
    except Exception as e:
        info["analysis_error"] = str(e)
    
    return info

def analyze_video_file_independent(path: Path) -> Dict[str, Any]:
    """Independent video file analysis using ffprobe"""
    info = {"resolution": None, "width": None, "height": None, "aspect_ratio": None, 
            "fps": None, "duration": None, "bitrate": None, "video_codec": None, 
            "audio_codec": None, "container": None}
    
    if not VIDEO_FFPROBE_AVAILABLE:
        return info
    
    try:
        cmd = ['ffprobe', '-v', 'quiet', '-print_format', 'json', 
               '-show_format', '-show_streams', str(path)]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        
        if result.returncode != 0:
            return info
        
        data = json.loads(result.stdout)
        
        # Format info
        if 'format' in data:
            format_data = data['format']
            info["duration"] = float(format_data.get('duration', 0)) or None
            info["bitrate"] = int(format_data.get('bit_rate', 0)) or None
            info["container"] = format_data.get('format_name', '').split(',')[0]
        
        # Stream info
        if 'streams' in data:
            for stream in data['streams']:
                if stream.get('codec_type') == 'video':
                    info["width"] = stream.get('width')
                    info["height"] = stream.get('height')
                    if info["width"] and info["height"]:
                        info["resolution"] = f"{info['width']}x{info['height']}"
                        info["aspect_ratio"] = round(info["width"] / info["height"], 2)
                    
                    # FPS
                    fps_str = stream.get('r_frame_rate', '0/1')
                    if '/' in fps_str:
                        num, den = fps_str.split('/')
                        if int(den) > 0:
                            info["fps"] = round(int(num) / int(den), 2)
                    
                    info["video_codec"] = stream.get('codec_name')
                
                elif stream.get('codec_type') == 'audio':
                    if not info["audio_codec"]:  # Use first audio stream
                        info["audio_codec"] = stream.get('codec_name')
        
    except Exception as e:
        info["analysis_error"] = str(e)
    
    return info

def analyze_image_file_independent(path: Path) -> Dict[str, Any]:
    """Independent image file analysis using PIL"""
    info = {"width": None, "height": None, "resolution": None, "color_mode": None, 
            "format": None, "dpi": None, "exif_data": {}}
    
    if not IMAGE_PIL_AVAILABLE:
        return info
    
    try:
        from PIL import Image
        from PIL.ExifTags import TAGS
        
        with Image.open(path) as img:
            info["width"], info["height"] = img.size
            info["resolution"] = f"{info['width']}x{info['height']}"
            info["color_mode"] = img.mode
            info["format"] = img.format
            
            # DPI info
            if hasattr(img, 'info') and 'dpi' in img.info:
                info["dpi"] = img.info['dpi']
            
            # EXIF data
            if hasattr(img, '_getexif') and img._getexif():
                exif = img._getexif()
                for tag_id, value in exif.items():
                    tag = TAGS.get(tag_id, tag_id)
                    info["exif_data"][tag] = str(value) if value else None
                    
                    # Extract common EXIF fields
                    if tag == 'DateTime':
                        info["datetime"] = str(value)
                    elif tag == 'Make':
                        info["camera_make"] = str(value)
                    elif tag == 'Model':
                        info["camera_model"] = str(value)
                    elif tag == 'ExposureTime':
                        info["exposure_time"] = str(value)
                    elif tag == 'FNumber':
                        info["f_number"] = str(value)
                    elif tag == 'ISOSpeedRatings':
                        info["iso"] = str(value)
        
    except Exception as e:
        info["analysis_error"] = str(e)
    
    return info

def analyze_document_file_independent(path: Path) -> Dict[str, Any]:
    """Independent document file analysis using various libraries"""
    info = {"pages": None, "words": None, "chars": None, "title": None, 
            "author": None, "subject": None, "creator": None, "producer": None, 
            "creation_date": None, "modification_date": None, "language": None, "encoding": None}
    
    ext = path.suffix.lower()
    
    try:
        if ext == '.pdf' and DOCUMENT_PDF_AVAILABLE:
            import PyPDF2
            
            with open(path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                info["pages"] = len(reader.pages)
                
                # Metadata
                if reader.metadata:
                    meta = reader.metadata
                    info["title"] = meta.get('/Title', '').strip() or None
                    info["author"] = meta.get('/Author', '').strip() or None
                    info["subject"] = meta.get('/Subject', '').strip() or None
                    info["creator"] = meta.get('/Creator', '').strip() or None
                    info["producer"] = meta.get('/Producer', '').strip() or None
                    
                    # Dates
                    if '/CreationDate' in meta:
                        info["creation_date"] = str(meta['/CreationDate'])
                    if '/ModDate' in meta:
                        info["modification_date"] = str(meta['/ModDate'])
                
                # Count text
                text_content = ""
                try:
                    for page in reader.pages:
                        text_content += page.extract_text()
                    info["words"] = len(text_content.split())
                    info["chars"] = len(text_content)
                except:
                    pass
                    
        elif ext in ['.docx', '.doc'] and DOCUMENT_DOCX_AVAILABLE:
            import docx
            
            doc = docx.Document(path)
            
            # Core properties
            props = doc.core_properties
            info["title"] = props.title or None
            info["author"] = props.author or None
            info["subject"] = props.subject or None
            info["creator"] = props.creator or None
            if props.created:
                info["creation_date"] = str(props.created)
            if props.modified:
                info["modification_date"] = str(props.modified)
            info["language"] = props.language or None
            
            # Count paragraphs and text
            paragraphs = doc.paragraphs
            text_content = "\n".join([p.text for p in paragraphs])
            info["words"] = len(text_content.split())
            info["chars"] = len(text_content)
            info["paragraphs"] = len(paragraphs)
                    
        elif ext == '.txt' and DOCUMENT_CHARDET_AVAILABLE:
            import chardet
            
            # Detect encoding
            with open(path, 'rb') as file:
                raw_data = file.read(10000)  # Read first 10KB
                encoding_info = chardet.detect(raw_data)
                info["encoding"] = encoding_info.get('encoding')
                info["encoding_confidence"] = encoding_info.get('confidence')
            
            # Count content
            try:
                with open(path, 'r', encoding=info["encoding"] or 'utf-8') as file:
                    content = file.read()
                    info["words"] = len(content.split())
                    info["chars"] = len(content)
                    info["lines"] = content.count('\n') + 1
            except:
                pass
        
    except Exception as e:
        info["analysis_error"] = str(e)
    
    return info

def analyze_3d_file_independent(path: Path) -> Dict[str, Any]:
    """Independent 3D file analysis using trimesh"""
    info = {"vertices": None, "faces": None, "triangles": None, "volume": None, 
            "surface_area": None, "bounds": None, "is_closed": None, "is_valid": None, 
            "materials": None, "textures": None}
    
    if not THREED_AVAILABLE:
        return info
    
    try:
        import trimesh
        import numpy as np
        
        mesh = trimesh.load(str(path))
        
        if hasattr(mesh, 'vertices'):
            info["vertices"] = len(mesh.vertices)
        if hasattr(mesh, 'faces'):
            info["faces"] = len(mesh.faces)
            info["triangles"] = info["faces"]  # Assuming triangular faces
        
        # Geometric properties
        try:
            if hasattr(mesh, 'volume'):
                info["volume"] = float(mesh.volume)
        except:
            pass
            
        try:
            if hasattr(mesh, 'area'):
                info["surface_area"] = float(mesh.area)
        except:
            pass
        
        # Bounds
        if hasattr(mesh, 'bounds'):
            bounds = mesh.bounds
            info["bounds"] = {
                "min": bounds[0].tolist() if hasattr(bounds[0], 'tolist') else list(bounds[0]),
                "max": bounds[1].tolist() if hasattr(bounds[1], 'tolist') else list(bounds[1])
            }
            
            # Calculate dimensions
            dimensions = bounds[1] - bounds[0]
            info["width"] = float(dimensions[0])
            info["height"] = float(dimensions[1])
            info["depth"] = float(dimensions[2])
        
        # Validation
        if hasattr(mesh, 'is_closed'):
            info["is_closed"] = bool(mesh.is_closed)
        if hasattr(mesh, 'is_valid'):
            info["is_valid"] = bool(mesh.is_valid)
        
        # Materials and textures
        if hasattr(mesh, 'materials') and mesh.materials:
            info["materials"] = len(mesh.materials)
        if hasattr(mesh, 'visual') and hasattr(mesh.visual, 'material'):
            if hasattr(mesh.visual.material, 'image'):
                info["textures"] = 1
        
    except Exception as e:
        info["analysis_error"] = str(e)
    
    return info

# Independent categorization functions

def categorize_audio_independent(info: Dict[str, Any]) -> Dict[str, str]:
    """Independent audio categorization"""
    categories = {}
    
    # Format category
    format_name = info.get("format", "").lower()
    if format_name in ["mp3", "mp3file"]:
        categories["format"] = "format_mp3"
    elif format_name in ["wav", "wave"]:
        categories["format"] = "format_wav"
    elif format_name in ["flac"]:
        categories["format"] = "format_flac"
    elif format_name in ["aac", "m4a"]:
        categories["format"] = "format_aac"
    elif format_name in ["aiff"]:
        categories["format"] = "format_aiff"
    elif format_name in ["ogg", "vorbis"]:
        categories["format"] = "format_ogg"
    else:
        categories["format"] = "format_other"
    
    # Sample rate category
    samplerate = info.get("samplerate")
    if samplerate:
        if samplerate <= 22050:
            categories["samplerate"] = "samplerate_low"
        elif samplerate <= 44100:
            categories["samplerate"] = "samplerate_cd"
        elif samplerate <= 48000:
            categories["samplerate"] = "samplerate_standard"
        elif samplerate <= 96000:
            categories["samplerate"] = "samplerate_high"
        else:
            categories["samplerate"] = "samplerate_very_high"
    else:
        categories["samplerate"] = "samplerate_unknown"
    
    # Channel category
    channels = info.get("channels")
    if channels == 1:
        categories["channels"] = "channels_mono"
    elif channels == 2:
        categories["channels"] = "channels_stereo"
    elif channels and channels > 2:
        categories["channels"] = f"channels_multichannel_{channels}"
    else:
        categories["channels"] = "channels_unknown"
    
    # Duration category
    duration = info.get("duration")
    if duration:
        if duration < 30:
            categories["duration"] = "duration_very_short"
        elif duration < 180:
            categories["duration"] = "duration_short"
        elif duration < 600:
            categories["duration"] = "duration_medium"
        elif duration < 1800:
            categories["duration"] = "duration_long"
        else:
            categories["duration"] = "duration_very_long"
    else:
        categories["duration"] = "duration_unknown"
    
    return categories

def categorize_video_independent(info: Dict[str, Any]) -> Dict[str, str]:
    """Independent video categorization"""
    categories = {}
    
    # Resolution category
    height = info.get("height")
    if height:
        if height <= 480:
            categories["resolution"] = "resolution_sd"
        elif height <= 720:
            categories["resolution"] = "resolution_hd"
        elif height <= 1080:
            categories["resolution"] = "resolution_fullhd"
        elif height <= 1440:
            categories["resolution"] = "resolution_qhd"
        elif height <= 2160:
            categories["resolution"] = "resolution_4k"
        else:
            categories["resolution"] = "resolution_8k_plus"
    else:
        categories["resolution"] = "resolution_unknown"
    
    # Aspect ratio category
    aspect_ratio = info.get("aspect_ratio")
    if aspect_ratio:
        if abs(aspect_ratio - 1.33) < 0.1:
            categories["aspect"] = "aspect_4_3"
        elif abs(aspect_ratio - 1.78) < 0.1:
            categories["aspect"] = "aspect_16_9"
        elif abs(aspect_ratio - 2.35) < 0.1:
            categories["aspect"] = "aspect_cinema"
        elif aspect_ratio < 1.2:
            categories["aspect"] = "aspect_square"
        else:
            categories["aspect"] = "aspect_other"
    else:
        categories["aspect"] = "aspect_unknown"
    
    # FPS category
    fps = info.get("fps")
    if fps:
        if fps <= 24:
            categories["fps"] = "fps_cinema"
        elif fps <= 30:
            categories["fps"] = "fps_standard"
        elif fps <= 60:
            categories["fps"] = "fps_smooth"
        else:
            categories["fps"] = "fps_high_speed"
    else:
        categories["fps"] = "fps_unknown"
    
    # Container category
    container = info.get("container", "").lower()
    if container in ["mp4"]:
        categories["container"] = "container_mp4"
    elif container in ["avi"]:
        categories["container"] = "container_avi"
    elif container in ["mkv", "matroska"]:
        categories["container"] = "container_mkv"
    elif container in ["mov"]:
        categories["container"] = "container_mov"
    else:
        categories["container"] = "container_other"
    
    return categories

def categorize_image_independent(info: Dict[str, Any]) -> Dict[str, str]:
    """Independent image categorization"""
    categories = {}
    
    # Format category
    format_name = info.get("format", "").upper()
    if format_name == "JPEG":
        categories["format"] = "format_jpeg"
    elif format_name == "PNG":
        categories["format"] = "format_png"
    elif format_name == "GIF":
        categories["format"] = "format_gif"
    elif format_name == "BMP":
        categories["format"] = "format_bmp"
    elif format_name == "TIFF":
        categories["format"] = "format_tiff"
    elif format_name == "WEBP":
        categories["format"] = "format_webp"
    else:
        categories["format"] = "format_other"
    
    # Size category
    width = info.get("width", 0)
    height = info.get("height", 0)
    if width and height:
        pixels = width * height
        if pixels < 100000:  # <100K pixels
            categories["size"] = "size_thumbnail"
        elif pixels < 1000000:  # <1M pixels
            categories["size"] = "size_small"
        elif pixels < 5000000:  # <5M pixels  
            categories["size"] = "size_medium"
        elif pixels < 20000000:  # <20M pixels
            categories["size"] = "size_large"
        else:
            categories["size"] = "size_huge"
    else:
        categories["size"] = "size_unknown"
    
    # Color mode category
    color_mode = info.get("color_mode", "")
    if color_mode == "L":
        categories["color"] = "color_grayscale"
    elif color_mode == "RGB":
        categories["color"] = "color_rgb"
    elif color_mode == "RGBA":
        categories["color"] = "color_rgba"
    elif color_mode == "CMYK":
        categories["color"] = "color_cmyk"
    else:
        categories["color"] = "color_other"
    
    return categories

def categorize_document_independent(info: Dict[str, Any]) -> Dict[str, str]:
    """Independent document categorization"""
    categories = {}
    
    # Pages category
    pages = info.get("pages")
    if pages:
        if pages == 1:
            categories["pages"] = "pages_single"
        elif pages <= 10:
            categories["pages"] = "pages_short"
        elif pages <= 50:
            categories["pages"] = "pages_medium"
        elif pages <= 200:
            categories["pages"] = "pages_long"
        else:
            categories["pages"] = "pages_very_long"
    else:
        categories["pages"] = "pages_unknown"
    
    # Words category
    words = info.get("words")
    if words:
        if words < 500:
            categories["length"] = "length_short"
        elif words < 2000:
            categories["length"] = "length_medium"
        elif words < 10000:
            categories["length"] = "length_long"
        else:
            categories["length"] = "length_very_long"
    else:
        categories["length"] = "length_unknown"
    
    # Has metadata category
    has_author = bool(info.get("author"))
    has_title = bool(info.get("title"))
    if has_author and has_title:
        categories["metadata"] = "metadata_rich"
    elif has_author or has_title:
        categories["metadata"] = "metadata_partial"
    else:
        categories["metadata"] = "metadata_minimal"
    
    return categories

def categorize_3d_independent(info: Dict[str, Any]) -> Dict[str, str]:
    """Independent 3D categorization"""
    categories = {}
    
    # Complexity category based on vertices
    vertices = info.get("vertices", 0)
    if vertices:
        if vertices < 1000:
            categories["complexity"] = "complexity_low"
        elif vertices < 10000:
            categories["complexity"] = "complexity_medium"
        elif vertices < 100000:
            categories["complexity"] = "complexity_high"
        else:
            categories["complexity"] = "complexity_very_high"
    else:
        categories["complexity"] = "complexity_unknown"
    
    # Quality category
    is_closed = info.get("is_closed")
    is_valid = info.get("is_valid")
    if is_closed and is_valid:
        categories["quality"] = "quality_good"
    elif is_closed or is_valid:
        categories["quality"] = "quality_fair"
    else:
        categories["quality"] = "quality_poor"
    
    # Size category based on volume
    volume = info.get("volume")
    if volume:
        if volume < 1:
            categories["volume"] = "volume_small"
        elif volume < 1000:
            categories["volume"] = "volume_medium"
        else:
            categories["volume"] = "volume_large"
    else:
        categories["volume"] = "volume_unknown"
    
    return categories

def analyze_archive_file(path: Path) -> Dict[str, Any]:
    """Analyze archive files (ZIP, RAR, etc.)"""
    info = {
        "archive_type": None,
        "compressed_size": 0,
        "uncompressed_size": 0,
        "file_count": 0,
        "compression_ratio": None,
        "has_encryption": False,
        "comment": None,
        "files": []
    }
    
    ext = path.suffix.lower()
    
    try:
        if ext == '.zip':
            with zipfile.ZipFile(path, 'r') as zf:
                info["archive_type"] = "ZIP"
                info["file_count"] = len(zf.namelist())
                info["comment"] = zf.comment.decode('utf-8', errors='ignore') if zf.comment else None
                
                total_uncompressed = 0
                total_compressed = 0
                
                for file_info_zip in zf.filelist:
                    total_uncompressed += file_info_zip.file_size
                    total_compressed += file_info_zip.compress_size
                    
                    if len(info["files"]) < 10:  # Limit for performance
                        info["files"].append({
                            "name": file_info_zip.filename,
                            "size": file_info_zip.file_size,
                            "compressed_size": file_info_zip.compress_size,
                            "modified": datetime(*file_info_zip.date_time) if file_info_zip.date_time else None
                        })
                
                info["uncompressed_size"] = total_uncompressed
                info["compressed_size"] = total_compressed
                
                if total_uncompressed > 0:
                    info["compression_ratio"] = round((1 - total_compressed / total_uncompressed) * 100, 1)
                    
        elif ext in ['.tar', '.tar.gz', '.tgz', '.tar.bz2', '.tar.xz']:
            with tarfile.open(path, 'r') as tf:
                info["archive_type"] = "TAR"
                members = tf.getmembers()
                info["file_count"] = len(members)
                
                total_size = 0
                for member in members:
                    if member.isfile():
                        total_size += member.size
                        
                        if len(info["files"]) < 10:
                            info["files"].append({
                                "name": member.name,
                                "size": member.size,
                                "modified": datetime.fromtimestamp(member.mtime) if member.mtime else None
                            })
                
                info["uncompressed_size"] = total_size
                info["compressed_size"] = path.stat().st_size
                
                if total_size > 0:
                    info["compression_ratio"] = round((1 - path.stat().st_size / total_size) * 100, 1)
                    
        elif ext == '.rar' and RARFILE_AVAILABLE:
            with rarfile.RarFile(path, 'r') as rf:
                info["archive_type"] = "RAR"
                members = rf.namelist()
                info["file_count"] = len(members)
                info["comment"] = rf.comment
                
                # RAR analysis is more limited due to library constraints
                info["has_encryption"] = rf.needs_password()
                
    except Exception as e:
        info["error"] = str(e)
    
    return info

def analyze_font_file(path: Path) -> Dict[str, Any]:
    """Analyze font files"""
    info = {
        "font_type": None,
        "font_family": None,
        "font_style": None,
        "version": None,
        "glyph_count": None,
        "is_monospace": None
    }
    
    ext = path.suffix.lower()
    
    try:
        if ext in ['.ttf', '.otf']:
            # Basic font analysis (simplified)
            with open(path, 'rb') as f:
                # Read font header (very basic)
                header = f.read(12)
                if len(header) >= 12:
                    # This is a very simplified analysis
                    info["font_type"] = "OpenType" if ext == '.otf' else "TrueType"
                    
                    # Try to read some basic info
                    f.seek(0)
                    font_data = f.read(1024)  # Read first 1KB
                    
                    # Look for common font name patterns (very basic)
                    if b'Regular' in font_data:
                        info["font_style"] = "Regular"
                    elif b'Bold' in font_data:
                        info["font_style"] = "Bold"
                    elif b'Italic' in font_data:
                        info["font_style"] = "Italic"
                        
        elif ext in ['.woff', '.woff2']:
            info["font_type"] = "Web Font"
            
    except Exception as e:
        info["error"] = str(e)
    
    return info

def multimedia_probe(path: Path) -> Dict[str, Any]:
    """Comprehensive file analysis using all available analyzers"""
    base_info = {
        "path": str(path),
        "name": path.name,
        "ext": path.suffix.lower(),
        "size": 0,
        "mtime": None,
        "file_hash": None,
        "category": "other",
        "mime_type": None,
        "analysis_engine": "basic",
        "anomaly_reasons": [],
        "has_name_anomaly": False,
        "is_corruption_suspected": False,
        "has_anomalies": False
    }
    
    try:
        stat = path.stat()
        base_info["size"] = stat.st_size
        base_info["mtime"] = stat.st_mtime
    except:
        pass
    
    # Calculate file hash for duplicate detection
    base_info["file_hash"] = get_file_hash(path)

    # Detect anomalous name patterns before deeper analysis
    name_anomalies = detect_name_anomalies(base_info["name"])
    if name_anomalies:
        base_info["has_name_anomaly"] = True
        base_info["anomaly_reasons"].extend(name_anomalies)
    
    # Detect MIME type
    try:
        if PYTHON_MAGIC_AVAILABLE:
            base_info["mime_type"] = magic.from_file(str(path), mime=True)
        else:
            base_info["mime_type"] = mimetypes.guess_type(str(path))[0]
    except:
        pass
    
    # Detect file category
    category = detect_file_category(path)
    base_info["category"] = category
    
    # Use specialized analyzers based on category - independent implementations
    try:
        if category == 'audio':
            audio_info = analyze_audio_file_independent(path)
            base_info.update(audio_info)
            base_info["analysis_engine"] = "audio"
            
        elif category == 'video':
            video_info = analyze_video_file_independent(path)
            base_info.update(video_info)
            base_info["analysis_engine"] = "video"
            
        elif category == 'image':
            image_info = analyze_image_file_independent(path)
            base_info.update(image_info)
            base_info["analysis_engine"] = "image"
            
        elif category == 'document':
            document_info = analyze_document_file_independent(path)
            base_info.update(document_info)
            base_info["analysis_engine"] = "document"
            
        elif category == '3d':
            threed_info = analyze_3d_file_independent(path)
            base_info.update(threed_info)
            base_info["analysis_engine"] = "3d"
            
        elif category == 'archive':
            archive_info = analyze_archive_file(path)
            base_info.update(archive_info)
            base_info["analysis_engine"] = "archive"
            
        elif category == 'font':
            font_info = analyze_font_file(path)
            base_info.update(font_info)
            base_info["analysis_engine"] = "font"
            
    except Exception as e:
        base_info["analysis_error"] = str(e)
    
    # Aggregate corruption indicators and finalize anomaly status
    corruption_reasons = detect_corruption_indicators(path, base_info)
    if corruption_reasons:
        base_info["is_corruption_suspected"] = True
        base_info["anomaly_reasons"].extend(corruption_reasons)
    
    if base_info["anomaly_reasons"]:
        base_info["anomaly_reasons"] = list(dict.fromkeys(base_info["anomaly_reasons"]))
        base_info["has_anomalies"] = True
    
    return base_info

def categorize_multimedia_file(info: Dict[str, Any]) -> Dict[str, str]:
    """Unified categorization system for all file types"""
    categories = {}
    
    # Primary category (file type)
    category = info.get("category", "other")
    categories["primary"] = f"primary_{category}"
    
    # Use specialized categorization - independent implementations
    try:
        if category == 'audio':
            audio_categories = categorize_audio_independent(info)
            categories.update({f"audio_{k}": v for k, v in audio_categories.items()})
            
        elif category == 'video':
            video_categories = categorize_video_independent(info)
            categories.update({f"video_{k}": v for k, v in video_categories.items()})
            
        elif category == 'image':
            image_categories = categorize_image_independent(info)
            categories.update({f"image_{k}": v for k, v in image_categories.items()})
            
        elif category == 'document':
            document_categories = categorize_document_independent(info)
            categories.update({f"document_{k}": v for k, v in document_categories.items()})
            
        elif category == '3d':
            threed_categories = categorize_3d_independent(info)
            categories.update({f"3d_{k}": v for k, v in threed_categories.items()})
            
    except Exception as e:
        pass
    
    # Universal categories that apply to all files
    
    # File size category
    size = info.get("size", 0)
    if size:
        size_mb = size / (1024 * 1024)
        if size_mb < 0.1:
            categories["size"] = "size_tiny"
        elif size_mb < 1:
            categories["size"] = "size_very_small"
        elif size_mb < 10:
            categories["size"] = "size_small"
        elif size_mb < 100:
            categories["size"] = "size_medium"
        elif size_mb < 1024:
            categories["size"] = "size_large"
        elif size_mb < 5120:  # 5GB
            categories["size"] = "size_very_large"
        else:
            categories["size"] = "size_huge"
    else:
        categories["size"] = "size_unknown"
    
    # Extension category (grouped)
    ext = info.get("ext", "").lower()
    if ext.startswith('.'):
        ext = ext[1:]  # Remove dot
        
    if len(ext) <= 3:
        categories["ext_length"] = "ext_short"
    elif len(ext) <= 6:
        categories["ext_length"] = "ext_medium"
    else:
        categories["ext_length"] = "ext_long"
    
    # Date category
    mtime = info.get("mtime")
    if mtime:
        date = datetime.fromtimestamp(mtime)
        categories["date"] = f"{date.year}-{date.month:02d}"
        
        # Age category
        age_days = (datetime.now() - date).days
        if age_days < 1:
            categories["age"] = "age_today"
        elif age_days < 7:
            categories["age"] = "age_week"
        elif age_days < 30:
            categories["age"] = "age_month"
        elif age_days < 365:
            categories["age"] = "age_year"
        else:
            categories["age"] = "age_old"
    else:
        categories["date"] = "date_unknown"
        categories["age"] = "age_unknown"
    
    # Analysis engine category
    engine = info.get("analysis_engine", "basic")
    categories["engine"] = f"engine_{engine}"
    
    return categories


class MultimediaAnalysisThread(QThread):
    """Multimedia analysis thread for comprehensive file processing"""
    
    progress_updated = Signal(str, int, int)  # message, current, total
    analysis_completed = Signal(dict)         # analysis results
    error_occurred = Signal(str)              # error message
    
    def __init__(self, paths: List[Path]):
        super().__init__()
        self.paths = paths if isinstance(paths, list) else [paths]
    
    def run(self):
        """Analyze all files in the given paths"""
        try:
            results = {}
            total_files = 0
            processed = 0
            suspect_files: List[Dict[str, Any]] = []
            
            # Count total files
            all_files = []
            for root_path in self.paths:
                if root_path.is_dir():
                    for file_path in root_path.rglob("*"):
                        if file_path.is_file():
                            all_files.append(file_path)
            
            total_files = len(all_files)
            if total_files == 0:
                self.analysis_completed.emit({})
                return
            
            # Process each file
            for file_path in all_files:
                self.progress_updated.emit(f"解析中: {file_path.name}", processed + 1, total_files)
                
                try:
                    # Get comprehensive file info
                    file_info = multimedia_probe(file_path)
                    categories = categorize_multimedia_file(file_info)
                    
                    if file_info.get("has_anomalies"):
                        suspect_files.append(file_info)
                    
                    # Organize by categories
                    for category_type, category_value in categories.items():
                        if category_type not in results:
                            results[category_type] = {}
                        
                        if category_value not in results[category_type]:
                            results[category_type][category_value] = {
                                "count": 0,
                                "total_size": 0,
                                "files": []
                            }
                        
                        category_data = results[category_type][category_value]
                        category_data["count"] += 1
                        category_data["total_size"] += file_info.get("size", 0)
                        category_data["files"].append(file_info)
                
                except Exception as e:
                    fallback_info = {
                        "path": str(file_path),
                        "name": file_path.name,
                        "ext": file_path.suffix.lower(),
                        "size": 0,
                        "mtime": None,
                        "file_hash": "",
                        "category": "other",
                        "analysis_engine": "basic",
                        "anomaly_reasons": [f"内部解析エラー: {e}"],
                        "analysis_error": str(e),
                        "has_anomalies": True,
                        "has_name_anomaly": False,
                        "is_corruption_suspected": True,
                        "mime_type": None
                    }
                    try:
                        stat = file_path.stat()
                        fallback_info["size"] = stat.st_size
                        fallback_info["mtime"] = stat.st_mtime
                    except OSError:
                        pass
                    suspect_files.append(fallback_info)
                    continue  # Skip further processing for this file
                
                finally:
                    processed += 1
            
            results["__suspect_files__"] = suspect_files
            
            self.analysis_completed.emit(results)
            
        except Exception as e:
            self.error_occurred.emit(str(e))


class MultimediaAnalyzerWindow(QMainWindow):
    """Ultimate multimedia analyzer with comprehensive analysis and processing capabilities"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("🎬 Multimedia解析・整理ツール【集大成版】")
        self.setGeometry(100, 100, 1600, 1000)
        self.setMinimumSize(1400, 900)
        
        # Data management
        self.selected_paths: List[Path] = []
        self.analysis_results: Dict[str, Any] = {}
        self.analysis_thread: Optional[MultimediaAnalysisThread] = None
        self.suspect_files: List[Dict[str, Any]] = []
        self.folder_placeholder_text = "ここにフォルダをドラッグ&ドロップ"

        # Check library availability and show detailed status
        self.check_library_dependencies()
        
        self.init_ui()
        self.apply_pro_theme()
        self.setAcceptDrops(True)
    
    def check_library_dependencies(self):
        """Check library dependencies and show detailed status"""
        missing_critical_libs = []
        
        # Check availability of analysis engines
        analysis_engines = {
            'Audio Analysis': AUDIO_AVAILABLE,
            'Video Analysis': VIDEO_AVAILABLE, 
            'Image Analysis': IMAGE_AVAILABLE,
            'Document Analysis': DOC_AVAILABLE,
            '3D Analysis': MODEL3D_AVAILABLE
        }
        
        missing_engines = [name for name, available in analysis_engines.items() if not available]
        
        if missing_engines:
            self.show_engine_status_dialog(analysis_engines, missing_engines)
    
    def show_engine_status_dialog(self, engines: Dict[str, bool], missing: List[str]):
        """Show analysis engine status dialog"""
        dialog = QDialog(self)
        dialog.setWindowTitle("🎬 Multimedia解析エンジンの状況")
        dialog.setMinimumSize(700, 500)
        dialog.setModal(True)
        
        layout = QVBoxLayout(dialog)
        
        # Header
        header = QLabel("🎬 Multimedia解析エンジンの統合状況")
        header.setStyleSheet("font-size: 18px; font-weight: bold; color: #007acc; margin-bottom: 15px;")
        layout.addWidget(header)
        
        if missing:
            warning_label = QLabel(
                f"⚠️ {len(missing)}個の解析エンジンが利用できません。\n"
                f"これらのファイル形式では詳細解析が制限されます。"
            )
            warning_label.setStyleSheet("color: #d9534f; font-weight: bold; margin-bottom: 10px;")
            warning_label.setWordWrap(True)
            layout.addWidget(warning_label)
        
        # Engine status
        status_group = QGroupBox("🔧 解析エンジン状況")
        status_layout = QVBoxLayout(status_group)
        
        for engine_name, available in engines.items():
            status_text = f"• {engine_name}: "
            if available:
                status_text += "✅ 利用可能"
                status_color = "color: green;"
            else:
                status_text += "❌ 利用不可"
                status_color = "color: red;"
            
            status_label = QLabel(status_text)
            status_label.setStyleSheet(status_color)
            status_layout.addWidget(status_label)
        
        layout.addWidget(status_group)
        
        # Capabilities
        capabilities_group = QGroupBox("🎯 利用可能な機能")
        capabilities_layout = QVBoxLayout(capabilities_group)
        
        capabilities = [
            "✅ 全ファイル形式の基本解析（サイズ・日付・拡張子等）",
            "✅ ファイル種別の自動検出・分類",
            "✅ 重複ファイル検出（ハッシュベース）",
            "✅ 包括的なファイル整理・フラット化",
            "✅ アーカイブファイルの基本解析",
            "✅ フォント形式の基本判定"
        ]
        
        if AUDIO_AVAILABLE:
            capabilities.append("🎵 音声ファイルの詳細解析（メタデータ・品質等）")
        if VIDEO_AVAILABLE:
            capabilities.append("🎥 動画ファイルの詳細解析（解像度・コーデック等）")
        if IMAGE_AVAILABLE:
            capabilities.append("🖼️ 画像ファイルの詳細解析（EXIF・色情報等）")
        if DOC_AVAILABLE:
            capabilities.append("📄 文書ファイルの詳細解析（メタデータ・内容等）")
        if MODEL3D_AVAILABLE:
            capabilities.append("🎮 3Dモデルの詳細解析（メッシュ・材質等）")
        
        for capability in capabilities:
            cap_label = QLabel(capability)
            cap_label.setWordWrap(True)
            capabilities_layout.addWidget(cap_label)
        
        layout.addWidget(capabilities_group)
        
        # Library info
        lib_group = QGroupBox("📚 依存ライブラリ状況")
        lib_layout = QVBoxLayout(lib_group)
        
        critical_libs = ['mutagen', 'PyPDF2', 'python-docx', 'trimesh', 'numpy']
        available_count = sum(1 for lib in critical_libs if COMBINED_LIBRARY_STATUS.get(lib, {}).get('available', False))
        
        lib_summary = QLabel(f"重要ライブラリ: {available_count}/{len(critical_libs)} 利用可能")
        lib_summary.setStyleSheet("font-weight: bold;")
        lib_layout.addWidget(lib_summary)
        
        for lib_name in critical_libs:
            if lib_name in COMBINED_LIBRARY_STATUS:
                lib_info = COMBINED_LIBRARY_STATUS[lib_name]
                if lib_info['available']:
                    version = lib_info.get('version', 'unknown')
                    text = f"✅ {lib_name} (v{version})"
                    color = "color: green;"
                else:
                    text = f"❌ {lib_name}"
                    color = "color: red;"
                
                lib_label = QLabel(text)
                lib_label.setStyleSheet(color)
                lib_layout.addWidget(lib_label)
        
        layout.addWidget(lib_group)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        continue_btn = QPushButton("継続（現在の機能で使用）")
        continue_btn.clicked.connect(dialog.accept)
        continue_btn.setStyleSheet("background-color: #007acc; color: white; padding: 10px 20px; font-weight: bold;")
        button_layout.addWidget(continue_btn)
        
        layout.addLayout(button_layout)
        
        dialog.exec()
    
    def show_library_status_dialog(self):
        """Show current library status (for menu/toolbar access)"""
        engines = {
            'Audio Analysis': AUDIO_AVAILABLE,
            'Video Analysis': VIDEO_AVAILABLE, 
            'Image Analysis': IMAGE_AVAILABLE,
            'Document Analysis': DOC_AVAILABLE,
            '3D Analysis': MODEL3D_AVAILABLE
        }
        missing = [name for name, available in engines.items() if not available]
        self.show_engine_status_dialog(engines, missing)
    
    def init_ui(self):
        """Initialize the UI layout similar to other analyzers"""
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(5, 5, 5, 5)
        main_layout.setSpacing(5)
        
        # Main splitter (vertical)
        vsplitter = QSplitter(Qt.Vertical)
        
        # Top: File/folder tree
        folder_widget = self.create_folder_tree_widget()
        vsplitter.addWidget(folder_widget)
        
        # Middle: Toolbar and analysis results
        bottom_widget = QWidget()
        bottom_layout = QVBoxLayout(bottom_widget)
        bottom_layout.setContentsMargins(0, 0, 0, 0)
        bottom_layout.setSpacing(2)
        
        # Toolbar
        toolbar = self.create_toolbar()
        bottom_layout.addWidget(toolbar)
        
        # Analysis results and processing options in horizontal splitter
        hsplitter = QSplitter(Qt.Horizontal)
        
        # Left: Analysis results
        result_widget = self.create_result_widget()
        hsplitter.addWidget(result_widget)
        
        # Right: Processing options
        options_widget = self.create_options_widget()
        hsplitter.addWidget(options_widget)
        
        hsplitter.setSizes([900, 500])
        bottom_layout.addWidget(hsplitter)
        
        vsplitter.addWidget(bottom_widget)
        vsplitter.setSizes([300, 700])
        
        main_layout.addWidget(vsplitter)
        
        # Status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("フォルダを追加してマルチメディア解析を開始してください")
    
    def create_folder_tree_widget(self):
        """Create folder tree widget for all file types"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # Header
        header_layout = QHBoxLayout()
        header_label = QLabel("📁 解析対象フォルダ")
        header_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        header_layout.addWidget(header_label)
        
        # Statistics
        self.stats_label = QLabel("")
        self.stats_label.setStyleSheet("color: #666; font-size: 12px;")
        header_layout.addWidget(self.stats_label)
        
        header_layout.addStretch()
        layout.addLayout(header_layout)
        
        # Tree view
        self.folder_tree = QTreeWidget()
        self.folder_tree.setHeaderHidden(True)
        self.folder_tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.folder_tree.setAcceptDrops(True)
        self.folder_tree.setMinimumHeight(200)
        
        # Placeholder
        self._add_placeholder_if_empty()

        layout.addWidget(self.folder_tree)

        return widget

    def _add_placeholder_if_empty(self):
        """Ensure placeholder guidance item is present when tree is empty."""
        if self.folder_tree.topLevelItemCount() == 0:
            placeholder = QTreeWidgetItem(self.folder_tree)
            placeholder.setText(0, self.folder_placeholder_text)
            placeholder.setFlags(Qt.NoItemFlags)
            placeholder.setForeground(0, QBrush(QColor("#666666")))
    
    def create_toolbar(self):
        """Create toolbar with multimedia-specific options"""
        toolbar = QWidget()
        toolbar.setMaximumHeight(45)
        layout = QHBoxLayout(toolbar)
        layout.setContentsMargins(8, 3, 8, 3)
        layout.setSpacing(8)
        
        # Folder selection
        add_btn = QPushButton("📁 フォルダ追加")
        add_btn.clicked.connect(self.select_folders)
        add_btn.setStyleSheet("padding: 6px 12px;")
        layout.addWidget(add_btn)
        
        # Remove selected
        remove_btn = QPushButton("🗑️ 選択削除")
        remove_btn.clicked.connect(self.remove_selected_folders)
        layout.addWidget(remove_btn)

        name_remove_btn = QPushButton("🔍 名前で削除")
        name_remove_btn.clicked.connect(self.remove_folders_by_name)
        layout.addWidget(name_remove_btn)

        # Analysis
        analyze_btn = QPushButton("🚀 マルチメディア解析実行")
        analyze_btn.setStyleSheet("background-color: #28a745; color: white; font-weight: bold; padding: 8px 16px;")
        analyze_btn.clicked.connect(self.run_multimedia_analysis)
        layout.addWidget(analyze_btn)
        
        self.suspect_review_btn = QPushButton("⚠️ 異常ファイル確認")
        self.suspect_review_btn.setEnabled(False)
        self.suspect_review_btn.setToolTip("文字化けや破損が疑われるファイル一覧を表示します")
        self.suspect_review_btn.clicked.connect(lambda: self.show_suspect_review_dialog())
        layout.addWidget(self.suspect_review_btn)
        
        # Separator
        separator = QLabel("|")
        separator.setStyleSheet("color: #ccc; font-size: 16px;")
        layout.addWidget(separator)
        
        # Processing mode
        layout.addWidget(QLabel("処理モード:"))
        self.processing_mode = QComboBox()
        self.processing_mode.addItems(["統合整理", "フラット化", "カテゴリ別整理"])
        layout.addWidget(self.processing_mode)
        
        # Dry run
        self.dry_run_check = QCheckBox("シミュレーション")
        self.dry_run_check.setChecked(True)
        layout.addWidget(self.dry_run_check)
        
        layout.addStretch()
        
        # Library status button
        lib_status_btn = QPushButton("📊 エンジン状況")
        lib_status_btn.setStyleSheet("color: #007acc; padding: 6px 12px;")
        lib_status_btn.clicked.connect(self.show_library_status_dialog)
        lib_status_btn.setToolTip("解析エンジンの状況を確認")
        layout.addWidget(lib_status_btn)
        
        # Clear all
        clear_btn = QPushButton("🧹 全クリア")
        clear_btn.setStyleSheet("color: #dc3545; padding: 6px 12px;")
        clear_btn.clicked.connect(self.clear_all)
        layout.addWidget(clear_btn)
        
        return toolbar
    
    def create_result_widget(self):
        """Create analysis results widget"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # Header
        header = QLabel("📊 マルチメディア解析結果")
        header.setStyleSheet("font-weight: bold; font-size: 16px; margin-bottom: 10px;")
        layout.addWidget(header)
        
        # Category tabs
        self.result_tabs = QTabWidget()
        
        # Create tabs for different analysis categories
        self.create_analysis_tabs()
        
        layout.addWidget(self.result_tabs)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        return widget
    
    def create_analysis_tabs(self):
        """Create tabs for different multimedia analysis categories"""
        # Primary categories based on available analysis engines
        categories = [
            ("📂 ファイル種別", "primary"),
            ("📏 ファイルサイズ", "size"),
            ("📅 日付・年齢", "date"),
            ("🏷️ 拡張子", "ext_length"),
            ("⚙️ 解析エンジン", "engine"),
            ("🎵 音声", "audio_format") if AUDIO_AVAILABLE else None,
            ("🎥 動画", "video_format") if VIDEO_AVAILABLE else None,
            ("🖼️ 画像", "image_format") if IMAGE_AVAILABLE else None,
            ("📄 文書", "document_format") if DOC_AVAILABLE else None,
            ("🎮 3D", "3d_format") if MODEL3D_AVAILABLE else None
        ]
        
        # Filter out None values
        categories = [cat for cat in categories if cat is not None]
        
        self.category_trees = {}
        
        for tab_name, category_key in categories:
            tab_widget = QWidget()
            tab_layout = QVBoxLayout(tab_widget)
            
            tree = QTreeWidget()
            tree.setHeaderLabels(["カテゴリ", "ファイル数", "合計サイズ", "平均サイズ"])
            tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
            tree.setAlternatingRowColors(True)
            tree.setSortingEnabled(True)
            
            tab_layout.addWidget(tree)
            self.result_tabs.addTab(tab_widget, tab_name)
            self.category_trees[category_key] = tree
    
    def create_options_widget(self):
        """Create processing options widget"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # Processing options
        options_group = QGroupBox("🎛️ 処理オプション")
        options_layout = QVBoxLayout(options_group)
        
        # Category selection for processing
        category_group = QGroupBox("📂 処理対象カテゴリ")
        category_layout = QVBoxLayout(category_group)
        
        self.category_checks = {}
        categories = [
            ("audio", "🎵 音声ファイル", AUDIO_AVAILABLE),
            ("video", "🎥 動画ファイル", VIDEO_AVAILABLE),
            ("image", "🖼️ 画像ファイル", IMAGE_AVAILABLE),
            ("document", "📄 文書ファイル", DOC_AVAILABLE),
            ("3d", "🎮 3Dモデル", MODEL3D_AVAILABLE),
            ("code", "💻 コード/スクリプト", True),
            ("archive", "📦 アーカイブ", True),
            ("font", "🔤 フォント", True),
            ("data", "💾 データベース", True),
            ("other", "📄 その他", True)
        ]
        
        for cat_key, cat_label, available in categories:
            check = QCheckBox(cat_label)
            check.setChecked(available)
            check.setEnabled(available)
            if not available:
                check.setToolTip("この形式の詳細解析エンジンが利用できません")
            self.category_checks[cat_key] = check
            category_layout.addWidget(check)
        
        options_layout.addWidget(category_group)
        
        # Sorting criteria
        sort_group = QGroupBox("📊 整理基準")
        sort_layout = QVBoxLayout(sort_group)
        
        self.sort_criterion = QComboBox()
        self.sort_criterion.addItems([
            "ファイル種別",
            "ファイルサイズ", 
            "日付・年齢",
            "拡張子",
            "解析エンジン別",
            "音声品質別" if AUDIO_AVAILABLE else None,
            "動画解像度別" if VIDEO_AVAILABLE else None,
            "画像解像度別" if IMAGE_AVAILABLE else None,
            "文書種別" if DOC_AVAILABLE else None,
            "3D複雑度別" if MODEL3D_AVAILABLE else None
        ])
        # Remove None items
        for i in range(self.sort_criterion.count() - 1, -1, -1):
            if self.sort_criterion.itemText(i) is None:
                self.sort_criterion.removeItem(i)
        
        sort_layout.addWidget(self.sort_criterion)
        
        # Advanced options
        advanced_group = QGroupBox("🔧 高度なオプション")
        advanced_layout = QVBoxLayout(advanced_group)
        
        self.duplicate_check = QCheckBox("🔍 重複ファイルを検出・統合")
        advanced_layout.addWidget(self.duplicate_check)
        
        self.size_filter_check = QCheckBox("📏 サイズフィルター有効")
        advanced_layout.addWidget(self.size_filter_check)
        
        self.size_filter_layout = QHBoxLayout()
        self.size_filter_layout.addWidget(QLabel("最小サイズ (MB):"))
        self.size_min = QLineEdit("0")
        self.size_min.setEnabled(False)
        self.size_filter_layout.addWidget(self.size_min)
        
        self.size_filter_layout.addWidget(QLabel("最大サイズ (MB):"))
        self.size_max = QLineEdit("1000")
        self.size_max.setEnabled(False)
        self.size_filter_layout.addWidget(self.size_max)
        
        def toggle_size_filter(enabled):
            self.size_min.setEnabled(enabled)
            self.size_max.setEnabled(enabled)
        
        self.size_filter_check.toggled.connect(toggle_size_filter)
        advanced_layout.addLayout(self.size_filter_layout)
        
        self.preserve_structure_check = QCheckBox("🗂️ フォルダ構造を保持")
        advanced_layout.addWidget(self.preserve_structure_check)
        
        self.remove_empty_check = QCheckBox("🗑️ 空フォルダを削除")
        self.remove_empty_check.setChecked(True)
        advanced_layout.addWidget(self.remove_empty_check)
        
        sort_layout.addWidget(advanced_group)
        options_layout.addWidget(sort_group)
        
        layout.addWidget(options_group)
        
        # Execute buttons
        button_layout = QVBoxLayout()
        
        execute_btn = QPushButton("🚀 実行")
        execute_btn.setStyleSheet("background-color: #28a745; color: white; font-weight: bold; padding: 12px 24px; font-size: 14px;")
        execute_btn.clicked.connect(self.execute_processing)
        button_layout.addWidget(execute_btn)
        
        export_btn = QPushButton("📊 結果をCSVエクスポート")
        export_btn.setStyleSheet("background-color: #17a2b8; color: white; padding: 8px 16px;")
        export_btn.clicked.connect(self.export_results)
        button_layout.addWidget(export_btn)
        
        layout.addLayout(button_layout)
        layout.addStretch()
        
        return widget
    
    def apply_pro_theme(self):
        """Apply Pro (dark) theme with multimedia-specific styling"""
        pro_theme_file = Path("themes/pro.qss")
        if pro_theme_file.exists():
            with open(pro_theme_file, "r", encoding="utf-8") as f:
                base_style = f.read()
        else:
            base_style = self.get_fallback_theme()
            
        # Multimedia analyzer specific styles
        multimedia_style = """
            QTabWidget::pane {
                border: 2px solid #007acc;
                background-color: #2b2b2b;
                border-radius: 8px;
            }
            
            QTabBar::tab {
                background-color: #3c3c3c;
                color: #cccccc;
                padding: 10px 20px;
                margin-right: 2px;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
                font-weight: bold;
            }
            
            QTabBar::tab:selected {
                background-color: #007acc;
                color: #ffffff;
                font-weight: bold;
            }
            
            QTabBar::tab:hover:!selected {
                background-color: #4c4c4c;
            }
            
            QGroupBox {
                font-size: 13px;
                font-weight: bold;
                border: 2px solid #5c5c5c;
                border-radius: 8px;
                margin-top: 20px;
                padding-top: 10px;
            }
            
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 10px;
                color: #4ec9b0;
                font-size: 14px;
            }
        """
        
        self.setStyleSheet(base_style + multimedia_style)
    
    def get_fallback_theme(self) -> str:
        """Fallback theme for Pro style"""
        return """
            QMainWindow { 
                background-color: #1e1e1e; 
                color: #ffffff; 
            }
            QPushButton { 
                background-color: #3c3c3c; 
                color: white; 
                border: none; 
                padding: 8px 16px; 
                border-radius: 6px; 
                font-weight: bold;
            }
            QPushButton:hover { 
                background-color: #4c4c4c; 
            }
            QTreeWidget { 
                background-color: #2b2b2b; 
                color: #cccccc; 
                border: 1px solid #3c3c3c;
                border-radius: 4px;
            }
            QGroupBox { 
                color: #ffffff; 
                border: 1px solid #5c5c5c; 
                border-radius: 5px; 
                margin-top: 15px; 
                font-weight: bold;
            }
            QGroupBox::title {
                subcontrol-origin: margin; 
                left: 10px; 
                padding: 0 5px; 
                color: #4ec9b0;
            }
        """
    
    def select_folders(self):
        """Select folders for analysis"""
        folder = QFileDialog.getExistingDirectory(
            self, "解析対象フォルダを選択", 
            str(Path.home()),
            QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks
        )
        
        if folder:
            self.add_folder(Path(folder))
    
    def remove_selected_folders(self):
        """Remove selected folders from the list"""
        selected_items = self.folder_tree.selectedItems()
        if not selected_items:
            QMessageBox.information(self, "情報", "削除するフォルダを選択してください")
            return
            
        for item in selected_items:
            if item.parent() is None:  # Top level only
                path_str = item.data(0, Qt.UserRole)
                if path_str:
                    path_to_remove = Path(path_str)
                    if path_to_remove in self.selected_paths:
                        self.selected_paths.remove(path_to_remove)
                
                index = self.folder_tree.indexOfTopLevelItem(item)
                if index >= 0:
                    self.folder_tree.takeTopLevelItem(index)
        
        self.update_stats_display()
        
        # Add placeholder if empty
        self._add_placeholder_if_empty()

        self.status_bar.showMessage("選択したフォルダを削除しました")

    def remove_folders_by_name(self):
        """Remove folders by name (exact or partial match)."""
        dialog = FolderNameDeleteDialog(self)
        if dialog.exec() != QDialog.Accepted:
            return

        query = dialog.get_query()
        match_mode = dialog.get_match_mode()

        removed_paths = remove_folders_matching_query(
            self.folder_tree,
            self.selected_paths,
            query,
            match_mode=match_mode,
        )

        if not removed_paths:
            QMessageBox.information(self, "情報", f"『{query}』に該当するフォルダは見つかりませんでした。")
            return

        self.update_stats_display()
        self._add_placeholder_if_empty()

        match_label = "完全一致" if match_mode == MATCH_EXACT else "部分一致"
        preview_names = ", ".join(path.name for path in removed_paths[:3])
        if len(removed_paths) > 3:
            preview_names += " ..."

        message = (
            f"{len(removed_paths)}件のフォルダを削除 ({match_label}): {preview_names}"
            if preview_names else
            f"{len(removed_paths)}件のフォルダを削除 ({match_label})"
        )
        self.status_bar.showMessage(message)

    def run_multimedia_analysis(self):
        """Run comprehensive multimedia analysis"""
        if not self.selected_paths:
            QMessageBox.warning(self, "警告", "解析する対象フォルダがありません")
            return
        
        # Clear previous results
        self.suspect_files.clear()
        if hasattr(self, "suspect_review_btn"):
            self.suspect_review_btn.setEnabled(False)
        
        for tree in self.category_trees.values():
            tree.clear()
        
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)  # Indeterminate
        
        # Start analysis thread
        self.analysis_thread = MultimediaAnalysisThread(self.selected_paths)
        self.analysis_thread.progress_updated.connect(self.update_analysis_progress)
        self.analysis_thread.analysis_completed.connect(self.display_analysis_results)
        self.analysis_thread.error_occurred.connect(self.handle_analysis_error)
        self.analysis_thread.finished.connect(lambda: self.progress_bar.setVisible(False))
        self.analysis_thread.start()
    
    def update_analysis_progress(self, message: str, current: int, total: int):
        """Update analysis progress"""
        self.status_bar.showMessage(f"{message} ({current:,}/{total:,})")
        if total > 0:
            self.progress_bar.setRange(0, total)
            self.progress_bar.setValue(current)
    
    def display_analysis_results(self, results: Dict[str, Any]):
        """Display comprehensive analysis results"""
        suspect_files = results.get("__suspect_files__", [])
        category_results = {k: v for k, v in results.items() if k != "__suspect_files__"}
        
        self.analysis_results = category_results
        self.suspect_files = suspect_files
        
        if not category_results and not suspect_files:
            QMessageBox.information(self, "結果", "解析対象ファイルが見つかりませんでした")
            return
        
        # Category display names
        category_names = {
            "primary": {
                "primary_audio": "🎵 音声ファイル",
                "primary_video": "🎥 動画ファイル", 
                "primary_image": "🖼️ 画像ファイル",
                "primary_document": "📄 文書ファイル",
                "primary_code": "💻 コード・スクリプト",
                "primary_3d": "🎮 3Dモデル",
                "primary_archive": "📦 アーカイブ",
                "primary_font": "🔤 フォント",
                "primary_data": "💾 データベース",
                "primary_config": "⚙️ 設定ファイル",
                "primary_executable": "🚀 実行ファイル",
                "primary_virtual": "💿 仮想・ディスク",
                "primary_temp_log": "📋 ログ・一時",
                "primary_other": "📄 その他"
            },
            "size": {
                "size_tiny": "極小 (<0.1MB)",
                "size_very_small": "極小 (0.1-1MB)",
                "size_small": "小 (1-10MB)",
                "size_medium": "中 (10-100MB)", 
                "size_large": "大 (100MB-1GB)",
                "size_very_large": "特大 (1-5GB)",
                "size_huge": "巨大 (5GB+)",
                "size_unknown": "不明"
            },
            "age": {
                "age_today": "今日",
                "age_week": "1週間以内",
                "age_month": "1ヶ月以内",
                "age_year": "1年以内",
                "age_old": "1年以上前",
                "age_unknown": "不明"
            },
            "ext_length": {
                "ext_short": "短い拡張子 (≤3文字)",
                "ext_medium": "中程度 (4-6文字)",
                "ext_long": "長い拡張子 (7文字+)"
            },
            "engine": {
                "engine_audio": "🎵 音声エンジン",
                "engine_video": "🎥 動画エンジン",
                "engine_image": "🖼️ 画像エンジン", 
                "engine_document": "📄 文書エンジン",
                "engine_3d": "🎮 3Dエンジン",
                "engine_archive": "📦 アーカイブエンジン",
                "engine_font": "🔤 フォントエンジン",
                "engine_basic": "📄 基本エンジン"
            }
        }
        
        # Populate category trees
        for category_key, tree in self.category_trees.items():
            tree.clear()
            
            # Handle both direct categories and nested categories (e.g., audio_format)
            data_key = category_key
            if category_key not in category_results:
                # Try to find related category data
                possible_keys = [k for k in category_results.keys() if k.startswith(category_key.split('_')[0])]
                if possible_keys:
                    data_key = possible_keys[0]
                else:
                    continue
            
            if data_key not in category_results:
                continue
                
            category_data = category_results[data_key]
            names = category_names.get(category_key, {})
            
            for subcategory, data in category_data.items():
                # Create main item
                display_name = names.get(subcategory, subcategory.replace('_', ' ').title())
                item = QTreeWidgetItem(tree)
                item.setText(0, display_name)
                item.setText(1, f"{data['count']:,}")
                
                # Size
                total_size = data['total_size']
                size_mb = total_size / (1024 * 1024)
                if size_mb >= 1024:
                    size_gb = size_mb / 1024
                    if size_gb >= 1024:
                        size_tb = size_gb / 1024
                        item.setText(2, f"{size_tb:.1f} TB")
                    else:
                        item.setText(2, f"{size_gb:.1f} GB")
                else:
                    item.setText(2, f"{size_mb:.1f} MB" if size_mb >= 0.1 else "< 0.1 MB")
                
                # Average size
                if data['count'] > 0:
                    avg_size_mb = size_mb / data['count']
                    if avg_size_mb >= 1:
                        item.setText(3, f"{avg_size_mb:.1f} MB")
                    else:
                        avg_size_kb = avg_size_mb * 1024
                        item.setText(3, f"{avg_size_kb:.1f} KB" if avg_size_kb >= 0.1 else "< 0.1 KB")
                else:
                    item.setText(3, "N/A")
                
                # Store data for processing
                item.setData(0, Qt.UserRole, (data_key, subcategory))
        
        # Expand all trees and resize columns
        for tree in self.category_trees.values():
            tree.expandAll()
            for i in range(tree.columnCount()):
                tree.resizeColumnToContents(i)
        
        # Update status
        total_categories = sum(len(cat_data) for cat_data in category_results.values())
        total_files = sum(data['count'] for cat_data in category_results.values() for data in cat_data.values())
        suspect_summary = f"、異常候補 {len(self.suspect_files):,}件" if self.suspect_files else ""
        
        self.status_bar.showMessage(f"マルチメディア解析完了: {total_files:,}ファイル、{total_categories:,}カテゴリ{suspect_summary}")
        self.handle_suspect_files_post_analysis()
    
    def handle_suspect_files_post_analysis(self):
        """Handle suspect file detection results and prompt for export"""
        has_suspects = bool(self.suspect_files)
        
        if hasattr(self, "suspect_review_btn"):
            self.suspect_review_btn.setEnabled(has_suspects)
        
        if not has_suspects:
            return
        
        message = (
            f"{len(self.suspect_files):,}件のファイルで文字化けまたは破損の可能性を検出しました。\n"
            "一覧を確認してCSV/Excelへ出力しますか？"
        )
        reply = QMessageBox.question(
            self,
            "⚠️ 異常ファイルを検出しました",
            message,
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.show_suspect_review_dialog(prompt_export=True)
    
    def show_suspect_review_dialog(self, prompt_export: bool = False):
        """Show dialog listing files with potential anomalies"""
        if not self.suspect_files:
            QMessageBox.information(self, "情報", "異常が疑われるファイルは検出されていません。")
            return
        
        dialog = QDialog(self)
        dialog.setWindowTitle("⚠️ 文字化け／破損の可能性があるファイル")
        dialog.resize(920, 540)
        
        layout = QVBoxLayout(dialog)
        
        summary_text = (
            f"{len(self.suspect_files):,}件のファイルで異常が検出されました。"
            "必要に応じて下部のボタンからCSVまたはExcelへ出力できます。"
        )
        if prompt_export:
            summary_text += "\n\n（エクスポートを行う場合は下部ボタンを使用してください）"
        
        summary_label = QLabel(summary_text)
        summary_label.setWordWrap(True)
        layout.addWidget(summary_label)
        
        table = QTableWidget(len(self.suspect_files), 5)
        table.setHorizontalHeaderLabels(["ファイル名", "フォルダ", "カテゴリ", "判定理由", "サイズ"])
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setAlternatingRowColors(True)
        
        for row, info in enumerate(self.suspect_files):
            path_str = info.get("path", "")
            try:
                full_path = Path(path_str) if path_str else None
            except Exception:
                full_path = None
            
            name = info.get("name") or (full_path.name if full_path else "")
            category = info.get("category", "other")
            reasons = " / ".join(info.get("anomaly_reasons", [])) or "理由情報なし"
            size_value = info.get("size") or 0
            if size_value >= 1024 * 1024:
                size_str = f"{size_value / (1024 * 1024):.2f} MB"
            elif size_value >= 1024:
                size_str = f"{size_value / 1024:.2f} KB"
            else:
                size_str = f"{size_value} B"
            
            parent_text = str(full_path.parent) if full_path else ""
            
            table.setItem(row, 0, QTableWidgetItem(name))
            table.setItem(row, 1, QTableWidgetItem(parent_text))
            table.setItem(row, 2, QTableWidgetItem(category))
            reason_item = QTableWidgetItem(reasons)
            table.setItem(row, 3, reason_item)
            table.setItem(row, 4, QTableWidgetItem(size_str))
            
            if full_path:
                tooltip = str(full_path)
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    if item:
                        item.setToolTip(tooltip)
        
        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        
        layout.addWidget(table)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Close)
        export_csv_btn = QPushButton("CSV出力")
        export_excel_btn = QPushButton("Excel出力")
        button_box.addButton(export_csv_btn, QDialogButtonBox.ActionRole)
        button_box.addButton(export_excel_btn, QDialogButtonBox.ActionRole)
        layout.addWidget(button_box)
        
        export_csv_btn.clicked.connect(self.export_suspect_csv)
        export_excel_btn.clicked.connect(self.export_suspect_excel)
        button_box.rejected.connect(dialog.reject)
        
        dialog.exec()
    
    def export_suspect_csv(self):
        """Export suspect file list to CSV"""
        if not self.suspect_files:
            QMessageBox.information(self, "情報", "出力する異常ファイルはありません。")
            return
        
        default_name = Path.home() / f"suspect_files_{datetime.now():%Y%m%d_%H%M%S}.csv"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "文字化け／破損候補をCSV出力",
            str(default_name),
            "CSV files (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(["ファイル名", "フルパス", "カテゴリ", "判定理由", "サイズ(バイト)", "更新日時"])
                
                for info in self.suspect_files:
                    path_str = info.get("path", "")
                    reasons = " / ".join(info.get("anomaly_reasons", []))
                    size_value = info.get("size") or 0
                    mtime = info.get("mtime")
                    mtime_str = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S") if mtime else ""
                    writer.writerow([
                        info.get("name", ""),
                        path_str,
                        info.get("category", "other"),
                        reasons,
                        size_value,
                        mtime_str
                    ])
            
            QMessageBox.information(self, "出力完了", f"CSVとして保存しました:\n{file_path}")
        
        except Exception as exc:
            QMessageBox.critical(self, "出力エラー", f"CSV出力中にエラーが発生しました:\n{exc}")
    
    def export_suspect_excel(self):
        """Export suspect file list to Excel workbook"""
        if not self.suspect_files:
            QMessageBox.information(self, "情報", "出力する異常ファイルはありません。")
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.warning(
                self,
                "ライブラリ不足",
                "Excel出力には openpyxl が必要です。\n`pip install openpyxl` を実行してから再試行してください。"
            )
            return
        
        default_name = Path.home() / f"suspect_files_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "文字化け／破損候補をExcel出力",
            str(default_name),
            "Excel files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Suspicious Files"
            headers = ["ファイル名", "フルパス", "カテゴリ", "判定理由", "サイズ(バイト)", "更新日時"]
            sheet.append(headers)
            
            for info in self.suspect_files:
                path_str = info.get("path", "")
                reasons = "\n".join(info.get("anomaly_reasons", []))
                size_value = info.get("size") or 0
                mtime = info.get("mtime")
                mtime_str = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S") if mtime else ""
                
                sheet.append([
                    info.get("name", ""),
                    path_str,
                    info.get("category", "other"),
                    reasons,
                    size_value,
                    mtime_str
                ])
            
            for idx, header in enumerate(headers, start=1):
                column_letter = get_column_letter(idx)
                sheet.column_dimensions[column_letter].width = 40 if idx in (2, 4) else 20
            
            workbook.save(file_path)
            QMessageBox.information(self, "出力完了", f"Excelファイルとして保存しました:\n{file_path}")
        
        except Exception as exc:
            QMessageBox.critical(self, "出力エラー", f"Excel出力中にエラーが発生しました:\n{exc}")
    
    def handle_analysis_error(self, error_message: str):
        """Handle analysis errors"""
        self.progress_bar.setVisible(False)
        QMessageBox.critical(self, "解析エラー", f"マルチメディア解析中にエラーが発生しました:\n\n{error_message}")
        self.status_bar.showMessage("マルチメディア解析エラー")
    
    def execute_processing(self):
        """Execute multimedia processing based on settings"""
        if not self.analysis_results:
            QMessageBox.warning(self, "警告", "先にマルチメディア解析を実行してください")
            return
        
        # Get output directory
        output_dir = QFileDialog.getExistingDirectory(
            self, "出力先フォルダを選択", 
            str(Path.home()),
            QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks
        )
        
        if not output_dir:
            return
        
        # Get processing parameters
        mode = self.processing_mode.currentText()
        is_dry_run = self.dry_run_check.isChecked()
        
        # Collect selected files from all tabs
        selected_files = []
        for category_key, tree in self.category_trees.items():
            selected_items = tree.selectedItems()
            for item in selected_items:
                data_info = item.data(0, Qt.UserRole)
                if data_info:
                    data_key, subcategory = data_info
                    if data_key in self.analysis_results and subcategory in self.analysis_results[data_key]:
                        files = self.analysis_results[data_key][subcategory].get('files', [])
                        selected_files.extend(files)
        
        if not selected_files:
            QMessageBox.warning(self, "警告", "処理対象ファイルがありません\n各タブでファイルカテゴリを選択してください")
            return
        
        # Execute processing
        self._execute_multimedia_processing(selected_files, Path(output_dir), mode, is_dry_run)
    
    def _execute_multimedia_processing(self, files: List[Dict], output_dir: Path, mode: str, is_dry_run: bool):
        """Execute the actual multimedia processing"""
        success_count = 0
        error_count = 0
        skipped_count = 0
        
        # Apply filters
        filtered_files = []
        for file_info in files:
            # Size filter
            if self.size_filter_check.isChecked():
                try:
                    size_mb = file_info.get("size", 0) / (1024 * 1024)
                    min_size = float(self.size_min.text())
                    max_size = float(self.size_max.text())
                    
                    if size_mb < min_size or size_mb > max_size:
                        skipped_count += 1
                        continue
                except:
                    pass
            
            # Category filter
            file_category = file_info.get("category", "other")
            if file_category in self.category_checks:
                if not self.category_checks[file_category].isChecked():
                    skipped_count += 1
                    continue
            
            filtered_files.append(file_info)
        
        # Process filtered files
        duplicate_hashes = {}
        
        for file_info in filtered_files:
            try:
                source_path = Path(file_info['path'])
                if not source_path.exists():
                    error_count += 1
                    continue
                
                # Handle duplicates
                if self.duplicate_check.isChecked():
                    file_hash = file_info.get('file_hash')
                    if file_hash and file_hash in duplicate_hashes:
                        # Skip duplicate, but note it
                        skipped_count += 1
                        continue
                    elif file_hash:
                        duplicate_hashes[file_hash] = source_path
                
                # Determine target path based on mode
                if mode == "フラット化":
                    target_path = unique_name(output_dir, source_path.name)
                    
                elif mode == "統合整理":
                    # Create category-based subdirectory
                    category = file_info.get("category", "other")
                    category_names = {
                        "audio": "🎵音声", "video": "🎥動画", "image": "🖼️画像", 
                        "document": "📄文書", "3d": "🎮3D", "code": "💻コード",
                        "archive": "📦アーカイブ", "font": "🔤フォント", "data": "💾データ",
                        "config": "⚙️設定", "executable": "🚀実行", "virtual": "💿仮想",
                        "temp_log": "📋ログ", "other": "📄その他"
                    }
                    subdir_name = category_names.get(category, category)
                    subdir = output_dir / subdir_name
                    target_path = unique_name(subdir, source_path.name)
                    
                elif mode == "カテゴリ別整理":
                    # More detailed categorization
                    analysis_engine = file_info.get("analysis_engine", "basic")
                    category = file_info.get("category", "other")
                    
                    subdir = output_dir / f"{category}_{analysis_engine}"
                    target_path = unique_name(subdir, source_path.name)
                    
                else:
                    target_path = unique_name(output_dir, source_path.name)
                
                if not is_dry_run:
                    target_path.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(source_path, target_path)
                
                success_count += 1
                
            except Exception as e:
                error_count += 1
                continue
        
        # Show results
        mode_text = "シミュレーション" if is_dry_run else "実行"
        result_text = (f"{mode} {mode_text}が完了しました\n\n"
                      f"✅ 成功: {success_count:,}ファイル\n"
                      f"❌ エラー: {error_count:,}ファイル\n"
                      f"⏭️ スキップ: {skipped_count:,}ファイル")
        QMessageBox.information(self, "処理完了", result_text)
        
        self.status_bar.showMessage(f"処理完了: 成功{success_count:,}、エラー{error_count:,}、スキップ{skipped_count:,}")
    
    def export_results(self):
        """Export analysis results to CSV"""
        if not self.analysis_results:
            QMessageBox.warning(self, "警告", "先にマルチメディア解析を実行してください")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "CSVファイルとして保存",
            str(Path.home() / "multimedia_analysis_results.csv"),
            "CSV files (*.csv)"
        )
        
        if file_path:
            try:
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    
                    # Header
                    writer.writerow(["カテゴリ種別", "カテゴリ名", "ファイル数", "合計サイズ(MB)", "平均サイズ(MB)"])
                    
                    # Data
                    for category_type, category_data in self.analysis_results.items():
                        for subcategory, data in category_data.items():
                            writer.writerow([
                                category_type,
                                subcategory,
                                data['count'],
                                round(data['total_size'] / (1024 * 1024), 2),
                                round(data['total_size'] / (1024 * 1024) / data['count'], 2) if data['count'] > 0 else 0
                            ])
                
                QMessageBox.information(self, "エクスポート完了", f"結果をCSVファイルに保存しました:\n{file_path}")
                
            except Exception as e:
                QMessageBox.critical(self, "エクスポートエラー", f"CSVファイルの保存中にエラーが発生しました:\n{e}")
    
    def update_stats_display(self):
        """Update folder statistics display"""
        if self.selected_paths:
            folder_count = len(self.selected_paths)
            self.stats_label.setText(f"({folder_count}フォルダ選択中)")
        else:
            self.stats_label.setText("")
    
    def clear_all(self):
        """Clear all data"""
        reply = QMessageBox.question(self, "確認", "すべてをクリアしますか？")
        if reply == QMessageBox.Yes:
            self.selected_paths.clear()
            self.analysis_results.clear()
            self.suspect_files.clear()
            self.folder_tree.clear()
            for tree in self.category_trees.values():
                tree.clear()
            
            if hasattr(self, "suspect_review_btn"):
                self.suspect_review_btn.setEnabled(False)
            
            # Add placeholder
            self._add_placeholder_if_empty()
            
            self.update_stats_display()
            self.status_bar.showMessage("すべてクリアしました")
    
    # Drag and drop support
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
    
    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
    
    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            for url in urls:
                path = Path(url.toLocalFile())
                if path.is_dir():
                    self.add_folder(path)
            event.acceptProposedAction()
    
    def add_folder(self, folder_path: Path):
        """Add folder to the analysis list"""
        # Remove placeholder if present
        if self.folder_tree.topLevelItemCount() == 1:
            item = self.folder_tree.topLevelItem(0)
            if item.text(0) == self.folder_placeholder_text:
                self.folder_tree.clear()
        
        # Check if already exists
        if folder_path in self.selected_paths:
            return
        
        # Add to paths list
        self.selected_paths.append(folder_path)
        
        # Add to tree
        root_item = QTreeWidgetItem(self.folder_tree, [f"📁 {folder_path.name}"])
        root_item.setData(0, Qt.UserRole, str(folder_path))
        root_item.setToolTip(0, str(folder_path))
        
        # Count files by category (limited sample for performance)
        category_counts = {}
        file_count = 0
        
        try:
            for file_path in folder_path.rglob("*"):
                if file_path.is_file():
                    file_count += 1
                    if file_count <= 1000:  # Limit for performance
                        category = detect_file_category(file_path)
                        category_counts[category] = category_counts.get(category, 0) + 1
                    elif file_count > 1000:
                        break  # Stop counting after 1000 files
            
            # Add category summaries as children
            category_icons = {
                'audio': '🎵', 'video': '🎥', 'image': '🖼️', 'document': '📄',
                '3d': '🎮', 'code': '💻', 'archive': '📦', 'font': '🔤',
                'data': '💾', 'config': '⚙️', 'executable': '🚀', 'virtual': '💿',
                'temp_log': '📋', 'other': '📄'
            }
            
            for category, count in sorted(category_counts.items()):
                if count > 0:
                    icon = category_icons.get(category, '📄')
                    child_item = QTreeWidgetItem(root_item)
                    child_item.setText(0, f"{icon} {category}: {count:,}")
                    child_item.setFlags(Qt.NoItemFlags)
            
            if file_count > 1000:
                more_item = QTreeWidgetItem(root_item)
                more_item.setText(0, f"... 総計{file_count:,}+ファイル (サンプル表示)")
                more_item.setFlags(Qt.NoItemFlags)
                more_item.setForeground(0, QBrush(QColor("#888888")))
        
        except Exception:
            pass
        
        root_item.setExpanded(True)
        self.update_stats_display()
        self.status_bar.showMessage(f"フォルダを追加: {folder_path.name} ({file_count:,}+ファイル)")


if __name__ == "__main__":
    from PySide6.QtWidgets import QApplication
    import sys
    
    app = QApplication(sys.argv)
    window = MultimediaAnalyzerWindow()
    window.show()
    sys.exit(app.exec())
